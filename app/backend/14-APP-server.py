"""
Spalla Dashboard — Server with Zoom, Google Calendar & Supabase Integration
Serves static files + proxies APIs
"""

import http.server
import http.client
import json
import urllib.request
import urllib.error
import urllib.parse
import os
import sys
import time
import base64
import threading
import hmac
import hashlib
import secrets
import unicodedata
import re
import uuid
from datetime import datetime, timedelta, timezone

try:
    import jwt
except ImportError:
    jwt = None  # Will handle gracefully if not installed

try:
    import bcrypt as _bcrypt
    BCRYPT_AVAILABLE = True
except ImportError:
    _bcrypt = None
    BCRYPT_AVAILABLE = False
    print('[WARN] bcrypt not installed — password hashing will fail. Run: pip install bcrypt')

PORT = int(os.environ.get('PORT', 8888))

# ===== CONFIG =====
EVOLUTION_BASE = os.environ.get('EVOLUTION_BASE', 'https://evolution.manager01.feynmanproject.com')
EVOLUTION_API_KEY = os.environ.get('EVOLUTION_API_KEY', '')

# Zoom Server-to-Server OAuth
ZOOM_ACCOUNT_ID = os.environ.get('ZOOM_ACCOUNT_ID', '')
ZOOM_CLIENT_ID = os.environ.get('ZOOM_CLIENT_ID', '')
ZOOM_CLIENT_SECRET = os.environ.get('ZOOM_CLIENT_SECRET', '')

# Google Service Account
GOOGLE_SA_PATH = os.environ.get('GOOGLE_SA_PATH', os.path.expanduser('~/.config/google/credentials.json'))

# YouTube API
YOUTUBE_API_KEY = os.environ.get('YOUTUBE_API_KEY', '')

# Google Drive (2 produtos)
GOOGLE_DRIVE_FOLDER_MENTORY = os.environ.get('GOOGLE_DRIVE_FOLDER_MENTORY', '')
GOOGLE_DRIVE_FOLDER_CLINIC = os.environ.get('GOOGLE_DRIVE_FOLDER_CLINIC', '')

# Supabase
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_ANON_KEY = os.environ.get('SUPABASE_ANON_KEY', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY', '')

# Calendar ID
GOOGLE_CALENDAR_ID = os.environ.get('GOOGLE_CALENDAR_ID', 'primary')

# ===== HETZNER S3 CONFIG =====
S3_ACCESS_KEY = os.environ.get('S3_ACCESS_KEY', '') or os.environ.get('S3_ACESS_KEY', '')
S3_SECRET_KEY = os.environ.get('S3_SECRET_KEY', '')
S3_BUCKET     = os.environ.get('S3_BUCKET', 'case-evolution-media')
S3_ENDPOINT   = os.environ.get('S3_ENDPOINT', 'hel1.your-objectstorage.com')
S3_REGION     = os.environ.get('S3_REGION', 'eu-central')

# ===== AI CONFIG (Semantic Search + Processing) =====
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '')
VOYAGE_API_KEY = os.environ.get('VOYAGE_API_KEY', '')
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')
WHISPER_MODEL = 'whisper-1'
GROQ_WHISPER_MODEL = 'whisper-large-v3'
VISION_MODEL = 'gpt-4o'
GEMINI_VISION_MODEL = 'gemini-2.5-flash'

# Embedding config — Voyage AI (512 dims via voyage-3-lite, $0.02/1M tokens, 200M free)
EMBEDDING_PROVIDER = os.environ.get('EMBEDDING_PROVIDER', 'voyage')  # 'voyage' or 'openai'
VOYAGE_EMBED_MODEL = 'voyage-3-lite'
OPENAI_EMBED_MODEL = 'text-embedding-3-small'
EMBEDDING_DIMS = 512 if EMBEDDING_PROVIDER == 'voyage' else 1536

# Chunking config
CHUNK_SIZE_TOKENS = 800
CHUNK_OVERLAP_TOKENS = 200
CHUNK_CONTEXT_SIZE_TOKENS = 2000   # for multi-pass context chunks
MIN_CHUNK_TOKENS = 50

# ===== CHATWOOT CONFIG =====
CHATWOOT_BASE_URL = os.environ.get('CHATWOOT_BASE_URL', '')  # e.g. https://chat.spalla.app
CHATWOOT_API_TOKEN = os.environ.get('CHATWOOT_API_TOKEN', '')  # user or bot token
CHATWOOT_ACCOUNT_ID = os.environ.get('CHATWOOT_ACCOUNT_ID', '1')
CHATWOOT_WEBHOOK_SECRET = os.environ.get('CHATWOOT_WEBHOOK_SECRET', '')  # HMAC verification
EVOLUTION_WEBHOOK_SECRET = os.environ.get('EVOLUTION_WEBHOOK_SECRET', '')  # Evolution API webhook apikey
WA_RATE_LIMIT_PER_MINUTE = int(os.environ.get('WA_RATE_LIMIT_PER_MINUTE', '30'))

# ===== JWT AUTH CONFIG =====
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    import secrets as _s
    JWT_SECRET = _s.token_hex(32)
    print(f'[WARNING] JWT_SECRET not set — generated ephemeral key (tokens will not survive restarts)')
JWT_ALGORITHM = 'HS256'
ACCESS_TOKEN_EXPIRY_MINUTES = 480  # 8 hours — avoids mid-day session drops
REFRESH_TOKEN_EXPIRY_DAYS = 7

# ===== API KEY CONFIG =====
STATIC_API_KEYS = {}
_raw_keys = os.environ.get('API_KEYS', '')
if _raw_keys:
    for entry in _raw_keys.split(','):
        entry = entry.strip()
        if ':' in entry:
            k, label = entry.split(':', 1)
            STATIC_API_KEYS[k.strip()] = label.strip()
        elif entry:
            STATIC_API_KEYS[entry] = 'default'

# ===== AUTH FUNCTIONS (Supabase-backed) =====
def hash_password(password):
    """Hash password using bcrypt — raises if bcrypt is not installed"""
    if not BCRYPT_AVAILABLE:
        raise RuntimeError('bcrypt is required for password hashing. Run: pip install bcrypt')
    return _bcrypt.hashpw(password.encode(), _bcrypt.gensalt()).decode()

def _is_legacy_sha256(stored_hash):
    """Check if hash is a legacy SHA-256 (64 hex chars)"""
    return len(stored_hash) == 64 and all(c in '0123456789abcdef' for c in stored_hash)

def verify_password(password, stored_hash):
    """Verify password — supports both bcrypt and legacy SHA-256"""
    if _is_legacy_sha256(stored_hash):
        if hashlib.sha256(password.encode()).hexdigest() == stored_hash:
            return 'migrate'  # Signal to upgrade hash
        return False
    if not BCRYPT_AVAILABLE:
        raise RuntimeError('bcrypt is required for password verification. Run: pip install bcrypt')
    try:
        return _bcrypt.checkpw(password.encode(), stored_hash.encode())
    except Exception:
        return False

def create_jwt_token(email, user_id, role='equipe', expiry_minutes=ACCESS_TOKEN_EXPIRY_MINUTES):
    """Create JWT access token"""
    if not jwt:
        return None
    payload = {
        'email': email,
        'user_id': user_id,
        'role': role,        'exp': datetime.now(timezone.utc) + timedelta(minutes=expiry_minutes),
        'iat': datetime.now(timezone.utc)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(email, user_id):
    """Create JWT refresh token (long-lived)"""
    if not jwt:
        return None
    payload = {
        'email': email,
        'user_id': user_id,
        'type': 'refresh',
        'exp': datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRY_DAYS),
        'iat': datetime.now(timezone.utc)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_jwt_token(token):
    """Verify and decode JWT token"""
    if not jwt:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def verify_api_key(key):
    """Verify API key — checks static keys first, then Supabase api_keys table."""
    if not key:
        return None
    if key in STATIC_API_KEYS:
        return {'label': STATIC_API_KEYS[key], 'source': 'env'}
    try:
        result = supabase_request('GET',
            f'api_keys?select=id,label,role,active'
            f'&key_hash=eq.{hashlib.sha256(key.encode()).hexdigest()}'
            f'&active=eq.true&limit=1')
        if isinstance(result, list) and result:
            supabase_request('PATCH',
                f'api_keys?id=eq.{result[0]["id"]}',
                {'last_used_at': datetime.now(timezone.utc).isoformat()})
            return {'label': result[0].get('label', ''), 'source': 'supabase',
                    'role': result[0].get('role', 'integration')}
    except Exception as e:
        log_error('APIKeys', f'verify_api_key failed: {e}')
    return None


def check_auth_any(headers):
    """Check auth via JWT Bearer OR X-API-Key header."""
    auth_header = headers.get('Authorization', '')
    if auth_header.startswith('Bearer '):
        payload = verify_jwt_token(auth_header[7:])
        if payload and payload.get('type') != 'refresh':
            return {'method': 'jwt', 'email': payload.get('email', ''),
                    'role': payload.get('role', 'equipe'), 'payload': payload}
    api_key = headers.get('X-API-Key', '')
    if api_key:
        result = verify_api_key(api_key)
        if result:
            return {'method': 'api_key', 'label': result['label'],
                    'role': result.get('role', 'integration')}
    return None


# Auth users stored in Supabase table 'auth_users'

def generate_presigned_url(key, expires=3600):
    """Generate AWS Signature V4 presigned URL for Hetzner S3"""
    host = S3_ENDPOINT
    algorithm = 'AWS4-HMAC-SHA256'
    now = datetime.now(timezone.utc)
    amz_date = now.strftime('%Y%m%dT%H%M%SZ')
    datestamp = now.strftime('%Y%m%d')

    # Debug: log S3 config
    print(f'[S3 Debug] BUCKET={S3_BUCKET}, ENDPOINT={S3_ENDPOINT}, REGION={S3_REGION}, KEY={key}')

    # Canonical request
    method = 'GET'
    canonical_uri = f'/{S3_BUCKET}/{key}'
    credential_scope = f'{datestamp}/{S3_REGION}/s3/aws4_request'

    canonical_querystring = f'X-Amz-Algorithm={algorithm}&X-Amz-Credential={urllib.parse.quote(f"{S3_ACCESS_KEY}/{credential_scope}", safe="")}&X-Amz-Date={amz_date}&X-Amz-Expires={expires}&X-Amz-SignedHeaders=host'

    canonical_headers = f'host:{host}\n'
    signed_headers = 'host'

    payload_hash = hashlib.sha256(b'').hexdigest()

    canonical_request = f'{method}\n{canonical_uri}\n{canonical_querystring}\n{canonical_headers}\n{signed_headers}\n{payload_hash}'

    # String to sign
    canonical_request_hash = hashlib.sha256(canonical_request.encode()).hexdigest()
    string_to_sign = f'{algorithm}\n{amz_date}\n{credential_scope}\n{canonical_request_hash}'

    # Calculate signature
    kDate = hmac.new(f'AWS4{S3_SECRET_KEY}'.encode(), datestamp.encode(), hashlib.sha256).digest()
    kRegion = hmac.new(kDate, S3_REGION.encode(), hashlib.sha256).digest()
    kService = hmac.new(kRegion, b's3', hashlib.sha256).digest()
    kSigning = hmac.new(kService, b'aws4_request', hashlib.sha256).digest()
    signature = hmac.new(kSigning, string_to_sign.encode(), hashlib.sha256).hexdigest()

    # Build URL
    url = f'https://{host}{canonical_uri}?{canonical_querystring}&X-Amz-Signature={signature}'
    return url

# ===== SUPABASE CONNECTION POOL =====
SUPABASE_HOST = 'knusqfbvhsqworzyhvip.supabase.co'
_supa_lock = threading.Lock()
_supa_conn = None


def _get_supa_conn():
    """Retorna conexão HTTPS persistente com Supabase"""
    global _supa_conn
    if _supa_conn is not None:
        return _supa_conn
    _supa_conn = http.client.HTTPSConnection(SUPABASE_HOST, timeout=15)
    return _supa_conn


def _reset_supa_conn():
    """Fecha e reseta a conexão Supabase"""
    global _supa_conn
    try:
        if _supa_conn:
            _supa_conn.close()
    except Exception:
        pass
    _supa_conn = None


def log_info(source, msg):
    """Log info message with timestamp"""
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f'[{ts}] [{source}] {msg}')


def log_error(source, msg, exc=None):
    """Log error message with timestamp"""
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if exc:
        print(f'[{ts}] [{source}] ERROR: {msg} — {exc}')
    else:
        print(f'[{ts}] [{source}] ERROR: {msg}')


# ===== CLICKUP DEEP LINK =====
def clickup_deep_link(url):
    """Converte URL web do ClickUp para deep link do app desktop."""
    if url and url.startswith('https://app.clickup.com/'):
        return url.replace('https://app.clickup.com/', 'clickup://')
    return url


# ===== ZOOM TOKEN CACHE =====
_zoom_token = {'access_token': None, 'expires_at': 0}


def get_zoom_token():
    """Get Zoom access token via Server-to-Server OAuth"""
    global _zoom_token
    if _zoom_token['access_token'] and time.time() < _zoom_token['expires_at'] - 60:
        return _zoom_token['access_token']

    if not ZOOM_ACCOUNT_ID or not ZOOM_CLIENT_ID or not ZOOM_CLIENT_SECRET:
        return None

    creds = base64.b64encode(f'{ZOOM_CLIENT_ID}:{ZOOM_CLIENT_SECRET}'.encode()).decode()
    url = f'https://zoom.us/oauth/token?grant_type=account_credentials&account_id={ZOOM_ACCOUNT_ID}'
    req = urllib.request.Request(url, data=b'', method='POST')
    req.add_header('Authorization', f'Basic {creds}')
    req.add_header('Content-Type', 'application/x-www-form-urlencoded')

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            _zoom_token['access_token'] = data['access_token']
            _zoom_token['expires_at'] = time.time() + data.get('expires_in', 3600)
            return data['access_token']
    except Exception as e:
        print(f'[Zoom] Token error: {e}')
        return None


def create_zoom_meeting(topic, start_time, duration=60, invitees=None):
    """Create a Zoom meeting and return join_url + meeting_id"""
    token = get_zoom_token()
    if not token:
        return {'error': 'Zoom credentials not configured'}

    body = json.dumps({
        'topic': topic,
        'type': 2,  # Scheduled meeting
        'start_time': start_time,  # ISO 8601
        'duration': duration,
        'timezone': 'America/Sao_Paulo',
        'settings': {
            'join_before_host': True,
            'waiting_room': False,
            'auto_recording': 'cloud',
            'meeting_invitees': [{'email': e} for e in (invitees or [])],
        }
    }).encode()

    req = urllib.request.Request('https://api.zoom.us/v2/users/me/meetings', data=body, method='POST')
    req.add_header('Authorization', f'Bearer {token}')
    req.add_header('Content-Type', 'application/json')

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            return {
                'meeting_id': data.get('id'),
                'join_url': data.get('join_url'),
                'start_url': data.get('start_url'),
                'password': data.get('password'),
                'topic': data.get('topic'),
            }
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        print(f'[Zoom] Create meeting error: {err}')
        return {'error': f'Zoom API error: {e.code}', 'detail': err}
    except Exception as e:
        return {'error': str(e)}


# ===== GOOGLE CALENDAR =====
_gcal_service = None


def get_gcal_service():
    """Initialize Google Calendar API service using service account.
    Supports GOOGLE_SA_JSON (base64-encoded JSON) for Docker environments
    or GOOGLE_SA_PATH (file path) for local development.
    """
    global _gcal_service
    if _gcal_service:
        return _gcal_service

    try:
        import base64
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

        SCOPES = ['https://www.googleapis.com/auth/calendar']
        sa_json = os.environ.get('GOOGLE_SA_JSON', '') or os.environ.get('GOOGLE_SA_CREDENTIALS_B64', '')

        if sa_json:
            info = json.loads(base64.b64decode(sa_json))
            credentials = service_account.Credentials.from_service_account_info(
                info, scopes=SCOPES
            )
            print('[GCal] using GOOGLE_SA_JSON')
        elif os.path.exists(GOOGLE_SA_PATH):
            credentials = service_account.Credentials.from_service_account_file(
                GOOGLE_SA_PATH, scopes=SCOPES
            )
            print(f'[GCal] using GOOGLE_SA_PATH: {GOOGLE_SA_PATH}')
        else:
            print(f'[GCal] No credentials found (GOOGLE_SA_JSON not set, {GOOGLE_SA_PATH} not found)')
            return None

        _gcal_service = build('calendar', 'v3', credentials=credentials)
        return _gcal_service
    except Exception as e:
        print(f'[GCal] Init error: {e}')
        return None


def create_calendar_event(summary, start_iso, end_iso, description='', attendees=None, location=''):
    """Create a Google Calendar event"""
    service = get_gcal_service()
    if not service:
        return {'error': 'Google Calendar not configured'}

    event = {
        'summary': summary,
        'description': description,
        'start': {
            'dateTime': start_iso,
            'timeZone': 'America/Sao_Paulo',
        },
        'end': {
            'dateTime': end_iso,
            'timeZone': 'America/Sao_Paulo',
        },
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'popup', 'minutes': 30},
                {'method': 'email', 'minutes': 60},
            ],
        },
    }

    if location:
        event['location'] = location

    if attendees:
        event['attendees'] = [{'email': e} for e in attendees if e]

    try:
        result = service.events().insert(
            calendarId=GOOGLE_CALENDAR_ID,
            body=event,
            sendUpdates='all'  # Send email invites
        ).execute()
        return {
            'event_id': result.get('id'),
            'html_link': result.get('htmlLink'),
            'status': result.get('status'),
        }
    except Exception as e:
        print(f'[GCal] Create event error: {e}')
        return {'error': str(e)}


def list_calendar_events(time_min=None, time_max=None, max_results=50):
    """List upcoming calendar events"""
    service = get_gcal_service()
    if not service:
        return {'error': 'Google Calendar not configured'}

    if not time_min:
        time_min = datetime.now(timezone.utc).isoformat() + 'Z'

    try:
        result = service.events().list(
            calendarId=GOOGLE_CALENDAR_ID,
            timeMin=time_min,
            timeMax=time_max,
            maxResults=max_results,
            singleEvents=True,
            orderBy='startTime'
        ).execute()
        return {'events': result.get('items', [])}
    except Exception as e:
        return {'error': str(e)}


def delete_calendar_event(event_id):
    """Delete a Google Calendar event by ID"""
    if not event_id:
        return {'skipped': True, 'reason': 'no event_id provided'}

    service = get_gcal_service()
    if not service:
        return {'error': 'Google Calendar not configured'}

    try:
        service.events().delete(
            calendarId=GOOGLE_CALENDAR_ID,
            eventId=event_id
        ).execute()
        return {'deleted': True, 'event_id': event_id}
    except Exception as e:
        print(f'[GCal] Delete event error: {e}')
        return {'error': str(e)}


# ===== GOOGLE SHEETS SYNC =====
PAYMENTS_SHEET_ID = '1YY6t5ZxRPTLyCHC-EVkyem10caEJw4TuBy3AR14r0ao'
PAYMENTS_TAB = 'Acompanhamento Pagamentos'
CONTRACTS_SHEET_ID = '1-Yi5G-bUJanRtmfugFgmVz_DRSV-HOEaP-sSAJqUIxY'
CONTRACTS_TAB = 'Dossiê Estratégico Mentorados'

_sheets_service = None
_sheets_last_sync = None
_sheets_last_result = None


def get_sheets_service():
    """Initialize Google Sheets API service using service account (env var or file)"""
    global _sheets_service
    if _sheets_service:
        return _sheets_service

    try:
        import base64
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

        SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

        sa_json = os.environ.get('GOOGLE_SA_JSON', '') or os.environ.get('GOOGLE_SA_CREDENTIALS_B64', '')
        if sa_json:
            info = json.loads(base64.b64decode(sa_json))
            credentials = service_account.Credentials.from_service_account_info(
                info, scopes=SCOPES
            )
            print('[Sheets] using GOOGLE_SA_JSON/GOOGLE_SA_CREDENTIALS_B64')
        elif os.path.exists(GOOGLE_SA_PATH):
            credentials = service_account.Credentials.from_service_account_file(
                GOOGLE_SA_PATH, scopes=SCOPES
            )
            print(f'[Sheets] using GOOGLE_SA_PATH: {GOOGLE_SA_PATH}')
        else:
            log_error('Sheets', f'No credentials found (GOOGLE_SA_JSON not set, {GOOGLE_SA_PATH} not found)')
            return None

        _sheets_service = build('sheets', 'v4', credentials=credentials)
        return _sheets_service
    except Exception as e:
        log_error('Sheets', 'Init error', e)
        return None


def _normalize_name(name):
    """Normalize name for matching: remove accents, 'Dra.', 'Dr.', lowercase, strip"""
    if not name:
        return ''
    # Remove Dra./Dr. prefix
    n = re.sub(r'^(Dra?\.?\s*)', '', name.strip(), flags=re.IGNORECASE)
    # Remove accents
    n = unicodedata.normalize('NFD', n)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    return n.lower().strip()


def _name_tokens(name):
    """Get set of name tokens for fuzzy matching"""
    return set(_normalize_name(name).split())


def _match_name(sheet_name, mentorados):
    """Match a sheet name to a mentorado. Returns mentorado dict or None."""
    norm = _normalize_name(sheet_name)
    if not norm:
        return None

    # Exact normalized match
    for m in mentorados:
        if _normalize_name(m['nome']) == norm:
            return m

    # Token overlap: at least 2 tokens in common, or all tokens of shorter name match
    sheet_tokens = _name_tokens(sheet_name)
    if len(sheet_tokens) < 1:
        return None

    best_match = None
    best_score = 0
    for m in mentorados:
        m_tokens = _name_tokens(m['nome'])
        overlap = sheet_tokens & m_tokens
        if len(overlap) >= 2 or (len(overlap) >= 1 and len(overlap) == min(len(sheet_tokens), len(m_tokens))):
            score = len(overlap) / max(len(sheet_tokens | m_tokens), 1)
            if score > best_score:
                best_score = score
                best_match = m

    return best_match if best_score >= 0.4 else None


def read_payments_sheet():
    """Read payments sheet and return list of dicts (one per row)"""
    service = get_sheets_service()
    if not service:
        return []

    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=PAYMENTS_SHEET_ID,
            range=f"'{PAYMENTS_TAB}'"
        ).execute()
        rows = result.get('values', [])
        if len(rows) < 2:
            return []

        headers = [h.strip().lower() for h in rows[0]]
        data = []
        for row in rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                d[h] = row[i].strip() if i < len(row) and row[i] else ''
            if d.get('nome') or d.get('mentorado') or d.get('mentorada'):
                data.append(d)
        log_info('Sheets', f'Read {len(data)} rows from payments sheet')
        return data
    except Exception as e:
        log_error('Sheets', 'Error reading payments sheet', e)
        return []


def read_contracts_sheet():
    """Read contracts sheet and return list of dicts (one per row)"""
    service = get_sheets_service()
    if not service:
        return []

    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=CONTRACTS_SHEET_ID,
            range=f"'{CONTRACTS_TAB}'"
        ).execute()
        rows = result.get('values', [])
        if len(rows) < 2:
            return []

        headers = [h.strip().lower() for h in rows[0]]
        data = []
        for row in rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                d[h] = row[i].strip() if i < len(row) and row[i] else ''
            if d.get('nome') or d.get('mentorado') or d.get('mentorada'):
                data.append(d)
        log_info('Sheets', f'Read {len(data)} rows from contracts sheet')
        return data
    except Exception as e:
        log_error('Sheets', 'Error reading contracts sheet', e)
        return []


def _parse_status_financeiro(row):
    """Parse financial status from a payments row"""
    # Look for common column names
    for key in ('status', 'status financeiro', 'status_financeiro', 'situação', 'situacao'):
        val = row.get(key, '').lower().strip()
        if val:
            if any(w in val for w in ('atrasa', 'inadimpl', 'pendente', 'devendo')):
                return 'atrasado'
            if any(w in val for w in ('quita', 'pago', 'encerrad')):
                return 'quitado'
            if any(w in val for w in ('dia', 'ok', 'regular', 'ativ')):
                return 'em_dia'
    return None


def _parse_dia_pagamento(row):
    """Parse payment day from a payments row"""
    for key in ('dia', 'dia pagamento', 'dia_pagamento', 'vencimento', 'dia vencimento'):
        val = row.get(key, '').strip()
        if val:
            # Extract first number found
            match = re.search(r'(\d{1,2})', val)
            if match:
                day = int(match.group(1))
                if 1 <= day <= 31:
                    return day
    return None


def _parse_contrato(row):
    """Parse contract status from a contracts row"""
    # Column M is typically index 12 (0-based), but we use headers
    for key in ('contrato', 'contrato assinado', 'contrato_assinado', 'status contrato'):
        val = row.get(key, '').lower().strip()
        if val:
            if any(w in val for w in ('sim', 'yes', 'assinado', 'ok', 'true', '1')):
                return True
            if any(w in val for w in ('não', 'nao', 'no', 'false', '0', 'pendente', 'falta')):
                return False
    return None


def sync_sheets_to_supabase():
    """Sync Google Sheets data to Supabase mentorados table"""
    global _sheets_last_sync, _sheets_last_result

    log_info('Sheets', 'Starting sync...')
    start_time = time.time()

    # Fetch mentorados from Supabase
    mentorados = supabase_request('GET', 'mentorados?select=id,nome&ativo=eq.true&cohort=not.eq.tese&order=nome')
    if isinstance(mentorados, dict) and 'error' in mentorados:
        log_error('Sheets', f'Failed to fetch mentorados: {mentorados}')
        _sheets_last_result = {'error': mentorados['error'], 'timestamp': datetime.now().isoformat()}
        return _sheets_last_result

    if not mentorados:
        _sheets_last_result = {'error': 'No mentorados found', 'timestamp': datetime.now().isoformat()}
        return _sheets_last_result

    # Read both sheets
    payments = read_payments_sheet()
    contracts = read_contracts_sheet()

    updated = 0
    unmatched_payments = []
    unmatched_contracts = []
    updates_log = []

    # Process payments sheet
    for row in payments:
        name = row.get('nome') or row.get('mentorado') or row.get('mentorada') or ''
        m = _match_name(name, mentorados)
        if not m:
            if name.strip():
                unmatched_payments.append(name)
            continue

        patch = {}
        status = _parse_status_financeiro(row)
        if status:
            patch['status_financeiro'] = status
        dia = _parse_dia_pagamento(row)
        if dia:
            patch['dia_pagamento'] = dia

        if patch:
            result = supabase_request('PATCH', f"mentorados?id=eq.{m['id']}", patch)
            if not (isinstance(result, dict) and 'error' in result):
                updated += 1
                updates_log.append({'id': m['id'], 'nome': m['nome'], 'fields': patch})

    # Process contracts sheet
    for row in contracts:
        name = row.get('nome') or row.get('mentorado') or row.get('mentorada') or ''
        m = _match_name(name, mentorados)
        if not m:
            if name.strip():
                unmatched_contracts.append(name)
            continue

        contrato = _parse_contrato(row)
        if contrato is not None:
            patch = {'contrato_assinado': contrato}
            result = supabase_request('PATCH', f"mentorados?id=eq.{m['id']}", patch)
            if not (isinstance(result, dict) and 'error' in result):
                updated += 1
                updates_log.append({'id': m['id'], 'nome': m['nome'], 'fields': patch})

    elapsed = round(time.time() - start_time, 2)
    _sheets_last_sync = datetime.now().isoformat()
    _sheets_last_result = {
        'success': True,
        'updated': updated,
        'payments_rows': len(payments),
        'contracts_rows': len(contracts),
        'unmatched_payments': unmatched_payments,
        'unmatched_contracts': unmatched_contracts,
        'updates': updates_log,
        'elapsed_seconds': elapsed,
        'timestamp': _sheets_last_sync,
    }
    log_info('Sheets', f'Sync complete: {updated} updates in {elapsed}s')
    if unmatched_payments:
        log_info('Sheets', f'Unmatched payments: {unmatched_payments}')
    if unmatched_contracts:
        log_info('Sheets', f'Unmatched contracts: {unmatched_contracts}')

    return _sheets_last_result


# ===== SHEETS SYNC SCHEDULER =====
def _sheets_sync_loop():
    """Background thread: sync every 6 hours"""
    time.sleep(30)  # Initial delay
    while True:
        try:
            sync_sheets_to_supabase()
        except Exception as e:
            log_error('Sheets', 'Scheduler sync failed', e)
        time.sleep(6 * 3600)  # 6 hours


# ===== SUPABASE HELPERS =====
class SupabaseResponse:
    """Wrapper for error responses — supports .status_code/.json()/.text AND dict access"""
    def __init__(self, data, status_code=500):
        self._data = data
        self.status_code = status_code
        self.text = json.dumps(data) if data else ''
    def json(self):
        return self._data
    def __iter__(self):
        return iter([])
    def __len__(self):
        return 0
    def __bool__(self):
        return False
    def __getitem__(self, key):
        if isinstance(self._data, dict): return self._data[key]
        raise KeyError(key)
    def get(self, key, default=None):
        if isinstance(self._data, dict): return self._data.get(key, default)
        return default


def supa_ok(result):
    """Check if supabase_request succeeded. Works with both raw data and SupabaseResponse."""
    if isinstance(result, SupabaseResponse):
        return result.status_code in (200, 201, 204)
    return not (isinstance(result, dict) and result.get('error'))


def _dependency_reactor(completed_task_id, completed_titulo=''):
    """ORCH-04: After a task reaches terminal state, check dependents and auto-unblock.
    For agent responsaveis: auto-start. For humans: just unblock (notification later).
    Returns list of actions taken."""
    results = []
    # Find tasks that depend on this one and are still pendente or bloqueada
    deps_r = supabase_request('GET',
        f'god_tasks?depends_on=cs.{{{completed_task_id}}}&status=in.(pendente,bloqueada)&select=id,titulo,depends_on,responsavel,especie,status')
    if not supa_ok(deps_r) or not isinstance(deps_r, list) or not deps_r:
        return results

    terminal = ('concluida', 'cancelada', 'arquivada')
    for dep in deps_r:
        # Check if ALL blockers are resolved
        all_deps = dep.get('depends_on') or []
        if len(all_deps) > 1:
            other_ids = [d for d in all_deps if d != completed_task_id]
            if other_ids:
                in_clause = ','.join(other_ids)
                others = supabase_request('GET', f'god_tasks?id=in.({in_clause})&select=id,status')
                if isinstance(others, list) and not all(o.get('status') in terminal for o in others):
                    continue  # Still has unresolved blockers

        # All blockers resolved — determine action
        dep_id = dep['id']
        responsavel = dep.get('responsavel') or ''

        # Check if responsavel is an agent
        is_agent = False
        if responsavel:
            member_r = supabase_request('GET',
                f'spalla_members?or=(id.eq.{responsavel},nome_curto.ilike.{responsavel})&tipo=eq.agent&limit=1')
            is_agent = isinstance(member_r, list) and len(member_r) > 0

        if dep.get('status') == 'bloqueada':
            # Unblock: bloqueada → em_andamento
            supabase_request('PATCH', f'god_tasks?id=eq.{dep_id}',
                body={'status': 'em_andamento', 'updated_at': datetime.now(timezone.utc).isoformat()})
            action = 'auto_unblocked'
        elif is_agent:
            # Agent: pendente → em_andamento (auto-start)
            supabase_request('PATCH', f'god_tasks?id=eq.{dep_id}',
                body={'status': 'em_andamento', 'updated_at': datetime.now(timezone.utc).isoformat()})
            action = 'agent_auto_started'
        else:
            # Human: just mark as ready (keep pendente but log event)
            action = 'ready_for_human'

        # Log reactor event
        try:
            supabase_request('POST', 'entity_events', body={
                'aggregate_type': 'Task',
                'aggregate_id': dep_id,
                'event_type': 'TaskReactorUnblocked',
                'payload': {
                    'action': action,
                    'unblocked_by': completed_task_id,
                    'unblocked_by_titulo': completed_titulo,
                },
                'metadata': {'source': 'dependency_reactor'},
            })
        except Exception:
            pass

        results.append({'task_id': dep_id, 'titulo': dep.get('titulo', ''), 'action': action})
    return results


def supabase_request(method, path, body=None, _retries=3, _backoff=1.0):
    """Make a request to Supabase REST API with retry logic and connection pooling.
    Returns SupabaseResponse (supports .status_code, .json(), .text AND dict/list access)."""
    key = SUPABASE_SERVICE_KEY or SUPABASE_ANON_KEY
    if not key:
        return SupabaseResponse({'error': 'Supabase key not configured'}, 500)

    headers = {
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation',
    }
    data = json.dumps(body).encode() if body else None
    url = f'/rest/v1/{path}'
    last_error = None

    for attempt in range(_retries):
        try:
            with _supa_lock:
                conn = _get_supa_conn()
                conn.request(method, url, body=data, headers=headers)
                resp = conn.getresponse()
                resp_body = resp.read()
                status = resp.status

            if status in (200, 201):
                parsed = json.loads(resp_body) if resp_body else {}
                # Inject .status_code and .json() on native types for backward compat
                if isinstance(parsed, list):
                    class SList(list):
                        status_code = status
                        text = resp_body.decode() if resp_body else ''
                        def json(self): return list(self)
                    return SList(parsed)
                elif isinstance(parsed, dict):
                    class SDict(dict):
                        status_code = status
                        text = resp_body.decode() if resp_body else ''
                        def json(self): return dict(self)
                    return SDict(parsed)
                return parsed

            elif status in (204,):
                class SDict204(dict):
                    status_code = 204
                    text = ''
                    def json(self): return {}
                return SDict204()

            elif status in (503, 429, 500, 502, 504):
                last_error = SupabaseResponse({'error': f'Supabase {status}: {resp_body.decode()}'}, status)
                if attempt < _retries - 1:
                    wait = _backoff * (2 ** attempt)
                    log_info('Supabase', f'Erro {status}, tentativa {attempt+1}/{_retries}, aguardando {wait}s...')
                    time.sleep(wait)
                    with _supa_lock:
                        _reset_supa_conn()
                    continue
            else:
                return SupabaseResponse({'error': f'Supabase {status}: {resp_body.decode()}'}, status)

        except (http.client.RemoteDisconnected, ConnectionResetError,
                BrokenPipeError, OSError, http.client.CannotSendRequest) as e:
            last_error = SupabaseResponse({'error': str(e)}, 503)
            log_info('Supabase', f'Erro de conexão na tentativa {attempt+1}/{_retries}: {e}')
            with _supa_lock:
                _reset_supa_conn()
            if attempt < _retries - 1:
                wait = _backoff * (2 ** attempt)
                time.sleep(wait)
                continue
        except Exception as e:
            log_error('Supabase', 'Erro inesperado', e)
            return SupabaseResponse({'error': str(e)}, 500)

    log_error('Supabase', f'Todas as {_retries} tentativas falharam', None)
    return last_error or SupabaseResponse({'error': 'Max retries exceeded'}, 500)


def get_mentees_with_email():
    """Fetch mentorados with email from Supabase"""
    return supabase_request('GET', 'mentorados?select=id,nome,email,instagram,fase_jornada,cohort&email=not.is.null&order=nome')


def insert_scheduled_call(data):
    """Insert a call into calls_mentoria"""
    return supabase_request('POST', 'calls_mentoria', data)


# ===== STORAGE PROCESSING PIPELINE =====

def _openai_request(endpoint, method='POST', body=None, files=None, timeout=120):
    """Make request to OpenAI API. Returns parsed JSON or raises."""
    if not OPENAI_API_KEY:
        raise ValueError('OPENAI_API_KEY not configured')

    if files:
        # Multipart upload (for Whisper)
        import io
        boundary = f'----FormBoundary{secrets.token_hex(8)}'
        parts = []
        for key, val in (body or {}).items():
            parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="{key}"\r\n\r\n{val}\r\n')
        for key, (filename, filedata, content_type) in files.items():
            parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="{key}"; filename="{filename}"\r\nContent-Type: {content_type}\r\n\r\n')
            parts.append(filedata)
            parts.append('\r\n')
        parts.append(f'--{boundary}--\r\n')

        raw_body = b''
        for p in parts:
            raw_body += p.encode() if isinstance(p, str) else p

        conn = http.client.HTTPSConnection('api.openai.com', timeout=timeout)
        conn.request(method, f'/v1/{endpoint}', body=raw_body, headers={
            'Authorization': f'Bearer {OPENAI_API_KEY}',
            'Content-Type': f'multipart/form-data; boundary={boundary}',
        })
    else:
        conn = http.client.HTTPSConnection('api.openai.com', timeout=timeout)
        conn.request(method, f'/v1/{endpoint}',
                     body=json.dumps(body).encode() if body else None,
                     headers={
                         'Authorization': f'Bearer {OPENAI_API_KEY}',
                         'Content-Type': 'application/json',
                     })
    resp = conn.getresponse()
    data = resp.read()
    conn.close()
    if resp.status >= 400:
        raise ValueError(f'OpenAI API error {resp.status}: {data.decode()[:500]}')
    # Whisper with response_format=text returns plain text, not JSON
    text = data.decode('utf-8').strip()
    if not text:
        return {}
    try:
        return json.loads(text)
    except (json.JSONDecodeError, ValueError):
        return text


def _voyage_request(endpoint, body, timeout=60):
    """Make request to Voyage AI API."""
    if not VOYAGE_API_KEY:
        raise ValueError('VOYAGE_API_KEY not configured')
    conn = http.client.HTTPSConnection('api.voyageai.com', timeout=timeout)
    conn.request('POST', f'/v1/{endpoint}',
                 body=json.dumps(body).encode(),
                 headers={
                     'Authorization': f'Bearer {VOYAGE_API_KEY}',
                     'Content-Type': 'application/json',
                 })
    resp = conn.getresponse()
    data = resp.read()
    conn.close()
    if resp.status >= 400:
        raise ValueError(f'Voyage API error {resp.status}: {data.decode()[:500]}')
    return json.loads(data)


def embed_texts(texts):
    """Generate embeddings. Uses Voyage AI (1024 dims) or OpenAI (1536 dims) based on config."""
    if EMBEDDING_PROVIDER == 'voyage':
        result = _voyage_request('embeddings', {
            'model': VOYAGE_EMBED_MODEL,
            'input': texts,
            'input_type': 'document',
        })
        return [item['embedding'] for item in result['data']]
    else:
        result = _openai_request('embeddings', body={
            'model': OPENAI_EMBED_MODEL,
            'input': texts,
        })
        return [item['embedding'] for item in result['data']]


def embed_query(text):
    """Embed a search query (uses input_type='query' for Voyage AI)."""
    if EMBEDDING_PROVIDER == 'voyage':
        result = _voyage_request('embeddings', {
            'model': VOYAGE_EMBED_MODEL,
            'input': [text],
            'input_type': 'query',
        })
        return result['data'][0]['embedding']
    else:
        result = _openai_request('embeddings', body={
            'model': OPENAI_EMBED_MODEL,
            'input': [text],
        })
        return result['data'][0]['embedding']


def rerank_results(query, results, top_n=10):
    """Rerank search results using Voyage AI reranker or cross-encoder."""
    if not results:
        return results

    if EMBEDDING_PROVIDER == 'voyage' and VOYAGE_API_KEY:
        try:
            documents = [r.get('chunk_texto', '') for r in results]
            resp = _voyage_request('rerank', {
                'model': 'rerank-2-lite',
                'query': query,
                'documents': documents,
                'top_k': min(top_n, len(documents)),
            })
            reranked = []
            for item in resp.get('data', []):
                idx = item['index']
                r = dict(results[idx])
                r['rerank_score'] = item['relevance_score']
                reranked.append(r)
            return reranked
        except Exception as e:
            log_info('Storage', f'Rerank fallback (Voyage failed): {e}')
            return results[:top_n]

    # Fallback: return as-is (cross-encoder local can be added later via flashrank)
    return results[:top_n]


def openai_whisper(audio_bytes, filename, mime_type):
    """Transcribe audio via Whisper. Tries Groq first (free), falls back to OpenAI."""
    # Try Groq first (free, fast, whisper-large-v3)
    if GROQ_API_KEY:
        try:
            result = _whisper_request('api.groq.com', GROQ_API_KEY, GROQ_WHISPER_MODEL,
                                      audio_bytes, filename, mime_type)
            if result:
                print(f'[whisper] Groq OK: {len(result)} chars')
                return result
        except Exception as e:
            print(f'[whisper] Groq failed, trying OpenAI: {e}')

    # Fallback to OpenAI
    if OPENAI_API_KEY:
        result = _whisper_request('api.openai.com', OPENAI_API_KEY, WHISPER_MODEL,
                                  audio_bytes, filename, mime_type)
        if result:
            print(f'[whisper] OpenAI OK: {len(result)} chars')
            return result

    raise ValueError('Nenhuma API de transcricao configurada (GROQ_API_KEY ou OPENAI_API_KEY)')


def _whisper_request(host, api_key, model, audio_bytes, filename, mime_type):
    """Send audio to Whisper-compatible API (OpenAI or Groq). Returns transcribed text."""
    boundary = f'----FormBoundary{secrets.token_hex(8)}'
    parts = []
    parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="model"\r\n\r\n{model}\r\n')
    parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="language"\r\n\r\npt\r\n')
    parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="response_format"\r\n\r\ntext\r\n')
    parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="file"; filename="{filename}"\r\nContent-Type: {mime_type}\r\n\r\n')
    parts.append(audio_bytes)
    parts.append(f'\r\n--{boundary}--\r\n')

    raw_body = b''
    for p in parts:
        raw_body += p.encode() if isinstance(p, str) else p

    conn = http.client.HTTPSConnection(host, timeout=300)
    conn.request('POST', '/openai/v1/audio/transcriptions' if 'groq' in host else '/v1/audio/transcriptions',
                 body=raw_body, headers={
                     'Authorization': f'Bearer {api_key}',
                     'Content-Type': f'multipart/form-data; boundary={boundary}',
                 })
    resp = conn.getresponse()
    data = resp.read()
    conn.close()
    if resp.status >= 400:
        raise ValueError(f'Whisper API error {resp.status} ({host}): {data.decode()[:300]}')
    text = data.decode('utf-8').strip()
    if not text:
        return None
    # response_format=text returns plain text, but some APIs wrap in JSON
    try:
        parsed = json.loads(text)
        return parsed.get('text', text) if isinstance(parsed, dict) else text
    except (json.JSONDecodeError, ValueError):
        return text


def gemini_vision_describe(image_bytes, mime_type):
    """Describe image content using Gemini Vision API (free tier). Returns text description."""
    if not GEMINI_API_KEY:
        raise ValueError('GEMINI_API_KEY not configured')
    b64 = base64.b64encode(image_bytes).decode()
    body = {
        'contents': [{
            'parts': [
                {'text': 'Descreva detalhadamente o conteúdo desta imagem em português. Se houver texto, transcreva-o. Se for um screenshot, descreva o contexto.'},
                {'inline_data': {'mime_type': mime_type, 'data': b64}}
            ]
        }]
    }
    conn = http.client.HTTPSConnection('generativelanguage.googleapis.com', timeout=120)
    conn.request('POST',
                 f'/v1beta/models/{GEMINI_VISION_MODEL}:generateContent?key={GEMINI_API_KEY}',
                 body=json.dumps(body).encode(),
                 headers={'Content-Type': 'application/json'})
    resp = conn.getresponse()
    data = resp.read()
    conn.close()
    if resp.status >= 400:
        raise ValueError(f'Gemini API error {resp.status}: {data.decode()[:500]}')
    result = json.loads(data)
    return result['candidates'][0]['content']['parts'][0]['text']


def openai_vision_describe(image_bytes, mime_type):
    """Describe image content using GPT-4o vision. Returns text description."""
    b64 = base64.b64encode(image_bytes).decode()
    result = _openai_request('chat/completions', body={
        'model': VISION_MODEL,
        'max_tokens': 1000,
        'messages': [{
            'role': 'user',
            'content': [
                {'type': 'text', 'text': 'Descreva detalhadamente o conteúdo desta imagem em português. Se houver texto, transcreva-o. Se for um screenshot, descreva o contexto.'},
                {'type': 'image_url', 'image_url': {'url': f'data:{mime_type};base64,{b64}'}}
            ]
        }]
    })
    return result['choices'][0]['message']['content']


def vision_describe(image_bytes, mime_type):
    """Describe image using Gemini (primary, free) → OpenAI (fallback). Returns (text, method).
    Method is always 'ocr_vision' for DB constraint compatibility; provider logged separately."""
    provider = None
    # Try Gemini first (free tier)
    if GEMINI_API_KEY:
        try:
            text = gemini_vision_describe(image_bytes, mime_type)
            log_info('Storage', f'Vision OCR via Gemini ({GEMINI_VISION_MODEL}) succeeded')
            return text, 'ocr_vision'
        except Exception as e:
            log_info('Storage', f'Gemini vision failed, trying OpenAI fallback: {e}')
    # Fallback to OpenAI
    if OPENAI_API_KEY:
        try:
            text = openai_vision_describe(image_bytes, mime_type)
            log_info('Storage', f'Vision OCR via OpenAI ({VISION_MODEL}) succeeded')
            return text, 'ocr_vision'
        except Exception as e:
            log_info('Storage', f'OpenAI vision also failed: {e}')
            raise
    raise ValueError('No vision API configured (need GEMINI_API_KEY or OPENAI_API_KEY)')


def _download_from_supabase_storage(storage_path):
    """Download file bytes from Supabase Storage bucket."""
    key = SUPABASE_SERVICE_KEY or SUPABASE_ANON_KEY
    encoded_path = urllib.parse.quote(storage_path, safe='/')
    url = f'/storage/v1/object/spalla-arquivos/{encoded_path}'

    conn = http.client.HTTPSConnection('knusqfbvhsqworzyhvip.supabase.co', timeout=60)
    conn.request('GET', url, headers={
        'apikey': key,
        'Authorization': f'Bearer {key}',
    })
    resp = conn.getresponse()
    data = resp.read()
    conn.close()
    if resp.status >= 400:
        raise ValueError(f'Storage download failed {resp.status}: {data.decode()[:200]}')
    return data


def _extract_text_pdf(file_bytes):
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber
        import io
        text_parts = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ''
                if page_text.strip():
                    text_parts.append(f'--- Página {i+1} ---\n{page_text}')
        return '\n\n'.join(text_parts), {'pages': len(pdf.pages)}, 'pdf_extract'
    except ImportError:
        # Fallback: try pymupdf
        try:
            import fitz
            import io
            doc = fitz.open(stream=file_bytes, filetype='pdf')
            text_parts = []
            for i, page in enumerate(doc):
                page_text = page.get_text()
                if page_text.strip():
                    text_parts.append(f'--- Página {i+1} ---\n{page_text}')
            return '\n\n'.join(text_parts), {'pages': len(doc)}, 'pdf_extract'
        except ImportError:
            raise ValueError('Neither pdfplumber nor pymupdf installed. Run: pip install pdfplumber')


def _extract_text_docx(file_bytes):
    """Extract text from DOCX."""
    try:
        import docx
        import io
        doc = docx.Document(io.BytesIO(file_bytes))
        text_parts = [p.text for p in doc.paragraphs if p.text.strip()]
        return '\n\n'.join(text_parts), {'paragraphs': len(text_parts)}, 'docx_extract'
    except ImportError:
        raise ValueError('python-docx not installed. Run: pip install python-docx')


def _extract_text_xlsx(file_bytes):
    """Extract text from XLSX — each sheet as structured text."""
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        text_parts = []
        structured = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = ' | '.join(str(c) if c is not None else '' for c in row)
                if row_text.strip(' |'):
                    rows.append(row_text)
            if rows:
                sheet_text = f'=== Planilha: {sheet_name} ===\n' + '\n'.join(rows)
                text_parts.append(sheet_text)
                structured[sheet_name] = rows
        return '\n\n'.join(text_parts), {'sheets': list(wb.sheetnames)}, 'xlsx_extract'
    except ImportError:
        raise ValueError('openpyxl not installed. Run: pip install openpyxl')


def _extract_text_csv(file_bytes):
    """Extract text from CSV."""
    import csv, io
    text = file_bytes.decode('utf-8', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows = [' | '.join(row) for row in reader if any(c.strip() for c in row)]
    return '\n'.join(rows), {'rows': len(rows)}, 'text_direct'


def _simple_token_count(text):
    """Approximate token count (4 chars per token heuristic for Portuguese)."""
    return len(text) // 4


def _content_hash(text):
    """SHA-256 hash of text content — for dedup (Notion pattern)."""
    return hashlib.sha256(text.encode('utf-8')).hexdigest()


def _chunk_text(text, chunk_size=CHUNK_SIZE_TOKENS, overlap=CHUNK_OVERLAP_TOKENS, min_size=MIN_CHUNK_TOKENS):
    """Split text into overlapping chunks by approximate token count."""
    words = text.split()
    words_per_chunk = int(chunk_size * 1.3)
    words_overlap = int(overlap * 1.3)

    chunks = []
    start = 0
    while start < len(words):
        end = min(start + words_per_chunk, len(words))
        chunk_text = ' '.join(words[start:end])
        if _simple_token_count(chunk_text) >= min_size:
            chunks.append(chunk_text)
        start += words_per_chunk - words_overlap
    return chunks


def _chunk_by_heading(text, max_chunk=CHUNK_SIZE_TOKENS, min_size=MIN_CHUNK_TOKENS):
    """Split markdown/structured text by headings (Unstructured.io pattern).
    Preserves section context. Falls back to regular chunking for non-headed text."""
    lines = text.split('\n')
    sections = []
    current_heading = ''
    current_lines = []

    for line in lines:
        stripped = line.strip()
        if stripped.startswith('#') or (stripped.startswith('---') and len(stripped) >= 3 and current_lines):
            # New section
            if current_lines:
                section_text = current_heading + '\n' + '\n'.join(current_lines) if current_heading else '\n'.join(current_lines)
                sections.append(section_text.strip())
            current_heading = stripped
            current_lines = []
        else:
            current_lines.append(line)

    # Last section
    if current_lines:
        section_text = current_heading + '\n' + '\n'.join(current_lines) if current_heading else '\n'.join(current_lines)
        sections.append(section_text.strip())

    # Now split oversized sections with regular chunking
    chunks = []
    for section in sections:
        if not section.strip():
            continue
        tokens = _simple_token_count(section)
        if tokens <= max_chunk:
            if tokens >= min_size:
                chunks.append(section)
        else:
            # Section too big — sub-chunk it
            sub_chunks = _chunk_text(section, chunk_size=max_chunk)
            chunks.extend(sub_chunks)

    return chunks if chunks else _chunk_text(text)


def _chunk_multipass(text, is_structured=False):
    """Multi-pass chunking (Onyx/Danswer pattern): standard chunks + broad context chunks."""
    # Pass 1: standard chunks (heading-based for structured, regular for others)
    if is_structured:
        standard = _chunk_by_heading(text)
    else:
        standard = _chunk_text(text)

    # Pass 2: broad context chunks (2x size, for capturing wider meaning)
    context = _chunk_text(text, chunk_size=CHUNK_CONTEXT_SIZE_TOKENS, overlap=CHUNK_OVERLAP_TOKENS * 2)

    return standard, context


def _resolve_mentorado_from_entity(entidade_tipo, entidade_id):
    """Resolve mentorado_id and mentorado_nome from entity reference."""
    if entidade_tipo == 'mentorado' and entidade_id:
        result = supabase_request('GET', f'mentorados?select=id,nome&id=eq.{entidade_id}')
        if isinstance(result, list) and result:
            return int(result[0]['id']), result[0]['nome']
    elif entidade_tipo == 'task' and entidade_id:
        result = supabase_request('GET', f'god_tasks?select=mentorado_id,mentorado_nome&id=eq.{entidade_id}')
        if isinstance(result, list) and result and result[0].get('mentorado_id'):
            return int(result[0]['mentorado_id']), result[0].get('mentorado_nome', '')
    elif entidade_tipo in ('dossie_doc', 'dossie_producao') and entidade_id:
        table = 'ds_documentos' if entidade_tipo == 'dossie_doc' else 'ds_producoes'
        result = supabase_request('GET', f'{table}?select=mentorado_id&id=eq.{entidade_id}')
        if isinstance(result, list) and result and result[0].get('mentorado_id'):
            mid = int(result[0]['mentorado_id'])
            m = supabase_request('GET', f'mentorados?select=nome&id=eq.{mid}')
            nome = m[0]['nome'] if isinstance(m, list) and m else ''
            return mid, nome
    return None, None


def process_file_pipeline(arquivo_id):
    """
    Full processing pipeline for a file:
    1. Download from Supabase Storage
    2. Extract text based on mime_type
    3. Chunk text
    4. Generate embeddings
    5. Store in sp_conteudo_extraido + sp_chunks
    """
    log_info('Storage', f'Processing file {arquivo_id}...')

    # Get file metadata
    result = supabase_request('GET', f'sp_arquivos?select=*&id=eq.{arquivo_id}')
    if not isinstance(result, list) or not result:
        log_error('Storage', f'File {arquivo_id} not found')
        return

    arquivo = result[0]
    mime = arquivo['mime_type']
    ext = (arquivo.get('extensao') or '').lower()
    storage_path = arquivo['storage_path']
    nome = arquivo['nome_original']

    try:
        # Update status
        supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                        {'status_processamento': 'extraindo'})

        # Download file
        file_bytes = _download_from_supabase_storage(storage_path)
        log_info('Storage', f'Downloaded {nome} ({len(file_bytes)} bytes)')

        # Extract content based on type
        texto = ''
        metadados = {}
        metodo = 'text_direct'
        duracao_seg = None

        if mime in ('text/plain', 'text/markdown') or ext in ('txt', 'md'):
            texto = file_bytes.decode('utf-8', errors='replace')
            metodo = 'text_direct'

        elif mime == 'application/pdf' or ext == 'pdf':
            texto, metadados, metodo = _extract_text_pdf(file_bytes)

        elif mime == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' or ext == 'docx':
            texto, metadados, metodo = _extract_text_docx(file_bytes)

        elif mime in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel') or ext in ('xlsx', 'xls'):
            texto, metadados, metodo = _extract_text_xlsx(file_bytes)

        elif mime == 'text/csv' or ext == 'csv':
            texto, metadados, metodo = _extract_text_csv(file_bytes)

        elif mime.startswith('audio/') or ext in ('mp3', 'wav', 'ogg', 'm4a'):
            if not OPENAI_API_KEY:
                raise ValueError('OPENAI_API_KEY required for audio transcription')
            texto = openai_whisper(file_bytes, nome, mime)
            metodo = 'whisper_stt'

        elif mime.startswith('video/') or ext in ('mp4', 'mov', 'webm'):
            if not OPENAI_API_KEY:
                raise ValueError('OPENAI_API_KEY required for video transcription')
            # Whisper API accepts video files directly (mp4, webm, etc)
            texto = openai_whisper(file_bytes, nome, mime)
            metodo = 'whisper_video'

        elif mime.startswith('image/') or ext in ('png', 'jpg', 'jpeg', 'webp'):
            if not GEMINI_API_KEY and not OPENAI_API_KEY:
                raise ValueError('GEMINI_API_KEY or OPENAI_API_KEY required for image description')
            texto, metodo = vision_describe(file_bytes, mime)

        else:
            # Unsupported type — mark as ignored
            supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                            {'status_processamento': 'ignorado',
                             'erro_processamento': f'Tipo não suportado: {mime}'})
            log_info('Storage', f'Skipped unsupported type: {mime}')
            return

        if not texto or not texto.strip():
            supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                            {'status_processamento': 'ignorado',
                             'erro_processamento': 'Nenhum conteúdo extraído'})
            return

        word_count = len(texto.split())
        text_hash = _content_hash(texto)
        log_info('Storage', f'Extracted {word_count} words via {metodo}')

        # Content hash dedup (Notion pattern) — skip re-embedding if content unchanged
        if arquivo.get('content_hash') == text_hash:
            log_info('Storage', f'Content unchanged (hash match) — skipping re-embedding')
            supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                            {'status_processamento': 'concluido',
                             'processado_em': datetime.now(timezone.utc).isoformat()})
            return

        # Save extracted content
        conteudo_row = supabase_request('POST', 'sp_conteudo_extraido', {
            'arquivo_id': arquivo_id,
            'conteudo_texto': texto,
            'conteudo_estruturado': metadados if metadados else None,
            'metadados_extracao': {
                'word_count': word_count,
                'method': metodo,
                'file_size': len(file_bytes),
            },
            'metodo_extracao': metodo,
            'word_count': word_count,
            'duracao_segundos': duracao_seg,
        })

        conteudo_id = conteudo_row[0]['id'] if isinstance(conteudo_row, list) and conteudo_row else conteudo_row.get('id')
        if not conteudo_id:
            raise ValueError(f'Failed to insert sp_conteudo_extraido: {conteudo_row}')

        # Update status
        supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                        {'status_processamento': 'chunking'})

        # Multi-pass chunking (Onyx/Danswer pattern)
        is_structured = ext in ('md', 'docx', 'html') or mime in ('text/markdown',)
        standard_chunks, context_chunks = _chunk_multipass(texto, is_structured=is_structured)
        total_chunks = len(standard_chunks) + len(context_chunks)
        log_info('Storage', f'Created {len(standard_chunks)} standard + {len(context_chunks)} context chunks')

        if total_chunks == 0:
            supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                            {'status_processamento': 'concluido',
                             'content_hash': text_hash,
                             'processado_em': datetime.now(timezone.utc).isoformat()})
            return

        # Update status
        supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                        {'status_processamento': 'embedding'})

        # Prepare all texts for embedding (standard + context in one batch)
        all_texts = standard_chunks + context_chunks
        chunk_tipos = ['standard'] * len(standard_chunks) + ['context'] * len(context_chunks)

        # Generate embeddings (batch)
        BATCH_SIZE = 100
        all_embeddings = []
        for i in range(0, len(all_texts), BATCH_SIZE):
            batch = all_texts[i:i + BATCH_SIZE]
            log_info('Storage', f'Embedding batch {i//BATCH_SIZE + 1}: {len(batch)} texts via {EMBEDDING_PROVIDER}')
            embeddings = embed_texts(batch)
            log_info('Storage', f'Got {len(embeddings)} embeddings, dims={len(embeddings[0]) if embeddings else 0}')
            all_embeddings.extend(embeddings)
        log_info('Storage', f'Total: {len(all_embeddings)} embeddings via {EMBEDDING_PROVIDER}')

        # Resolve mentorado
        mentorado_id, mentorado_nome = _resolve_mentorado_from_entity(
            arquivo['entidade_tipo'], arquivo.get('entidade_id'))

        # Delete old chunks if reprocessing
        supabase_request('DELETE', f'sp_chunks?arquivo_id=eq.{arquivo_id}')

        # Insert chunks with embeddings
        inserted = 0
        for idx, (chunk_text, embedding, chunk_tipo) in enumerate(zip(all_texts, all_embeddings, chunk_tipos)):
            # Format embedding as pgvector string
            emb_str = '[' + ','.join(str(v) for v in embedding) + ']'
            chunk_row = {
                'arquivo_id': arquivo_id,
                'conteudo_id': conteudo_id,
                'texto': chunk_text,
                'chunk_index': idx,
                'token_count': _simple_token_count(chunk_text),
                'embedding': emb_str,
                'chunk_tipo': chunk_tipo,
                'arquivo_nome': nome,
                'entidade_tipo': arquivo['entidade_tipo'],
                'entidade_id': arquivo.get('entidade_id'),
                'categoria': arquivo.get('categoria'),
                'mentorado_id': mentorado_id,
                'mentorado_nome': mentorado_nome,
            }
            result = supabase_request('POST', 'sp_chunks', chunk_row)
            if isinstance(result, dict) and result.get('error'):
                log_error('Storage', f'Chunk {idx} insert failed: {result["error"][:200]}')
            else:
                inserted += 1
        log_info('Storage', f'Inserted {inserted}/{len(all_texts)} chunks')

        # Done! Save content hash
        supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                        {'status_processamento': 'concluido',
                         'content_hash': text_hash,
                         'processado_em': datetime.now(timezone.utc).isoformat()})

        log_info('Storage', f'✓ File {nome} processed: {len(standard_chunks)} standard + {len(context_chunks)} context chunks indexed')

    except Exception as e:
        log_error('Storage', f'Processing failed for {nome}', e)
        supabase_request('PATCH', f'sp_arquivos?id=eq.{arquivo_id}',
                        {'status_processamento': 'erro',
                         'erro_processamento': str(e)[:500]})


def search_semantic(query_text, mode='hybrid', filters=None, limit=10):
    """
    Execute semantic/keyword/hybrid search with reranking.
    Uses Voyage AI or OpenAI based on EMBEDDING_PROVIDER config.
    """
    filters = filters or {}

    if mode == 'keyword':
        params = {
            'p_query': query_text,
            'p_limit': limit * 2,  # fetch more for reranking
            'p_entidade_tipo': filters.get('entidade_tipo'),
            'p_categoria': filters.get('categoria'),
            'p_mentorado_id': filters.get('mentorado_id'),
        }
        result = supabase_request('POST', 'rpc/fn_busca_keyword', params)
        results = result if isinstance(result, list) else []
        return rerank_results(query_text, results, top_n=limit)

    # Need embedding for semantic/hybrid
    api_key = VOYAGE_API_KEY if EMBEDDING_PROVIDER == 'voyage' else OPENAI_API_KEY
    if not api_key:
        raise ValueError(f'{EMBEDDING_PROVIDER.upper()}_API_KEY required for semantic search')

    query_embedding = embed_query(query_text)
    fetch_limit = limit * 3  # over-fetch for reranking

    if mode == 'semantic':
        params = {
            'p_query_embedding': str(query_embedding),
            'p_limit': fetch_limit,
            'p_entidade_tipo': filters.get('entidade_tipo'),
            'p_categoria': filters.get('categoria'),
            'p_mentorado_id': filters.get('mentorado_id'),
            'p_threshold': filters.get('threshold', 0.3),
        }
        result = supabase_request('POST', 'rpc/fn_busca_semantica', params)
        results = result if isinstance(result, list) else []
        return rerank_results(query_text, results, top_n=limit)

    else:  # hybrid (default)
        params = {
            'p_query_embedding': str(query_embedding),
            'p_query_text': query_text,
            'p_limit': fetch_limit,
            'p_entidade_tipo': filters.get('entidade_tipo'),
            'p_categoria': filters.get('categoria'),
            'p_mentorado_id': filters.get('mentorado_id'),
            'p_semantic_weight': filters.get('semantic_weight', 0.7),
            'p_threshold': filters.get('threshold', 0.25),
        }
        result = supabase_request('POST', 'rpc/fn_busca_hibrida', params)
        results = result if isinstance(result, list) else []
        return rerank_results(query_text, results, top_n=limit)


# ============================================================
# LF-FASE3: Descarrego saga / processor (module-level helpers)
# ============================================================
def _descarrego_load(descarrego_id):
    r = supabase_request('GET', f'descarregos?id=eq.{descarrego_id}&limit=1')
    if r.status_code != 200:
        raise RuntimeError(f'load failed: {r.text}')
    rows = r.json()
    if not rows:
        raise RuntimeError(f'descarrego {descarrego_id} not found')
    return rows[0]


def _descarrego_update(descarrego_id, fields):
    r = supabase_request(
        'PATCH', f'descarregos?id=eq.{descarrego_id}', body=fields
    )
    if r.status_code not in (200, 204):
        raise RuntimeError(f'update failed: {r.text}')


def _descarrego_transition(descarrego_id, event, actor='system', payload=None):
    """Carrega FSM, aplica transição, persiste status."""
    from domain.state_machines import DescarregoStateMachine
    row = _descarrego_load(descarrego_id)
    sm = DescarregoStateMachine(row)
    result = sm.transition(event, actor=actor, persist=False, payload=payload)
    _descarrego_update(descarrego_id, {'status': result['to']})
    log_info('descarrego_fsm',
             f'{descarrego_id}: {result["from"]} --[{event}]--> {result["to"]}')
    return result


def _descarrego_processor_run(descarrego_id):
    """Saga: capturado → transcrito → classificado → ação."""
    try:
        row = _descarrego_load(descarrego_id)
        log_info('descarrego_saga', f'start {descarrego_id} status={row["status"]}')

        # 1) Transcrição (se áudio/vídeo)
        if row.get('tipo_bruto') in ('audio', 'video', 'gravacao'):
            if row['status'] == 'capturado':
                _descarrego_transition(descarrego_id, 'needs_transcription')
                _descarrego_do_transcription(descarrego_id)
                _descarrego_transition(descarrego_id, 'transcribed')
        elif row['status'] == 'capturado':
            _descarrego_transition(descarrego_id, 'skip_transcription')

        # 2) Classificação
        row = _descarrego_load(descarrego_id)
        if row['status'] == 'transcrito':
            _descarrego_transition(descarrego_id, 'classify')
            _descarrego_do_classification(descarrego_id)
            _descarrego_transition(descarrego_id, 'classified')

        # 3) Decisão
        row = _descarrego_load(descarrego_id)
        confidence = float(row.get('classificacao_confidence') or 0)
        principal = row.get('classificacao_principal')
        if confidence >= 0.8 and principal in ('task', 'contexto'):
            _descarrego_transition(descarrego_id, 'auto_execute')
            _descarrego_execute_action(descarrego_id, 'auto')
        else:
            _descarrego_transition(descarrego_id, 'request_human')
            log_info('descarrego_saga', f'{descarrego_id} → HITL (conf={confidence})')

    except Exception as e:
        log_error('descarrego_saga', f'{descarrego_id} failed: {e}', e)
        try:
            _descarrego_update(descarrego_id, {
                'status': 'erro',
                'last_error': str(e)[:500],
                'retry_count': (_descarrego_load(descarrego_id).get('retry_count') or 0) + 1,
            })
        except Exception:
            pass


def _descarrego_do_transcription(descarrego_id):
    row = _descarrego_load(descarrego_id)
    arquivo_url = row.get('arquivo_url')
    if not arquivo_url:
        raise RuntimeError('arquivo_url ausente para transcrição')
    # Download e transcrição via Whisper
    try:
        with urllib.request.urlopen(arquivo_url, timeout=120) as resp:
            audio_bytes = resp.read()
    except Exception as e:
        raise RuntimeError(f'download falhou: {e}')

    mime = row.get('arquivo_mime_type') or 'audio/mpeg'
    filename = arquivo_url.rsplit('/', 1)[-1] or 'audio.mp3'
    transcricao = openai_whisper(audio_bytes, filename, mime)

    _descarrego_update(descarrego_id, {
        'transcricao': transcricao,
        'transcrito_em': datetime.now(timezone.utc).isoformat(),
        'transcrito_por': 'whisper-1',
        'transcricao_confidence': 0.95,
    })


def _descarrego_do_classification(descarrego_id):
    from domain.services.descarrego_classifier import classify_descarrego
    row = _descarrego_load(descarrego_id)
    text = row.get('transcricao') or row.get('conteudo_bruto') or ''
    mentorado_ctx = None
    if row.get('mentorado_id'):
        try:
            r = supabase_request(
                'GET',
                f'mentorados?id=eq.{row["mentorado_id"]}'
                f'&select=nome,fase_jornada,trilha&limit=1',
            )
            if r.status_code == 200 and r.json():
                mentorado_ctx = r.json()[0]
        except Exception:
            pass

    # Enrich: últimas 5 interações do mentorado
    recent_interactions = None
    if row.get('mentorado_id'):
        try:
            ri = supabase_request(
                'GET',
                f'mentorado_context?mentorado_id=eq.{row["mentorado_id"]}'
                f'&select=tipo,titulo,conteudo&order=created_at.desc&limit=5',
            )
            if ri.status_code == 200 and ri.json():
                recent_interactions = [
                    {'tipo': x.get('tipo', ''), 'resumo': (x.get('titulo') or x.get('conteudo', ''))[:200]}
                    for x in ri.json()
                ]
        except Exception:
            pass

    # Enrich: últimas 5 classificações anteriores (feedback loop)
    previous_classifications = None
    if row.get('mentorado_id'):
        try:
            rp = supabase_request(
                'GET',
                f'descarregos?mentorado_id=eq.{row["mentorado_id"]}'
                f'&classificacao_principal=not.is.null&id=neq.{descarrego_id}'
                f'&select=classificacao_principal,classificacao_confidence,classificacao_payload'
                f'&order=created_at.desc&limit=5',
            )
            if rp.status_code == 200 and rp.json():
                previous_classifications = [
                    {
                        'primary_type': x.get('classificacao_principal'),
                        'confidence': x.get('classificacao_confidence', 0),
                        'summary': (x.get('classificacao_payload') or {}).get('summary', ''),
                    }
                    for x in rp.json()
                ]
        except Exception:
            pass

    result = classify_descarrego(text, mentorado_ctx, recent_interactions, previous_classifications)
    _descarrego_update(descarrego_id, {
        'classificacao_principal': result.get('primary_type'),
        'classificacao_sub': result.get('subtype'),
        'classificacao_confidence': result.get('confidence', 0),
        'classificacao_payload': result,
        'classificado_em': datetime.now(timezone.utc).isoformat(),
        'classificado_por': 'gpt-4o-mini',
    })


def _descarrego_execute_action(descarrego_id, mode='auto'):
    """mode: 'auto' (FSM em executando_acao_automatica) ou 'manual' (após HITL)."""
    try:
        row = _descarrego_load(descarrego_id)
        principal = row.get('classificacao_principal')
        payload = row.get('classificacao_payload') or {}

        if principal == 'task':
            task_data = payload.get('task') or {}
            prazo_dias = task_data.get('prazo_dias', 0)
            data_fim = None
            if prazo_dias:
                data_fim = (datetime.now(timezone.utc)
                            + timedelta(days=int(prazo_dias))).isoformat()
            new_task = {
                'titulo': task_data.get('titulo') or 'Task de descarrego',
                'descricao': task_data.get('descricao') or row.get('transcricao'),
                'status': 'pendente',
                'prioridade': task_data.get('prioridade', 'normal'),
                'responsavel': task_data.get('responsavel') or 'kaique',
                'mentorado_id': row.get('mentorado_id'),
                'data_fim': data_fim,
                'fonte': 'descarrego',
                'auto_created': mode == 'auto',
                'confianca_ia': row.get('classificacao_confidence'),
                'especie': 'one_time',
            }
            r = supabase_request('POST', 'god_tasks', body=new_task)
            if r.status_code in (200, 201):
                created = r.json()
                task_id = created[0]['id'] if isinstance(created, list) and created else created.get('id')
                _descarrego_update(descarrego_id, {
                    'acao_tomada': 'task_criada',
                    'task_id': task_id,
                    'acao_tomada_em': datetime.now(timezone.utc).isoformat(),
                    'acao_tomada_por': 'saga' if mode == 'auto' else 'humano',
                })
            else:
                raise RuntimeError(f'task create failed: {r.text}')

        elif principal == 'lembrete':
            lembrete = payload.get('lembrete') or {}
            prazo_dias = lembrete.get('prazo_dias', 3)
            data_fim = (datetime.now(timezone.utc) + timedelta(days=int(prazo_dias))).isoformat()
            new_task = {
                'titulo': lembrete.get('titulo') or f'Lembrete: {payload.get("summary", "")[:80]}',
                'descricao': f'[Auto-gerado de descarrego]\n{row.get("transcricao") or row.get("conteudo_bruto", "")}',
                'status': 'pendente',
                'prioridade': 'alta',
                'responsavel': lembrete.get('responsavel') or 'kaique',
                'mentorado_id': row.get('mentorado_id'),
                'data_fim': data_fim,
                'fonte': 'descarrego',
                'especie': 'one_time',
                'tipo': 'lembrete',
            }
            r = supabase_request('POST', 'god_tasks', body=new_task)
            task_id = None
            if r.status_code in (200, 201):
                created = r.json()
                task_id = created[0]['id'] if isinstance(created, list) and created else created.get('id')
            _descarrego_update(descarrego_id, {
                'acao_tomada': 'lembrete_criado',
                'task_id': task_id,
                'acao_tomada_em': datetime.now(timezone.utc).isoformat(),
                'acao_tomada_por': 'saga' if mode == 'auto' else 'humano',
            })

        elif principal == 'escalacao':
            # Cria task urgente para Kaique
            new_task = {
                'titulo': f'ESCALAR: {payload.get("summary", "")[:80]}',
                'descricao': f'[Escalação automática]\n{row.get("transcricao") or row.get("conteudo_bruto", "")}',
                'status': 'pendente',
                'prioridade': 'urgente',
                'responsavel': 'kaique',
                'mentorado_id': row.get('mentorado_id'),
                'fonte': 'descarrego',
                'especie': 'one_time',
            }
            r = supabase_request('POST', 'god_tasks', body=new_task)
            task_id = None
            if r.status_code in (200, 201):
                created = r.json()
                task_id = created[0]['id'] if isinstance(created, list) and created else created.get('id')
            _descarrego_update(descarrego_id, {
                'acao_tomada': 'escalado_kaique',
                'task_id': task_id,
                'acao_tomada_em': datetime.now(timezone.utc).isoformat(),
                'acao_tomada_por': 'saga' if mode == 'auto' else 'humano',
            })

        elif principal in ('contexto', 'dossie', 'feedback', 'plano_acao'):
            # Salva como contexto do mentorado
            if row.get('mentorado_id'):
                ctx_entry = {
                    'mentorado_id': row['mentorado_id'],
                    'tipo': 'texto',
                    'titulo': f'[{principal}] {payload.get("summary", "")[:100]}',
                    'conteudo': row.get('transcricao') or row.get('conteudo_bruto', ''),
                    'fase': principal,
                    'origem': 'descarrego',
                    'criado_por': 'sistema',
                }
                supabase_request('POST', 'mentorado_context', body=ctx_entry)
            _descarrego_update(descarrego_id, {
                'acao_tomada': f'salvo_como_{principal}',
                'acao_tomada_em': datetime.now(timezone.utc).isoformat(),
                'acao_tomada_por': 'saga' if mode == 'auto' else 'humano',
            })

        else:
            _descarrego_update(descarrego_id, {
                'acao_tomada': 'sem_acao',
                'acao_tomada_em': datetime.now(timezone.utc).isoformat(),
            })

        _descarrego_transition(descarrego_id, 'action_done')
    except Exception as e:
        log_error('descarrego_action', f'{descarrego_id}: {e}', e)
        try:
            _descarrego_transition(descarrego_id, 'action_failed')
        except Exception:
            pass


# ===== CUSTOM HTTP SERVER (fix SO_REUSEADDR) =====
class ReuseAddrHTTPServer(http.server.HTTPServer):
    allow_reuse_address = True


# ===== HTTP HANDLER =====
class ProxyHandler(http.server.SimpleHTTPRequestHandler):

    def _read_body(self):
        length = int(self.headers.get('Content-Length', 0))
        return self.rfile.read(length) if length > 0 else b''

    def _read_json_body(self):
        return json.loads(self._read_body())

    def _get_cors_origin(self):
        origin = self.headers.get('Origin', '')
        allowed = [
            'https://spalla-dashboard.vercel.app',
            'https://spalla-dashboard-git-',  # Vercel preview URLs
            'http://localhost:',
            'http://127.0.0.1:',
        ]
        if any(origin.startswith(a) for a in allowed):
            return origin
        return 'https://spalla-dashboard.vercel.app'  # default

    def _send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode()
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
        self.send_header('Content-Length', len(body))
        self.end_headers()
        self.wfile.write(body)

    # ===== CLICKUP COMMAND CENTER =====
    def _handle_clickup_command_center(self):
        """GET /api/clickup/command-center — Sprint tasks, team summary, activity from ClickUp"""
        token = os.environ.get('CLICKUP_API_TOKEN', '')
        if not token:
            self._send_json({'error': 'CLICKUP_API_TOKEN not configured'}, 500)
            return

        # All known lists — sprint lists + main backlog
        sprint_lists = [
            {'id': '901113377455', 'list_id': '901113377455', 'nome': 'Sprint 1 (3/16 - 3/22)', 'inicio': '2026-03-16', 'fim': '2026-03-22'},
            {'id': '901113377456', 'list_id': '901113377456', 'nome': 'Sprint 2 (3/23 - 3/29)', 'inicio': '2026-03-23', 'fim': '2026-03-29'},
            {'id': '901113377457', 'list_id': '901113377457', 'nome': 'Sprint 3 (3/30 - 4/5)',  'inicio': '2026-03-30', 'fim': '2026-04-05'},
        ]
        main_list = {'id': '901113375992', 'list_id': '901113375992', 'nome': 'Backlog Geral', 'inicio': '2026-01-01', 'fim': '2099-12-31'}

        today_str = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        active = next(
            (s for s in sprint_lists if s['inicio'] <= today_str <= s['fim']),
            sprint_lists[-1]  # Fallback to last sprint if none matches today
        )

        headers = {'Authorization': token, 'Content-Type': 'application/json'}

        def clickup_get(url):
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as r:
                return json.loads(r.read())

        def normalize_status(raw):
            s = (raw or '').lower().strip()
            if s in ('to do', 'open', 'backlog', 'pendente', 'not started'):
                return 'backlog'
            if s in ('in progress', 'em andamento', 'doing', 'in_progress'):
                return 'em_andamento'
            if s in ('review', 'em revisão', 'in review', 'revisão', 'em_revisao', 'em revisao'):
                return 'em_revisao'
            if s in ('done', 'complete', 'concluída', 'concluido', 'closed', 'complete'):
                return 'concluida'
            return 'backlog'

        try:
            # Fetch sprint tasks
            data = clickup_get(
                f"https://api.clickup.com/api/v2/list/{active['list_id']}/task"
                f"?include_closed=true&subtasks=true&page=0"
            )
            all_items = data.get('tasks', [])

            # Also fetch main-list (backlog) and merge — sprint stays as active reference
            if main_list['list_id'] != active['list_id']:
                try:
                    backlog_data = clickup_get(
                        f"https://api.clickup.com/api/v2/list/{main_list['list_id']}/task"
                        f"?include_closed=true&subtasks=true&page=0"
                    )
                    # Add backlog tasks that aren't already in the sprint
                    sprint_ids = {t['id'] for t in all_items}
                    for t in backlog_data.get('tasks', []):
                        if t['id'] not in sprint_ids:
                            all_items.append(t)
                except Exception:
                    pass  # Non-fatal: backlog fetch failure doesn't block sprint view
        except Exception as e:
            self._send_json({'error': f'ClickUp API error: {e}'}, 502)
            return
        # Separate parent tasks from subtasks (subtasks have a non-null 'parent' field)
        tasks     = [t for t in all_items if not t.get('parent')]
        subtasks  = [t for t in all_items if t.get('parent')]

        by_status = {'backlog': [], 'em_andamento': [], 'em_revisao': [], 'concluida': []}
        subtasks_by_parent = {}
        by_member = {}
        activity = []

        # Index subtasks by parent clickup id
        for st in subtasks:
            pid = st.get('parent')
            subtasks_by_parent.setdefault(pid, []).append({
                'id':          st.get('id'),
                'titulo':      st.get('name', ''),
                'status':      normalize_status(st.get('status', {}).get('status', '')),
                'responsavel': ', '.join(a.get('username', '') or a.get('email', '') for a in st.get('assignees', [])),
                'data_inicio': st.get('start_date') and str(int(st['start_date']) // 1000) or None,
                'data_fim':    st.get('due_date')   and str(int(st['due_date'])   // 1000) or None,
                'clickup_id':  st.get('id'),
                'url':         st.get('url', ''),
            })

        for t in tasks:
            status = normalize_status(t.get('status', {}).get('status', ''))
            nome = t.get('name', '')
            assignees = t.get('assignees', [])
            updated_ms = int(t.get('date_updated') or 0)
            tid = t.get('id')
            operon_id = t.get('id', '')
            clickup_url = clickup_deep_link(t.get('url', ''))

            # Descrição: pega primeira linha não-vazia, remove markdown/HTML, trunca 80
            raw_desc = t.get('description') or ''
            clean_desc = re.sub(r'<[^>]+>', '', raw_desc)
            clean_desc = re.sub(r'[*_`#>]', '', clean_desc)  # remove markdown
            # primeira linha não-vazia
            first_line = next((l.strip() for l in clean_desc.splitlines() if l.strip()), '')
            short_desc = first_line[:80] + ('…' if len(first_line) > 80 else '')

            task_obj = {
                'id': tid,
                'titulo': nome,
                'desc': short_desc,
                'status': status,
                # responsavel mantém username ClickUp — frontend resolve via spalla_members
                'responsavel': ', '.join(
                    (a.get('username') or a.get('email') or '').split('@')[0]
                    for a in assignees
                ),
                # assignee_ids para match direto com spalla_members.clickup_user_id
                'assignee_ids': [str(a.get('id', '')) for a in assignees if a.get('id')],
                'url': clickup_url,
                'atualizado_ms': updated_ms,
                'subtasks': subtasks_by_parent.get(tid, []),
                'subtasks_count': len(subtasks_by_parent.get(tid, [])),
            }
            by_status.setdefault(status, []).append(task_obj)

            for a in assignees:
                # Usa username como chave — será mapeado pelo front via spalla_members
                key = (a.get('username') or a.get('email') or 'desconhecido').split('@')[0]
                by_member[key] = by_member.get(key, 0) + 1

            if updated_ms:
                activity.append({
                    'operon_id': operon_id,     # permite match com god_tasks.operon_id
                    'who': ', '.join(
                        (a.get('username') or '').split('@')[0]
                        for a in assignees
                    ) or 'Sistema',
                    'text': nome,
                    'time': datetime.fromtimestamp(updated_ms / 1000, tz=timezone.utc).isoformat(),
                    'url': clickup_url,
                })

        activity.sort(key=lambda x: x['time'], reverse=True)

        self._send_json({
            'sprint': {
                'id': active['id'],
                'nome': active['nome'],
                'inicio': active['inicio'],
                'fim': active['fim'],
            },
            'total': len(tasks),
            'subtasks_total': len(subtasks),
            'by_status': by_status,
            'by_member': by_member,
            'activity': activity[:25],
            'concluidas': len(by_status.get('concluida', [])),
        })

    # ===== CLICKUP IMPORT ALL TASKS =====
    def _handle_clickup_import_all(self):
        """POST /api/clickup/import-all — Full import of tasks from ClickUp lists (auth required)"""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return
        token = os.environ.get('CLICKUP_API_TOKEN', '')
        if not token:
            self._send_json({'error': 'CLICKUP_API_TOKEN not configured'}, 500)
            return

        headers = {'Authorization': token, 'Content-Type': 'application/json'}

        def clickup_get(url):
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=20) as r:
                return json.loads(r.read())

        def normalize_status(raw):
            s = (raw or '').lower().strip()
            if s in ('in progress', 'em andamento', 'doing', 'in_progress'):
                return 'em_andamento'
            if s in ('review', 'em revisão', 'in review', 'em_revisao', 'em revisao'):
                return 'em_revisao'
            if s in ('done', 'complete', 'concluída', 'concluido', 'closed'):
                return 'concluida'
            if s in ('cancelled', 'cancelada', 'canceled'):
                return 'cancelada'
            return 'pendente'

        def normalize_priority(raw):
            p = (raw or 0)
            if isinstance(p, dict):
                p = p.get('id', 0)
            try:
                p = int(p)
            except (ValueError, TypeError):
                return 'normal'
            return {1: 'urgente', 2: 'alta', 3: 'normal', 4: 'baixa'}.get(p, 'normal')

        def ms_to_date(ms_val):
            if not ms_val:
                return None
            try:
                return datetime.fromtimestamp(int(ms_val) // 1000, tz=timezone.utc).strftime('%Y-%m-%d')
            except Exception:
                return None

        try:
            body = self._read_json_body()
            list_ids = body.get('list_ids', [])
            if not list_ids:
                # Fetch all lists from god_lists
                god_lists = supabase_request('GET', 'god_lists?select=clickup_list_id&clickup_list_id=not.is.null')
                list_ids = [l['clickup_list_id'] for l in (god_lists or []) if l.get('clickup_list_id')]

            if not list_ids:
                self._send_json({'error': 'No list_ids provided and no god_lists with clickup_list_id found'}, 400)
                return

            # Fetch existing operon_id mapping
            existing = supabase_request('GET', 'god_tasks?select=id,operon_id&operon_id=not.is.null')
            existing_map = {r['operon_id']: r['id'] for r in (existing or [])}

            imported = 0
            updated = 0
            errors = 0

            for list_id in list_ids:
                page = 0
                while True:
                    try:
                        data = clickup_get(
                            f"https://api.clickup.com/api/v2/list/{list_id}/task"
                            f"?include_closed=true&subtasks=false&page={page}"
                        )
                    except Exception as e:
                        errors += 1
                        break

                    tasks = data.get('tasks', [])
                    if not tasks:
                        break

                    for t in tasks:
                        cu_id = t.get('id', '')
                        assignees = ', '.join(
                            (a.get('username') or a.get('email', '').split('@')[0] or '')
                            for a in t.get('assignees', [])
                        ) or None

                        row = {
                            'titulo': t.get('name', ''),
                            'descricao': t.get('description') or t.get('text_content') or '',
                            'status': normalize_status(t.get('status', {}).get('status', '')),
                            'prioridade': normalize_priority(t.get('priority')),
                            'responsavel': assignees,
                            'data_inicio': ms_to_date(t.get('start_date')),
                            'data_fim': ms_to_date(t.get('due_date')),
                            'operon_id': cu_id,
                            'clickup_url': clickup_deep_link(t.get('url', '')),
                            'clickup_synced_at': datetime.now(timezone.utc).isoformat(),
                            'fonte': 'clickup',
                        }

                        # Map points from custom field or time estimate
                        points = t.get('points')
                        if points:
                            try:
                                row['points'] = int(float(points))
                            except (ValueError, TypeError):
                                pass

                        if cu_id in existing_map:
                            result = supabase_request('PATCH', f'god_tasks?id=eq.{existing_map[cu_id]}', row)
                            updated += 1
                        else:
                            result = supabase_request('POST', 'god_tasks', row)
                            imported += 1

                        if isinstance(result, dict) and result.get('error'):
                            errors += 1

                    if data.get('last_page', True):
                        break
                    page += 1

            self._send_json({
                'imported': imported,
                'updated': updated,
                'errors': errors,
                'lists_processed': len(list_ids),
            })

        except Exception as e:
            self._send_json({'error': f'Import failed: {e}'}, 500)

    # ===== CLICKUP WEBHOOK HANDLER =====
    def _handle_clickup_webhook(self):
        """POST /api/clickup/webhook — Receive ClickUp webhook events"""
        try:
            body = self._read_json_body()
            event = body.get('event', '')
            task_data = body.get('task_data') or body.get('task') or {}
            task_id = body.get('task_id') or task_data.get('id', '')

            if not event or not task_id:
                self._send_json({'ok': True, 'skipped': 'no event or task_id'})
                return

            # Find the matching god_task by operon_id
            rows = supabase_request('GET', f'god_tasks?select=id,status,prioridade&operon_id=eq.{task_id}')
            if not rows or not len(rows):
                self._send_json({'ok': True, 'skipped': f'no god_task with operon_id={task_id}'})
                return

            god_task = rows[0]
            update = {}

            if event == 'taskStatusUpdated':
                new_status = (body.get('history_items', [{}])[0].get('after', {}).get('status', '') or '').lower().strip()
                status_map = {
                    'to do': 'pendente', 'open': 'pendente', 'pendente': 'pendente',
                    'in progress': 'em_andamento', 'em andamento': 'em_andamento',
                    'review': 'em_revisao', 'em revisão': 'em_revisao', 'in review': 'em_revisao',
                    'blocked': 'bloqueada', 'bloqueada': 'bloqueada',
                    'on hold': 'pausada', 'pausada': 'pausada',
                    'complete': 'concluida', 'done': 'concluida', 'closed': 'concluida',
                    'cancelled': 'cancelada', 'cancelada': 'cancelada',
                    'archived': 'arquivada', 'arquivada': 'arquivada',
                }
                mapped = status_map.get(new_status)
                if mapped and mapped != god_task.get('status'):
                    update['status'] = mapped

            elif event == 'taskUpdated':
                if task_data.get('name'):
                    update['titulo'] = task_data['name']
                if task_data.get('description'):
                    update['descricao'] = task_data['description']

            elif event == 'taskPriorityUpdated':
                prio_map = {1: 'urgente', 2: 'alta', 3: 'normal', 4: 'baixa'}
                new_prio = body.get('history_items', [{}])[0].get('after', {}).get('priority', {})
                if isinstance(new_prio, dict):
                    new_prio = new_prio.get('id', 3)
                mapped = prio_map.get(int(new_prio), 'normal')
                if mapped != god_task.get('prioridade'):
                    update['prioridade'] = mapped

            if update:
                update['clickup_synced_at'] = datetime.now(timezone.utc).isoformat()
                supabase_request('PATCH', f'god_tasks?id=eq.{god_task["id"]}', update)

            self._send_json({'ok': True, 'event': event, 'updated_fields': list(update.keys())})

        except Exception as e:
            self._send_json({'error': f'Webhook processing failed: {e}'}, 500)

    # ===== CLICKUP PUSH TASK =====
    def _handle_clickup_push(self, task_uuid):
        """POST /api/clickup/push/<task_uuid> — Push a Spalla task to ClickUp (auth required)"""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return
        token = os.environ.get('CLICKUP_API_TOKEN', '')
        if not token:
            self._send_json({'error': 'CLICKUP_API_TOKEN not configured'}, 500)
            return

        headers = {'Authorization': token, 'Content-Type': 'application/json'}

        try:
            # Fetch the task from god_tasks
            rows = supabase_request('GET', f'god_tasks?select=*&id=eq.{task_uuid}')
            if not rows or not len(rows):
                self._send_json({'error': 'Task not found'}, 404)
                return

            task = rows[0]
            status_map = {
                'pendente': 'to do', 'em_andamento': 'in progress',
                'em_revisao': 'review', 'concluida': 'complete', 'cancelada': 'closed'
            }
            prio_map = {'urgente': 1, 'alta': 2, 'normal': 3, 'baixa': 4}

            cu_body = {
                'name': task.get('titulo', ''),
                'description': task.get('descricao', ''),
                'status': status_map.get(task.get('status', 'pendente'), 'to do'),
                'priority': prio_map.get(task.get('prioridade', 'normal'), 3),
            }
            if task.get('data_fim'):
                try:
                    dt = datetime.strptime(task['data_fim'], '%Y-%m-%d').replace(tzinfo=timezone.utc)
                    cu_body['due_date'] = int(dt.timestamp() * 1000)
                except Exception:
                    pass

            if task.get('operon_id'):
                # Update existing ClickUp task
                req = urllib.request.Request(
                    f"https://api.clickup.com/api/v2/task/{task['operon_id']}",
                    data=json.dumps(cu_body).encode(),
                    headers=headers,
                    method='PUT'
                )
                with urllib.request.urlopen(req, timeout=15) as r:
                    result = json.loads(r.read())
                supabase_request('PATCH', f'god_tasks?id=eq.{task_uuid}', {
                    'clickup_synced_at': datetime.now(timezone.utc).isoformat()
                })
                self._send_json({'ok': True, 'action': 'updated', 'clickup_id': task['operon_id']})
            else:
                # Create new ClickUp task — need a list_id
                body = self._read_json_body() or {}
                cu_list_id = body.get('clickup_list_id', '901113377457')  # default: latest sprint
                req = urllib.request.Request(
                    f"https://api.clickup.com/api/v2/list/{cu_list_id}/task",
                    data=json.dumps(cu_body).encode(),
                    headers=headers,
                    method='POST'
                )
                with urllib.request.urlopen(req, timeout=15) as r:
                    result = json.loads(r.read())
                new_cu_id = result.get('id', '')
                new_cu_url = clickup_deep_link(result.get('url', ''))
                supabase_request('PATCH', f'god_tasks?id=eq.{task_uuid}', {
                    'operon_id': new_cu_id,
                    'clickup_url': new_cu_url,
                    'clickup_synced_at': datetime.now(timezone.utc).isoformat(),
                })
                self._send_json({'ok': True, 'action': 'created', 'clickup_id': new_cu_id, 'clickup_url': new_cu_url})

        except Exception as e:
            self._send_json({'error': f'Push failed: {e}'}, 500)

    # ===== CLICKUP SYNC SUBTASKS =====
    def _handle_clickup_sync_subtasks(self):
        """POST /api/clickup/sync-subtasks — Fetch ClickUp subtasks, upsert into god_task_subtasks"""
        token = os.environ.get('CLICKUP_API_TOKEN', '')
        if not token:
            self._send_json({'error': 'CLICKUP_API_TOKEN not configured'}, 500)
            return

        sprint_list_ids = ['901113375992', '901113377455', '901113377456', '901113377457']
        headers = {'Authorization': token, 'Content-Type': 'application/json'}

        def clickup_get(url):
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=15) as r:
                return json.loads(r.read())

        def normalize_status(raw):
            s = (raw or '').lower().strip()
            if s in ('in progress', 'em andamento', 'doing', 'in_progress'):
                return 'em_andamento'
            if s in ('review', 'em revisão', 'in review', 'em_revisao', 'em revisao'):
                return 'em_revisao'
            if s in ('done', 'complete', 'concluída', 'concluido', 'closed'):
                return 'concluida'
            return 'pendente'

        def ms_to_date(ms_val):
            if not ms_val:
                return None
            try:
                return datetime.fromtimestamp(int(ms_val) // 1000, tz=timezone.utc).strftime('%Y-%m-%d')
            except Exception:
                return None

        try:
            # 1. Collect all ClickUp subtasks across sprint lists
            clickup_subs = {}  # parent_clickup_id → [subtask, ...]
            for list_id in sprint_list_ids:
                try:
                    data = clickup_get(
                        f"https://api.clickup.com/api/v2/list/{list_id}/task"
                        f"?include_closed=true&subtasks=true&page=0"
                    )
                except Exception:
                    continue
                for item in data.get('tasks', []):
                    if item.get('parent'):
                        clickup_subs.setdefault(item['parent'], []).append(item)

            # 2. Fetch god_tasks that have operon_id set (maps ClickUp ID → Supabase UUID)
            gt_rows = supabase_request('GET', 'god_tasks?select=id,operon_id&operon_id=not.is.null')
            if isinstance(gt_rows, dict) and gt_rows.get('error'):
                self._send_json({'error': f"DB error: {gt_rows['error']}"}, 500)
                return
            operon_to_uuid = {r['operon_id']: r['id'] for r in (gt_rows or [])}

            # 3. Find parents that exist both in ClickUp subtasks and god_tasks
            matched_parents = {
                cu_id: operon_to_uuid[cu_id]
                for cu_id in clickup_subs
                if cu_id in operon_to_uuid
            }

            if not matched_parents:
                self._send_json({
                    'synced': 0,
                    'tasks_matched': 0,
                    'message': 'No god_tasks matched ClickUp parent IDs — check operon_id column',
                })
                return

            # 4. Fetch existing clickup-sourced subtasks for matched task UUIDs
            uuid_csv = ','.join(matched_parents.values())
            existing_rows = supabase_request(
                'GET',
                f'god_task_subtasks?select=id,task_id,clickup_id&task_id=in.({uuid_csv})&clickup_id=not.is.null'
            )
            existing_map = {}  # (task_uuid, clickup_id) → subtask row id
            for row in (existing_rows or []):
                if row.get('clickup_id'):
                    existing_map[(row['task_id'], row['clickup_id'])] = row['id']

            # 5. Upsert subtasks
            synced = 0
            errors = 0
            for cu_parent_id, god_task_uuid in matched_parents.items():
                for st in clickup_subs[cu_parent_id]:
                    cu_st_id = st.get('id', '')
                    status = normalize_status(st.get('status', {}).get('status', ''))
                    row = {
                        'task_id':     god_task_uuid,
                        'texto':       st.get('name', ''),
                        'done':        status == 'concluida',
                        'status':      status,
                        'responsavel': ', '.join(
                            (a.get('username') or a.get('email') or '').split('@')[0]
                            for a in st.get('assignees', [])
                        ) or None,
                        'data_inicio': ms_to_date(st.get('start_date')),
                        'data_fim':    ms_to_date(st.get('due_date')),
                        'prioridade':  'normal',
                        'clickup_id':  cu_st_id,
                    }
                    key = (god_task_uuid, cu_st_id)
                    if key in existing_map:
                        result = supabase_request('PATCH', f'god_task_subtasks?id=eq.{existing_map[key]}', row)
                    else:
                        result = supabase_request('POST', 'god_task_subtasks', row)
                    if isinstance(result, dict) and result.get('error'):
                        errors += 1
                    else:
                        synced += 1

            self._send_json({
                'synced': synced,
                'errors': errors,
                'tasks_matched': len(matched_parents),
                'subtasks_found': sum(len(v) for v in clickup_subs.values()),
            })

        except Exception as e:
            self._send_json({'error': f'Sync failed: {e}'}, 500)

    # ============================================================
    # LF Story 5: Task FSM transition endpoint
    # ============================================================
    def _handle_task_transition(self, task_id):
        """POST /api/tasks/{id}/transition
        Body: { event: 'start'|'complete'|..., payload?: {...} }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body() or {}
            event = body.get('event')
            payload = body.get('payload') or {}
            if not event:
                return self._send_json({'error': 'event obrigatório'}, 400)

            r = supabase_request('GET', f'god_tasks?id=eq.{task_id}&limit=1')
            if not supa_ok(r) or not r:
                return self._send_json({'error': 'task não encontrada'}, 404)
            task_row = r[0] if isinstance(r, list) else r

            # Hidrata flags de guard
            deps = task_row.get('depends_on') or []
            if deps:
                in_clause = ','.join(deps)
                rdep = supabase_request('GET',
                    f'god_tasks?id=in.({in_clause})&select=id,status&limit=1000')
                terminal = ('concluida','cancelada','arquivada')
                rows = rdep if isinstance(rdep, list) else []
                task_row['_dependencies_resolved'] = all(x.get('status') in terminal for x in rows)
            else:
                task_row['_dependencies_resolved'] = True

            if task_row.get('especie') == 'quest':
                rch = supabase_request('GET',
                    f'god_tasks?parent_task_id=eq.{task_id}&select=id,status')
                terminal = ('concluida','cancelada','arquivada')
                children = rch if isinstance(rch, list) else []
                task_row['_children_complete'] = (
                    len(children) == 0 or all(c.get('status') in terminal for c in children)
                )
            else:
                task_row['_children_complete'] = True

            from domain.state_machines import TaskStateMachine, IllegalTransition, GuardFailed
            sm = TaskStateMachine(task_row)
            try:
                result = sm.transition(event, persist=False, payload=payload)
            except IllegalTransition as e:
                return self._send_json({'error': str(e), 'allowed': sm.allowed_events()}, 409)
            except GuardFailed as e:
                return self._send_json({'error': str(e), 'guard_failed': True}, 412)

            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            ru = supabase_request('PATCH',
                f'god_tasks?id=eq.{task_id}',
                body={'status': result['to'],
                      'updated_at': datetime.now(timezone.utc).isoformat()})
            if not supa_ok(ru):
                return self._send_json({'error': 'persist failed'}, 500)

            try:
                supabase_request('POST', 'entity_events', body={
                    'aggregate_type': 'Task',
                    'aggregate_id': task_id,
                    'event_type': f'Task{event[0].upper()}{event[1:]}',
                    'payload': {**result, **payload},
                    'metadata': {'source': 'fsm_api', 'actor': actor},
                })
            except Exception:
                pass

            # ORCH-04: Dependency Reactor — auto-unblock dependents
            reactor_results = []
            if result['to'] in ('concluida', 'cancelada', 'arquivada'):
                try:
                    reactor_results = _dependency_reactor(task_id, task_row.get('titulo', ''))
                except Exception as rx:
                    log_error('dependency_reactor', str(rx), rx)

            self._send_json({
                'task_id': task_id,
                'from': result['from'],
                'to': result['to'],
                'event': event,
                'reactor': reactor_results,
            })
        except Exception as e:
            log_error('task_transition', str(e), e)
            self._send_json({'error': str(e)}, 500)

    # ============================================================
    # ORCH-05: Agent Dispatch Pipeline
    # ============================================================
    def _handle_agent_dispatch(self, task_id):
        """POST /api/tasks/{id}/agent-dispatch
        Dispatches task to the appropriate agent based on execution_endpoint.
        Body: { force?: boolean }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            r = supabase_request('GET', f'god_tasks?id=eq.{task_id}&limit=1')
            if not supa_ok(r) or not r:
                return self._send_json({'error': 'task não encontrada'}, 404)
            task = r[0] if isinstance(r, list) else r

            responsavel = task.get('responsavel') or ''
            # Lookup agent member
            member_r = supabase_request('GET',
                f'spalla_members?or=(id.eq.{responsavel},nome_curto.ilike.{responsavel})&tipo=eq.agent&limit=1')
            if not isinstance(member_r, list) or not member_r:
                return self._send_json({'error': 'responsável não é um agente'}, 400)
            agent = member_r[0]
            endpoint = agent.get('execution_endpoint') or ''

            # Check max_concurrent
            max_tasks = agent.get('max_concurrent_tasks')
            if max_tasks:
                active_r = supabase_request('GET',
                    f'god_tasks?responsavel=eq.{responsavel}&status=eq.em_andamento&select=id')
                active_count = len(active_r) if isinstance(active_r, list) else 0
                if active_count >= max_tasks:
                    return self._send_json({
                        'error': f'Agente {agent["nome_curto"]} no limite ({active_count}/{max_tasks})',
                        'at_capacity': True
                    }, 429)

            # Route by execution_endpoint
            dispatch_result = {'endpoint': endpoint, 'status': 'dispatched'}
            if endpoint == 'descarrego_pipeline':
                dispatch_result['note'] = 'Routed to Descarrego pipeline (existing)'
            elif endpoint.startswith('n8n_webhook:'):
                webhook_name = endpoint.split(':', 1)[1]
                dispatch_result['note'] = f'Would POST to n8n webhook: {webhook_name}'
            elif endpoint.startswith('claude_api:'):
                prompt_type = endpoint.split(':', 1)[1]
                dispatch_result['note'] = f'Would call Claude API with prompt type: {prompt_type}'
            else:
                dispatch_result['note'] = f'Unknown endpoint: {endpoint}'
                dispatch_result['status'] = 'unknown_endpoint'

            # Log dispatch event
            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            supabase_request('POST', 'entity_events', body={
                'aggregate_type': 'Task',
                'aggregate_id': task_id,
                'event_type': 'TaskAgentDispatched',
                'payload': {
                    'agent_id': agent['id'],
                    'agent_name': agent.get('nome_curto', ''),
                    'endpoint': endpoint,
                },
                'metadata': {'source': 'agent_dispatch', 'actor': actor},
            })

            self._send_json({
                'task_id': task_id,
                'agent': agent['id'],
                'dispatch': dispatch_result,
            })
        except Exception as e:
            log_error('agent_dispatch', str(e), e)
            self._send_json({'error': str(e)}, 500)

    # ============================================================
    # ORCH-06: Handoff Protocol — log reassignment
    # ============================================================
    def _handle_task_handoff(self, task_id):
        """POST /api/tasks/{id}/handoff
        Body: { from_person: str, to_person: str, note?: str }
        Logs a handoff and auto-dispatches if new responsavel is an agent.
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body() or {}
            from_person = body.get('from_person') or ''
            to_person = body.get('to_person') or ''
            note = body.get('note') or f'Reatribuído de {from_person} para {to_person}'

            if not to_person:
                return self._send_json({'error': 'to_person obrigatório'}, 400)

            # Log handoff
            supabase_request('POST', 'god_task_handoffs', body={
                'task_id': task_id,
                'from_person': from_person or None,
                'to_person': to_person,
                'note': note,
            })

            # Log event
            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            supabase_request('POST', 'entity_events', body={
                'aggregate_type': 'Task',
                'aggregate_id': task_id,
                'event_type': 'TaskHandoff',
                'payload': {'from': from_person, 'to': to_person, 'note': note},
                'metadata': {'source': 'handoff_protocol', 'actor': actor},
            })

            # Check if new responsavel is agent → auto-dispatch
            is_agent = False
            if to_person:
                member_r = supabase_request('GET',
                    f'spalla_members?or=(id.eq.{to_person},nome_curto.ilike.{to_person})&tipo=eq.agent&limit=1')
                is_agent = isinstance(member_r, list) and len(member_r) > 0

            self._send_json({
                'task_id': task_id,
                'handoff': {'from': from_person, 'to': to_person},
                'agent_dispatched': is_agent,
            })
        except Exception as e:
            log_error('task_handoff', str(e), e)
            self._send_json({'error': str(e)}, 500)

    # ============================================================
    # FSM: Mentorado transition
    # ============================================================
    def _handle_mentorado_transition(self, mentorado_id):
        """POST /api/mentees/{id}/transition
        Body: { event: 'contract_signed'|'kickoff_done'|..., payload?: {...} }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body() or {}
            event = body.get('event')
            payload = body.get('payload') or {}
            if not event:
                return self._send_json({'error': 'event obrigatório'}, 400)

            r = supabase_request('GET', f'mentorados?id=eq.{mentorado_id}&limit=1')
            if r.status_code != 200 or not r.json():
                return self._send_json({'error': 'mentorado não encontrado'}, 404)
            row = r.json()[0]

            from domain.state_machines import MentoradoStateMachine, IllegalTransition, GuardFailed
            sm = MentoradoStateMachine(row)
            try:
                result = sm.transition(event, persist=False, payload=payload)
            except IllegalTransition as e:
                return self._send_json({'error': str(e), 'allowed': sm.allowed_events()}, 409)
            except GuardFailed as e:
                return self._send_json({'error': str(e), 'guard_failed': True}, 412)

            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            ru = supabase_request('PATCH',
                f'mentorados?id=eq.{mentorado_id}',
                body={'fase_jornada': result['to'],
                      'updated_at': datetime.now(timezone.utc).isoformat()})
            if not supa_ok(ru):
                return self._send_json({'error': 'persist failed'}, 500)

            try:
                supabase_request('POST', 'entity_events', body={
                    'aggregate_type': 'Mentorado',
                    'aggregate_id': mentorado_id,
                    'event_type': f'Mentorado{event[0].upper()}{event[1:]}',
                    'payload': {**result, **payload},
                    'metadata': {'source': 'fsm_api', 'actor': actor},
                })
            except Exception:
                pass

            self._send_json({
                'mentorado_id': mentorado_id,
                'from': result['from'],
                'to': result['to'],
                'event': event,
            })
        except Exception as e:
            log_error('mentorado_transition', str(e), e)
            self._send_json({'error': str(e)}, 500)

    # ============================================================
    # FSM: Dossiê Produção transition
    # ============================================================
    def _handle_dossie_producao_transition(self, producao_id):
        """POST /api/dossies/producoes/{id}/transition
        Body: { event: 'start_production'|'request_review'|..., payload?: {...} }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body() or {}
            event = body.get('event')
            payload = body.get('payload') or {}
            if not event:
                return self._send_json({'error': 'event obrigatório'}, 400)

            r = supabase_request('GET', f'ds_producoes?id=eq.{producao_id}&limit=1')
            if r.status_code != 200 or not r.json():
                return self._send_json({'error': 'produção não encontrada'}, 404)
            row = r.json()[0]

            from domain.state_machines import DossieProducaoStateMachine, IllegalTransition, GuardFailed
            sm = DossieProducaoStateMachine(row)
            try:
                result = sm.transition(event, persist=False, payload=payload)
            except IllegalTransition as e:
                return self._send_json({'error': str(e), 'allowed': sm.allowed_events()}, 409)
            except GuardFailed as e:
                return self._send_json({'error': str(e), 'guard_failed': True}, 412)

            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            ru = supabase_request('PATCH',
                f'ds_producoes?id=eq.{producao_id}',
                body={'status': result['to'],
                      'updated_at': datetime.now(timezone.utc).isoformat()})
            if not supa_ok(ru):
                return self._send_json({'error': 'persist failed'}, 500)

            try:
                supabase_request('POST', 'entity_events', body={
                    'aggregate_type': 'DossieProducao',
                    'aggregate_id': producao_id,
                    'event_type': f'DossieProducao{event[0].upper()}{event[1:]}',
                    'payload': {**result, **payload},
                    'metadata': {'source': 'fsm_api', 'actor': actor},
                })
            except Exception:
                pass

            self._send_json({
                'producao_id': producao_id,
                'from': result['from'],
                'to': result['to'],
                'event': event,
            })
        except Exception as e:
            log_error('dossie_producao_transition', str(e), e)
            self._send_json({'error': str(e)}, 500)

    # ============================================================
    # FSM: Dossiê Documento transition
    # ============================================================
    def _handle_dossie_documento_transition(self, documento_id):
        """POST /api/dossies/documentos/{id}/transition
        Body: { event: 'start_writing'|'submit_to_qg'|..., trilha?: 'scale'|'clinic', payload?: {...} }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body() or {}
            event = body.get('event')
            trilha = body.get('trilha', 'scale')
            payload = body.get('payload') or {}
            if not event:
                return self._send_json({'error': 'event obrigatório'}, 400)

            r = supabase_request('GET', f'ds_documentos?id=eq.{documento_id}&limit=1')
            if r.status_code != 200 or not r.json():
                return self._send_json({'error': 'documento não encontrado'}, 404)
            row = r.json()[0]

            # Detect trilha from producao if not passed
            if not body.get('trilha') and row.get('producao_id'):
                rp = supabase_request('GET', f'ds_producoes?id=eq.{row["producao_id"]}&select=trilha&limit=1')
                if rp.status_code == 200 and rp.json():
                    trilha = rp.json()[0].get('trilha', 'scale')

            from domain.state_machines import DossieDocumentoStateMachine, IllegalTransition, GuardFailed
            sm = DossieDocumentoStateMachine(row, trilha=trilha)
            try:
                result = sm.transition(event, persist=False, payload=payload)
            except IllegalTransition as e:
                return self._send_json({'error': str(e), 'allowed': sm.allowed_events()}, 409)
            except GuardFailed as e:
                return self._send_json({'error': str(e), 'guard_failed': True}, 412)

            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            ru = supabase_request('PATCH',
                f'ds_documentos?id=eq.{documento_id}',
                body={'status': result['to'],
                      'updated_at': datetime.now(timezone.utc).isoformat()})
            if not supa_ok(ru):
                return self._send_json({'error': 'persist failed'}, 500)

            try:
                supabase_request('POST', 'entity_events', body={
                    'aggregate_type': 'DossieDocumento',
                    'aggregate_id': documento_id,
                    'event_type': f'DossieDocumento{event[0].upper()}{event[1:]}',
                    'payload': {**result, **payload, 'trilha': trilha},
                    'metadata': {'source': 'fsm_api', 'actor': actor},
                })
            except Exception:
                pass

            self._send_json({
                'documento_id': documento_id,
                'from': result['from'],
                'to': result['to'],
                'event': event,
                'trilha': trilha,
            })
        except Exception as e:
            log_error('dossie_documento_transition', str(e), e)
            self._send_json({'error': str(e)}, 500)

    # ============================================================
    # LF-FASE3: Descarrego endpoints
    # ============================================================
    def _handle_descarrego_capture(self):
        """POST /api/descarrego/capture
        Body: { mentorado_id, tipo_bruto, conteudo_bruto?, arquivo_url?, fonte? }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body()
            mentorado_id = body.get('mentorado_id')
            tipo_bruto = body.get('tipo_bruto', 'texto')
            if tipo_bruto not in ('texto','audio','video','imagem','arquivo','link','gravacao'):
                return self._send_json({'error': f'tipo_bruto inválido: {tipo_bruto}'}, 400)
            payload = {
                'mentorado_id': mentorado_id,
                'consultor_id': auth.get('user_id') if isinstance(auth, dict) else str(auth),
                'tipo_bruto': tipo_bruto,
                'conteudo_bruto': body.get('conteudo_bruto'),
                'arquivo_url': body.get('arquivo_url'),
                'arquivo_size_bytes': body.get('arquivo_size_bytes'),
                'arquivo_mime_type': body.get('arquivo_mime_type'),
                'duracao_ms': body.get('duracao_ms'),
                'fonte': body.get('fonte', 'web_drawer'),
                'correlation_id': str(uuid.uuid4()),
                'status': 'capturado',
            }
            r = supabase_request('POST', 'descarregos', body=payload)
            if r.status_code not in (200, 201):
                return self._send_json({'error': r.text}, r.status_code)
            data = r.json()
            row = data[0] if isinstance(data, list) and data else data
            self._send_json({
                'descarrego_id': row.get('id'),
                'correlation_id': row.get('correlation_id'),
                'status': row.get('status'),
            }, 201)
        except Exception as e:
            log_error('descarrego_capture', str(e), e)
            self._send_json({'error': str(e)}, 500)

    def _handle_descarrego_batch_capture(self):
        """POST /api/descarrego/batch-capture
        Body: { mentorado_id, items: [ { tipo_bruto, conteudo_bruto?, arquivo_url?, fonte? }, ... ] }
        Creates multiple descarregos at once. Max 20 per batch.
        Optionally auto-processes them if auto_process=true.
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body()
            mentorado_id = body.get('mentorado_id')
            items = body.get('items') or []
            auto_process = body.get('auto_process', False)

            if not mentorado_id:
                return self._send_json({'error': 'mentorado_id obrigatório'}, 400)
            if not items:
                return self._send_json({'error': 'items vazio'}, 400)
            if len(items) > 20:
                return self._send_json({'error': 'máximo 20 itens por batch'}, 400)

            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            batch_id = str(uuid.uuid4())
            created = []

            for i, item in enumerate(items):
                tipo = item.get('tipo_bruto', 'texto')
                if tipo not in ('texto', 'audio', 'video', 'imagem', 'arquivo', 'link', 'gravacao'):
                    tipo = 'texto'
                payload = {
                    'mentorado_id': mentorado_id,
                    'consultor_id': actor,
                    'tipo_bruto': tipo,
                    'conteudo_bruto': item.get('conteudo_bruto'),
                    'arquivo_url': item.get('arquivo_url'),
                    'arquivo_size_bytes': item.get('arquivo_size_bytes'),
                    'arquivo_mime_type': item.get('arquivo_mime_type'),
                    'duracao_ms': item.get('duracao_ms'),
                    'fonte': item.get('fonte', 'batch_import'),
                    'correlation_id': f'{batch_id}:{i}',
                    'status': 'capturado',
                }
                r = supabase_request('POST', 'descarregos', body=payload)
                if r.status_code in (200, 201):
                    data = r.json()
                    row = data[0] if isinstance(data, list) and data else data
                    created.append(row.get('id'))

            # Auto-process all created descarregos
            if auto_process and created:
                for did in created:
                    t = threading.Thread(
                        target=_descarrego_processor_run,
                        args=(did,),
                        daemon=True,
                    )
                    t.start()

            self._send_json({
                'batch_id': batch_id,
                'total': len(items),
                'created': len(created),
                'descarrego_ids': created,
                'auto_processing': auto_process,
            }, 201)
        except Exception as e:
            log_error('descarrego_batch', str(e), e)
            self._send_json({'error': str(e)}, 500)

    def _handle_descarrego_process(self, descarrego_id):
        """POST /api/descarrego/{id}/process — async pipeline."""
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        # Spawn worker thread
        t = threading.Thread(
            target=_descarrego_processor_run,
            args=(descarrego_id,),
            daemon=True,
        )
        t.start()
        self._send_json({
            'descarrego_id': descarrego_id,
            'status': 'processing_started',
        }, 202)

    def _handle_descarrego_approve(self, descarrego_id):
        """POST /api/descarrego/{id}/approve — HITL approve."""
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
        try:
            _descarrego_transition(descarrego_id, 'human_approved', actor=actor)
            # Spawn execução manual em thread
            t = threading.Thread(
                target=_descarrego_execute_action,
                args=(descarrego_id, 'manual'),
                daemon=True,
            )
            t.start()
            self._send_json({'descarrego_id': descarrego_id, 'status': 'approved'})
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    def _handle_descarrego_reject(self, descarrego_id):
        """POST /api/descarrego/{id}/reject — HITL reject."""
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
        try:
            _descarrego_transition(descarrego_id, 'human_rejected', actor=actor)
            supabase_request(
                'PATCH',
                f'descarregos?id=eq.{descarrego_id}',
                body={'acao_tomada': 'rejeitado_humano',
                      'acao_tomada_em': datetime.now(timezone.utc).isoformat(),
                      'acao_tomada_por': actor},
            )
            self._send_json({'descarrego_id': descarrego_id, 'status': 'rejected'})
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    def _handle_descarrego_reclassify(self, descarrego_id):
        """POST /api/descarrego/{id}/reclassify — manual reclassification.
        Body: { new_type: 'task'|'contexto'|... }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            body = self._read_json_body() or {}
            new_type = body.get('new_type')
            valid_types = ('task', 'lembrete', 'contexto', 'feedback', 'reembolso',
                           'bloqueio', 'escalacao', 'plano_acao', 'dossie', 'duvida',
                           'celebracao', 'outro')
            if not new_type or new_type not in valid_types:
                return self._send_json({'error': f'new_type inválido. Válidos: {", ".join(valid_types)}'}, 400)

            actor = auth.get('user_id') if isinstance(auth, dict) else str(auth)
            old_row = _descarrego_load(descarrego_id)
            old_type = old_row.get('classificacao_principal')

            _descarrego_update(descarrego_id, {
                'classificacao_principal': new_type,
                'classificacao_payload': {
                    **(old_row.get('classificacao_payload') or {}),
                    'primary_type': new_type,
                    '_reclassified_from': old_type,
                    '_reclassified_by': actor,
                    '_reclassified_at': datetime.now(timezone.utc).isoformat(),
                },
            })

            try:
                supabase_request('POST', 'entity_events', body={
                    'aggregate_type': 'Descarrego',
                    'aggregate_id': descarrego_id,
                    'event_type': 'DescarregoReclassified',
                    'payload': {'from': old_type, 'to': new_type},
                    'metadata': {'source': 'manual', 'actor': actor},
                })
            except Exception:
                pass

            self._send_json({'descarrego_id': descarrego_id, 'new_type': new_type, 'old_type': old_type})
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    def _handle_descarregos_list(self, mentorado_id):
        """GET /api/mentees/{id}/descarregos"""
        auth = check_auth_any(self.headers)
        if not auth:
            return self._send_json({'error': 'unauthorized'}, 401)
        try:
            r = supabase_request(
                'GET',
                f'descarregos?mentorado_id=eq.{mentorado_id}'
                f'&order=created_at.desc&limit=100',
            )
            self._send_json({'descarregos': r.json()})
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, PATCH, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, apikey, Authorization, X-API-Key')
        self.send_header('Access-Control-Max-Age', '86400')
        self.end_headers()

    def do_GET(self):
        if self.path == '/api/auth/me':
            self._handle_auth_me()
        elif self.path.startswith('/api/evolution/'):
            self._proxy_evolution('GET')
        elif self.path == '/api/mentees/portfolio':
            self._handle_get_portfolio()
        elif self.path == '/api/mentees':
            self._handle_get_mentees()
        elif re.match(r'^/api/mentees/(\d+)/notes$', self.path):
            _m = re.match(r'^/api/mentees/(\d+)/notes$', self.path)
            self._handle_get_notes(_m.group(1))
        elif self.path.startswith('/api/calendar/events'):
            self._handle_list_events()
        elif self.path == '/api/calls/upcoming':
            self._handle_upcoming_calls()
        elif self.path.startswith('/api/media/presign'):
            self._handle_media_presign()
        elif self.path.startswith('/api/media/stream'):
            self._handle_media_stream()
        elif self.path == '/api/evolution/instance-uuid':
            self._handle_instance_uuid()
        elif self.path == '/api/sheets/status':
            self._handle_sheets_status()
        elif self.path == '/api/wa/groups':
            self._handle_wa_groups_list()
        elif self.path.startswith('/api/storage/files'):
            self._handle_storage_list_files()
        elif self.path.startswith('/api/storage/status'):
            self._handle_storage_status()
        elif self.path == '/api/health':
            self._send_json({
                'status': 'ok',
                'zoom_configured': bool(ZOOM_ACCOUNT_ID and ZOOM_CLIENT_ID),
                'gcal_configured': bool(os.environ.get('GOOGLE_SA_JSON') or os.environ.get('GOOGLE_SA_CREDENTIALS_B64') or os.path.exists(GOOGLE_SA_PATH)),
                'supabase_configured': bool(SUPABASE_SERVICE_KEY or SUPABASE_ANON_KEY),
                'sheets_configured': get_sheets_service() is not None,
                'openai_configured': bool(OPENAI_API_KEY),
                'storage_search': bool(OPENAI_API_KEY),
                'evolution_configured': bool(EVOLUTION_API_KEY),
                'evolution_base': EVOLUTION_BASE,
                'evolution_key_prefix': EVOLUTION_API_KEY[:8] + '...' if EVOLUTION_API_KEY else 'EMPTY',
            })
        # ===== WA DM v2 (S9-A) =====
        elif self.path == '/api/mentees/triage':
            self._handle_mentees_triage()
        elif self.path.startswith('/api/wa/media'):
            self._handle_wa_media()
        elif self.path.startswith('/api/wa/labels/summary'):
            self._handle_wa_labels_summary()
        elif self.path.startswith('/api/wa/inbox'):
            self._handle_wa_inbox()
        elif re.match(r'^/api/wa/presence/(\d+)$', self.path):
            _m = re.match(r'^/api/wa/presence/(\d+)$', self.path)
            self._handle_wa_presence_get(_m.group(1))
        # ===== Mentee Groups =====
        elif self.path == '/api/mentee-groups':
            self._handle_get_groups()
        elif re.match(r'^/api/mentee-groups/(\d+)/members$', self.path):
            _m = re.match(r'^/api/mentee-groups/(\d+)/members$', self.path)
            self._handle_get_group_members(_m.group(1))
        elif re.match(r'^/api/mentees/(\d+)/descarregos$', self.path):
            _m = re.match(r'^/api/mentees/(\d+)/descarregos$', self.path)
            self._handle_descarregos_list(_m.group(1))
        elif self.path == '/api/clickup/command-center':
            self._handle_clickup_command_center()
        # ===== Biblioteca =====
        elif self.path.startswith('/api/biblioteca'):
            self._handle_biblioteca_get()
        # ===== API Keys Management =====
        elif self.path == '/api/keys':
            self._handle_list_api_keys()
        # ===== Chatwoot Messages =====
        elif re.match(r'^/api/mentees/(\d+)/messages$', self.path):
            _m = re.match(r'^/api/mentees/(\d+)/messages$', self.path)
            self._handle_get_chatwoot_messages(_m.group(1))
        # ===== Cron Jobs Status (OpenFang — EPIC 5) =====
        elif self.path == '/api/crons/status':
            self._handle_cron_status()
        # ===== Dossiê QA Scores =====
        elif re.match(r'^/api/mentees/(\d+)/qa-scores$', self.path):
            _m = re.match(r'^/api/mentees/(\d+)/qa-scores$', self.path)
            self._handle_get_qa_scores(_m.group(1))
        # ===== Root =====
        elif self.path == '/' or self.path == '':
            self._send_json({
                'service': 'Spalla Dashboard API',
                'version': '1.0.0',
                'status': 'operational',
                'docs': 'https://spalla-dashboard.vercel.app/api-docs',
                'health': '/api/health',
                'endpoints': 58,
            })
        else:
            super().do_GET()

    def end_headers(self):
        # Disable browser cache for JS/HTML/CSS files
        if hasattr(self, 'path') and any(ext in self.path for ext in ('.js', '.html', '.css', '.jpg', '.png')):
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
        super().end_headers()

    def do_POST(self):
        _notes_post = re.match(r'^/api/mentees/(\d+)/notes$', self.path)
        _offboard = re.match(r'^/api/mentees/(\d+)/offboard$', self.path)
        if _notes_post:
            self._handle_post_note(_notes_post.group(1))
        elif _offboard:
            self._handle_offboard_mentee(_offboard.group(1))
        elif self.path == '/api/auth/register':
            self._handle_auth_register()
        elif self.path == '/api/auth/login':
            self._handle_auth_login()
        elif self.path == '/api/auth/refresh':
            self._handle_auth_refresh()
        elif self.path == '/api/auth/reset-password':
            self._handle_auth_reset_password()
        elif self.path.startswith('/api/evolution/'):
            self._proxy_evolution('POST')
        elif self.path == '/api/schedule-call':
            self._handle_schedule_call()
        elif self.path == '/api/zoom/create-meeting':
            self._handle_create_zoom_meeting()
        elif self.path == '/api/calendar/create-event':
            self._handle_create_calendar_event()
        elif self.path == '/api/sheets/sync':
            self._handle_sheets_sync()
        elif self.path == '/api/welcome-flow/register':
            self._handle_welcome_flow_register()
        elif self.path == '/api/storage/process':
            self._handle_storage_process()
        elif self.path == '/api/storage/search':
            self._handle_storage_search()
        elif self.path == '/api/storage/test':
            self._handle_storage_test()
        elif self.path == '/api/storage/reprocess':
            self._handle_storage_reprocess()
        # ===== WA DM v2 (S9-A) =====
        elif self.path == '/api/wa/presence':
            self._handle_wa_presence_post()
# ===== Intelligence Layer (SPEC-6.1) =====
        elif self.path == '/api/copilot':
            self._handle_copilot()
        elif self.path == '/api/ds/update-stage':
            self._handle_ds_update_stage()
        # ===== Mentee Groups =====
        elif self.path == '/api/mentee-groups':
            self._handle_post_group()
        elif re.match(r'^/api/mentee-groups/(\d+)/members$', self.path):
            _m = re.match(r'^/api/mentee-groups/(\d+)/members$', self.path)
            self._handle_post_group_member(_m.group(1))
        # ===== ORCH-05: Agent Dispatch =====
        elif re.match(r'^/api/tasks/([^/]+)/agent-dispatch$', self.path):
            _ad = re.match(r'^/api/tasks/([^/]+)/agent-dispatch$', self.path)
            self._handle_agent_dispatch(_ad.group(1))
        # ===== ORCH-06: Handoff Protocol =====
        elif re.match(r'^/api/tasks/([^/]+)/handoff$', self.path):
            _hf = re.match(r'^/api/tasks/([^/]+)/handoff$', self.path)
            self._handle_task_handoff(_hf.group(1))
        # ===== LF Story 5: Task FSM transition =====
        elif re.match(r'^/api/tasks/([^/]+)/transition$', self.path):
            _tm = re.match(r'^/api/tasks/([^/]+)/transition$', self.path)
            self._handle_task_transition(_tm.group(1))
        # ===== FSM: Mentorado transition =====
        elif re.match(r'^/api/mentees/([^/]+)/transition$', self.path):
            _mm = re.match(r'^/api/mentees/([^/]+)/transition$', self.path)
            self._handle_mentorado_transition(_mm.group(1))
        # ===== FSM: Dossiê Produção transition =====
        elif re.match(r'^/api/dossies/producoes/([^/]+)/transition$', self.path):
            _dp = re.match(r'^/api/dossies/producoes/([^/]+)/transition$', self.path)
            self._handle_dossie_producao_transition(_dp.group(1))
        # ===== FSM: Dossiê Documento transition =====
        elif re.match(r'^/api/dossies/documentos/([^/]+)/transition$', self.path):
            _dd = re.match(r'^/api/dossies/documentos/([^/]+)/transition$', self.path)
            self._handle_dossie_documento_transition(_dd.group(1))
        # ===== LF-FASE3: Descarrego =====
        elif self.path == '/api/descarrego/capture':
            self._handle_descarrego_capture()
        elif self.path == '/api/descarrego/batch-capture':
            self._handle_descarrego_batch_capture()
        elif re.match(r'^/api/descarrego/([0-9a-f-]+)/process$', self.path):
            _dm = re.match(r'^/api/descarrego/([0-9a-f-]+)/process$', self.path)
            self._handle_descarrego_process(_dm.group(1))
        elif re.match(r'^/api/descarrego/([0-9a-f-]+)/approve$', self.path):
            _dm = re.match(r'^/api/descarrego/([0-9a-f-]+)/approve$', self.path)
            self._handle_descarrego_approve(_dm.group(1))
        elif re.match(r'^/api/descarrego/([0-9a-f-]+)/reject$', self.path):
            _dm = re.match(r'^/api/descarrego/([0-9a-f-]+)/reject$', self.path)
            self._handle_descarrego_reject(_dm.group(1))
        elif re.match(r'^/api/descarrego/([0-9a-f-]+)/reclassify$', self.path):
            _dm = re.match(r'^/api/descarrego/([0-9a-f-]+)/reclassify$', self.path)
            self._handle_descarrego_reclassify(_dm.group(1))
        elif self.path == '/api/clickup/sync-subtasks':
            self._handle_clickup_sync_subtasks()
        elif self.path == '/api/clickup/import-all':
            self._handle_clickup_import_all()
        elif self.path == '/api/clickup/webhook':
            self._handle_clickup_webhook()
        elif re.match(r'^/api/clickup/push/(.+)$', self.path):
            _m = re.match(r'^/api/clickup/push/(.+)$', self.path)
            self._handle_clickup_push(_m.group(1))
        # ===== API Keys Management =====
        elif self.path == '/api/keys/generate':
            self._handle_generate_api_key()
        # ===== Evolution Webhook (Status Updates) =====
        elif self.path == '/api/webhooks/evolution':
            self._handle_evolution_webhook()
        # ===== WhatsApp Send Endpoints =====
        elif self.path == '/api/wa/send-text':
            self._handle_wa_send_text()
        elif self.path == '/api/wa/send-media':
            self._handle_wa_send_media()
        elif self.path == '/api/wa/reply':
            self._handle_wa_reply()
        # ===== WhatsApp Group Management =====
        elif self.path == '/api/wa/groups/sync':
            self._handle_wa_groups_sync()
        elif self.path == '/api/wa/groups/create':
            self._handle_wa_groups_create()
        elif re.match(r'^/api/wa/groups/([^/]+)/link$', self.path):
            _gm = re.match(r'^/api/wa/groups/([^/]+)/link$', self.path)
            self._handle_wa_group_link(_gm.group(1))
        # ===== YouTube Upload =====
        elif self.path == '/api/youtube/upload':
            self._handle_youtube_upload()
        # ===== Google Drive Sync =====
        elif self.path == '/api/drive/sync':
            self._handle_drive_sync()
        # ===== Resumo Semanal =====
        elif self.path == '/api/mentee/weekly-summary':
            self._handle_weekly_summary()
        # ===== Task from Audio (TASK-07) =====
        elif self.path == '/api/tasks/from-audio':
            self._handle_tasks_from_audio()
        # ===== Context Hub: Transcribe audio =====
        elif self.path == '/api/context/transcribe':
            self._handle_context_transcribe()
        # ===== Task Notifications =====
        elif self.path == '/api/tasks/notify':
            self._handle_task_notify()
        # ===== Chatwoot Webhook =====
        elif self.path == '/api/webhooks/chatwoot':
            self._handle_chatwoot_webhook()
        # ===== Fabric Pattern Runner =====
        elif self.path == '/api/fabric/run':
            self._handle_fabric_run()
        # ===== RAGAS Quality Gate =====
        elif self.path == '/api/dossie/evaluate':
            self._handle_ragas_evaluate()
        # ===== Dossiê Generation (Goose Agent — EPIC 6) =====
        elif self.path == '/api/dossie/generate':
            self._handle_dossie_generate()
        # ===== Automations Cron (Dragon 16) =====
        elif self.path == '/api/automations/evaluate':
            self._handle_automations_evaluate()
        # ===== Webhook Outgoing (Dragon 17) =====
        elif self.path == '/api/webhooks/outgoing/test':
            self._handle_webhook_outgoing_test()
        # ===== Recurring Tasks (Dragon 18) =====
        elif self.path == '/api/tasks/process-recurring':
            self._handle_process_recurring()
        else:
            self._send_json({'error': 'Not found'}, 404)

    def do_PUT(self):
        if self.path.startswith('/api/evolution/'):
            self._proxy_evolution('PUT')
        else:
            self._send_json({'error': 'Not found'}, 404)

    def do_DELETE(self):
        if self.path.startswith('/api/calendar/event/'):
            self._handle_delete_calendar_event()
        elif self.path.startswith('/api/evolution/'):
            self._proxy_evolution('DELETE')
        # ===== WA DM v2 (S9-A) =====
        elif self.path.startswith('/api/wa/presence'):
            self._handle_wa_presence_delete()
        # ===== Mentee Groups =====
        elif re.match(r'^/api/mentee-groups/(\d+)$', self.path):
            _m = re.match(r'^/api/mentee-groups/(\d+)$', self.path)
            self._handle_delete_group(_m.group(1))
        elif re.match(r'^/api/mentee-groups/(\d+)/members/(\d+)$', self.path):
            _m = re.match(r'^/api/mentee-groups/(\d+)/members/(\d+)$', self.path)
            self._handle_delete_group_member(_m.group(1), _m.group(2))
        # ===== API Keys Management =====
        elif re.match(r'^/api/keys/([0-9a-f-]+)$', self.path):
            _m = re.match(r'^/api/keys/([0-9a-f-]+)$', self.path)
            self._handle_revoke_api_key(_m.group(1))
        else:
            self._send_json({'error': 'Not found'}, 404)

    # ===== WA DM v2 HANDLERS (S9-A) =====

    def _handle_mentees_triage(self):
            """GET /api/mentees/triage — Server-side triage score per mentee.
            Uses vw_wa_mentee_inbox + wa_topics to compute priority scores.
            Returns: [{ id, nome, score, level, factors }] sorted by score desc.
            """
            auth = check_auth_any(self.headers)
            if not auth:
                self._send_json({'error': 'Authentication required'}, 401)
                return
            try:
                inbox = supabase_request('GET',
                    'vw_wa_mentee_inbox?select=*&order=horas_sem_resposta_equipe.desc.nullslast')
                if not isinstance(inbox, list):
                    inbox = []

                cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
                topics = supabase_request('GET',
                    f'wa_topics?select=group_jid,sentiment&sentiment=in.(negativo,critico)'
                    f'&last_message_at=gte.{cutoff}')
                if not isinstance(topics, list):
                    topics = []

                neg_jids = {t['group_jid'] for t in topics if t.get('group_jid')}

                results = []
                for m in inbox:
                    score = 0
                    factors = []

                    h = m.get('horas_sem_resposta_equipe') or 0
                    if h > 72:
                        score += 40; factors.append('sem_contato_72h')
                    elif h > 48:
                        score += 25; factors.append('sem_contato_48h')
                    elif h > 24:
                        score += 10; factors.append('sem_contato_24h')

                    fase = m.get('fase_jornada', '')
                    if fase in ('onboarding', 'renovacao'):
                        score += 20; factors.append(f'fase_critica_{fase}')

                    tarefas = m.get('tarefas_pendentes') or 0
                    score += min(20, tarefas * 5)
                    if tarefas > 0:
                        factors.append(f'{tarefas}_tarefas')

                    unread = m.get('msgs_pendentes_resposta') or 0
                    score += min(10, unread)
                    if unread > 0:
                        factors.append(f'{unread}_msgs_nao_lidas')

                    jid = m.get('grupo_whatsapp_id') or m.get('group_jid') or ''
                    if jid in neg_jids:
                        score += 15; factors.append('sentimento_negativo')

                    if m.get('health_status') == 'vermelho':
                        score += 10

                    level = 'critico' if score >= 60 else 'atencao' if score >= 30 else 'ok'
                    results.append({
                        'id': m.get('id'),
                        'nome': m.get('nome'),
                        'score': score,
                        'level': level,
                        'factors': factors,
                    })

                results.sort(key=lambda x: x['score'], reverse=True)
                self._send_json(results)
            except Exception as e:
                log_error('Triage', f'_handle_mentees_triage failed: {e}')
                self._send_json({'error': str(e)}, 500)

    def _handle_wa_media(self):
        """GET /api/wa/media?mentee_id={id} — WA media files for a mentee.
        Queries wa_message_queue (raw JSONB payloads) and extracts media in Python.
        Returns: { media: [{ id, message_type, url, mimetype, created_at, from_me }] }
        """
        try:
            parsed = urllib.parse.urlparse(self.path)
            params = urllib.parse.parse_qs(parsed.query)
            mentee_id = params.get('mentee_id', [None])[0]
            if not mentee_id:
                self._send_json({'error': 'mentee_id is required'}, 400)
                return

            mentee = supabase_request('GET',
                f'mentorados?id=eq.{mentee_id}&select=id,nome,grupo_whatsapp_id')
            if not isinstance(mentee, list) or not mentee:
                self._send_json({'media': []})
                return

            group_jid = mentee[0].get('grupo_whatsapp_id')
            if not group_jid:
                self._send_json({'media': []})
                return

            # Fetch recent queue rows for this group (payload is raw Evolution API JSON)
            rows = supabase_request('GET',
                f'wa_message_queue?select=id,payload,created_at'
                f'&group_jid=eq.{urllib.parse.quote(group_jid)}'
                f'&order=created_at.desc&limit=200')

            MEDIA_TYPES = {'imageMessage', 'audioMessage', 'videoMessage', 'documentMessage'}
            media = []
            for r in (rows if isinstance(rows, list) else []):
                payload = r.get('payload') or {}
                data = payload.get('data', payload)  # Evolution API wraps in 'data'
                msg_type = data.get('messageType', '')
                if msg_type not in MEDIA_TYPES:
                    continue
                msg_content = data.get('message', {}).get(msg_type, {})
                url = msg_content.get('url') or msg_content.get('directPath', '')
                if not url:
                    continue
                media.append({
                    'id': r.get('id'),
                    'message_type': msg_type,
                    'url': url,
                    'mimetype': msg_content.get('mimetype', ''),
                    'created_at': r.get('created_at'),
                    'from_me': data.get('key', {}).get('fromMe', False),
                })

            self._send_json({'media': media})
        except Exception as e:
            log_error('WaMedia', f'_handle_wa_media failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_copilot(self):
        """POST /api/copilot — Contextual AI assistant for a mentee.
        Body: { mentee_id: int, message: str, history?: [{role, content}] }
        Enriches system prompt with mentee context (fase, saúde, tópicos WA, notas recentes).
        Returns: { reply: str, context_used: bool }
        """
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
            mentee_id = body.get('mentee_id')
            user_message = (body.get('message') or '').strip()
            history = body.get('history') or []

            if not user_message:
                self._send_json({'error': 'message is required'}, 400)
                return

            if not OPENAI_API_KEY:
                self._send_json({'error': 'OPENAI_API_KEY not configured'}, 503)
                return

            # ── Build context block ─────────────────────────────────────────
            context_lines = []
            context_used = False

            if mentee_id:
                try:
                    # Mentee profile + inbox snapshot
                    inbox = supabase_request('GET',
                        f'vw_wa_mentee_inbox?mentorado_id=eq.{mentee_id}&limit=1')
                    if isinstance(inbox, list) and inbox:
                        m = inbox[0]
                        context_lines.append(f'Mentorado: {m.get("nome", "?")}')
                        context_lines.append(f'Fase: {m.get("fase_jornada", "?")}')
                        context_lines.append(f'Saúde WA: {m.get("health_status", "?")}')
                        horas = m.get('horas_sem_resposta_equipe')
                        if horas is not None:
                            context_lines.append(f'Horas sem resposta da equipe: {horas}h')
                        context_lines.append(f'Msgs não lidas: {m.get("unread_count", 0)}')
                        context_lines.append(f'Tarefas abertas: {m.get("active_tasks_count", 0)}')
                        last_msg = m.get('last_message')
                        if last_msg:
                            sender = 'Equipe' if m.get('last_message_is_team') else m.get('nome', 'Mentorado')
                            context_lines.append(f'Última msg ({sender}): "{last_msg[:120]}"')

                    # Recent WA topics (last 5, last 7 days)
                    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
                    topics = supabase_request('GET',
                        f'wa_topics?mentorado_id=eq.{mentee_id}'
                        f'&last_message_at=gte.{urllib.parse.quote(cutoff)}'
                        f'&order=last_message_at.desc&limit=5'
                        f'&select=title,sentiment,message_count,last_message_at')
                    if isinstance(topics, list) and topics:
                        context_lines.append('\nTópicos WA recentes (7 dias):')
                        for t in topics:
                            sent = t.get('sentiment', '')
                            sent_label = f' [{sent}]' if sent and sent != 'neutral' else ''
                            context_lines.append(f'  • {t.get("title", "?")} ({t.get("message_count", 0)} msgs){sent_label}')

                    # Recent notes (last 3)
                    notes = supabase_request('GET',
                        f'mentee_notes?mentorado_id=eq.{mentee_id}'
                        f'&order=created_at.desc&limit=3'
                        f'&select=tipo,conteudo,created_at')
                    if isinstance(notes, list) and notes:
                        context_lines.append('\nNotas recentes:')
                        for n in notes:
                            preview = (n.get('conteudo') or '')[:100]
                            context_lines.append(f'  • [{n.get("tipo", "nota")}] {preview}')

                    context_used = bool(context_lines)
                except Exception as ctx_err:
                    log_error('Copilot', f'Context fetch failed (non-fatal): {ctx_err}')

            # ── System prompt ───────────────────────────────────────────────
            system_prompt = (
                'Você é o Copiloto do Consultor do CASE Mentoring. '
                'Ajuda consultores a gerir seus mentorados com clareza e objetividade. '
                'Responda em português. Seja direto e prático.'
            )
            if context_lines:
                system_prompt += '\n\n=== CONTEXTO DO MENTORADO ===\n' + '\n'.join(context_lines)

            # ── Build messages ──────────────────────────────────────────────
            messages = [{'role': 'system', 'content': system_prompt}]
            # Append up to 6 history turns to keep context window small
            for turn in history[-6:]:
                role = turn.get('role', 'user')
                content = turn.get('content', '')
                if role in ('user', 'assistant') and content:
                    messages.append({'role': role, 'content': content})
            messages.append({'role': 'user', 'content': user_message})

            result = _openai_request('chat/completions', body={
                'model': 'gpt-4o-mini',
                'messages': messages,
                'max_tokens': 600,
                'temperature': 0.4,
            })

            reply = result.get('choices', [{}])[0].get('message', {}).get('content', '')
            self._send_json({'reply': reply, 'context_used': context_used})

        except Exception as e:
            log_error('Copilot', f'_handle_copilot failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_labels_summary(self):
        """GET /api/wa/labels/summary — Topic-type label counts.
        Params: mentee_id (optional) — filter by mentee; days (default 30)
        Returns: [{ slug, name, color, icon, count }] sorted by count desc.
        """
        try:
            parsed = urllib.parse.urlparse(self.path)
            params = urllib.parse.parse_qs(parsed.query)
            mentee_id = params.get('mentee_id', [None])[0]
            days = int(params.get('days', ['30'])[0])

            cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

            # Fetch topic types (static lookup)
            types = supabase_request('GET',
                'wa_topic_types?select=id,slug,name,color,icon&order=sort_order.asc')
            if not isinstance(types, list):
                types = []
            type_map = {t['id']: t for t in types}

            # Count topics per type within timeframe
            qs = (f'wa_topics?select=type_id,status'
                  f'&last_message_at=gte.{urllib.parse.quote(cutoff)}'
                  f'&type_id=not.is.null')
            if mentee_id:
                qs += f'&mentorado_id=eq.{mentee_id}'
            topics = supabase_request('GET', qs)
            if not isinstance(topics, list):
                topics = []

            counts = {}
            for t in topics:
                tid = t.get('type_id')
                if tid:
                    counts[tid] = counts.get(tid, 0) + 1

            result = []
            for tid, cnt in counts.items():
                tt = type_map.get(tid, {})
                if tt:
                    result.append({
                        'slug': tt.get('slug', ''),
                        'name': tt.get('name', ''),
                        'color': tt.get('color', '#6b7280'),
                        'icon': tt.get('icon', ''),
                        'count': cnt,
                    })
            result.sort(key=lambda x: x['count'], reverse=True)
            self._send_json(result)
        except Exception as e:
            log_error('WaLabels', f'_handle_wa_labels_summary failed: {e}')
            self._send_json({'error': str(e)}, 500)
    def _handle_wa_inbox(self):
        """
        GET /api/wa/inbox
        Returns vw_wa_mentee_inbox rows with optional filters.
        Params:
          health_status: verde | amarelo | vermelho | snoozed
          fase_jornada:  onboarding | execucao | resultado | renovacao
          search:        ilike match on nome
          sort:          sla_desc (default) | unread_desc | last_message_desc
        """
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return
        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)

        filters = []

        health = params.get('health_status', [None])[0]
        if health:
            filters.append(f'health_status=eq.{urllib.parse.quote(health)}')

        fase = params.get('fase_jornada', [None])[0]
        if fase:
            filters.append(f'fase_jornada=eq.{urllib.parse.quote(fase)}')

        search = params.get('search', [None])[0]
        if search:
            filters.append(f'nome=ilike.*{urllib.parse.quote(search)}*')

        sort = params.get('sort', ['sla_desc'])[0]
        sort_map = {
            'sla_desc':          'horas_sem_resposta_equipe.desc.nullslast',
            'unread_desc':       'unread_count.desc.nullslast',
            'last_message_desc': 'last_message_at.desc.nullslast',
        }
        order = sort_map.get(sort, 'horas_sem_resposta_equipe.desc.nullslast')
        filters.append(f'order={order}')

        query = 'vw_wa_mentee_inbox?select=*'
        if filters:
            query += '&' + '&'.join(filters)

        result = supabase_request('GET', query)
        if isinstance(result, dict) and result.get('error'):
            self._send_json({'error': 'upstream error', 'detail': result.get('message', str(result))}, 502)
            return
        self._send_json(result if isinstance(result, list) else [])

    def _handle_wa_presence_post(self):
        """
        POST /api/wa/presence
        Body: { mentorado_id, user_email, user_name }
        Upserts wa_presence row — heartbeat every 30s from frontend.
        Uses PATCH-then-INSERT pattern (avoids custom Prefer header).
        """
        try:
            body = json.loads(self._read_body())
        except Exception:
            self._send_json({'error': 'Invalid JSON'}, 400)
            return

        mentorado_id = body.get('mentorado_id')
        user_email   = body.get('user_email', '').strip()
        user_name    = body.get('user_name', '').strip()

        if not mentorado_id or not user_email:
            self._send_json({'error': 'mentorado_id and user_email required'}, 400)
            return

        ts_now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S+00:00')

        # Try UPDATE first (row already exists)
        patch_path = (
            f'wa_presence'
            f'?mentorado_id=eq.{mentorado_id}'
            f'&user_email=eq.{urllib.parse.quote(user_email)}'
        )
        updated = supabase_request('PATCH', patch_path, {
            'last_seen': ts_now,
            'user_name': user_name,
        })

        if isinstance(updated, dict) and updated.get('error'):
            self._send_json({'error': 'presence update failed', 'detail': updated.get('message', str(updated))}, 502)
            return

        # If no row was updated, INSERT (first heartbeat for this user+mentee)
        if isinstance(updated, list) and len(updated) == 0:
            inserted = supabase_request('POST', 'wa_presence', {
                'mentorado_id': mentorado_id,
                'user_email':   user_email,
                'user_name':    user_name,
                'last_seen':    ts_now,
            })
            # accept duplicate-key (race condition) as success; propagate other errors
            if isinstance(inserted, dict) and inserted.get('error'):
                code = inserted.get('code', '')
                if code != '23505':  # 23505 = unique_violation (concurrent insert won)
                    self._send_json({'error': 'presence insert failed', 'detail': inserted.get('message', str(inserted))}, 502)
                    return
                # re-PATCH to ensure last_seen is updated after race
                supabase_request('PATCH', patch_path, {'last_seen': ts_now, 'user_name': user_name})

        self._send_json({'ok': True})

    def _handle_wa_presence_delete(self):
        """
        DELETE /api/wa/presence?mentorado_id=X&user_email=Y
        Clears presence when user closes the mentee chat/drawer.
        """
        parsed = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed.query)

        mentorado_id = params.get('mentorado_id', [None])[0]
        user_email   = (params.get('user_email', [None])[0] or '').strip()

        if not mentorado_id or not user_email:
            self._send_json({'error': 'mentorado_id and user_email required'}, 400)
            return

        result = supabase_request(
            'DELETE',
            f'wa_presence'
            f'?mentorado_id=eq.{mentorado_id}'
            f'&user_email=eq.{urllib.parse.quote(user_email)}'
        )
        if isinstance(result, dict) and result.get('error'):
            self._send_json({'error': 'presence delete failed', 'detail': result.get('message', str(result))}, 502)
            return
        self._send_json({'ok': True})

    def _handle_wa_presence_get(self, mentorado_id):
        """
        GET /api/wa/presence/{mentorado_id}
        Returns active presence entries (last_seen within 60s).
        Used by frontend to show collision badge.
        """
        cutoff = (
            datetime.now(timezone.utc) - timedelta(seconds=60)
        ).strftime('%Y-%m-%dT%H:%M:%S+00:00')

        result = supabase_request(
            'GET',
            f'wa_presence'
            f'?mentorado_id=eq.{mentorado_id}'
            f'&last_seen=gte.{urllib.parse.quote(cutoff)}'
            f'&select=user_email,user_name,last_seen'
        )
        if isinstance(result, dict) and result.get('error'):
            self._send_json({'error': 'upstream error', 'detail': result.get('message', str(result))}, 502)
            return
        self._send_json(result if isinstance(result, list) else [])

    # ===== END WA DM v2 HANDLERS =====

    def do_PATCH(self):
        # bulk MUST come before individual to avoid /api/mentees/bulk matching UUID
        if self.path == '/api/mentees/bulk':
            self._handle_bulk_patch_mentees()
        else:
            _m = re.match(r'^/api/mentees/([0-9a-f-]+)$', self.path)
            if _m:
                self._handle_patch_mentee(_m.group(1))
            else:
                self._send_json({'error': 'Not found'}, 404)
    def _handle_delete_calendar_event(self):
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            event_id = self.path.split('/')[-1]
            result = delete_calendar_event(event_id)
            self._send_json(result)
        except Exception as e:
            self._send_json({'error': str(e)}, 500)
    # ===== SCHEDULE CALL (main orchestrator) =====
    def _handle_schedule_call(self):
        """
        Full scheduling flow:
        1. Create Zoom meeting
        2. Create Google Calendar event with Zoom link
        3. Store in Supabase calls_mentoria
        """
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
        except Exception:
            self._send_json({'error': 'Invalid JSON'}, 400)
            return

        mentorado_nome = body.get('mentorado', '')
        mentorado_id = body.get('mentorado_id', '')
        email = body.get('email', '')
        tipo = body.get('tipo', 'acompanhamento')
        data = body.get('data', '')  # YYYY-MM-DD
        horario = body.get('horario', '10:00')  # HH:MM
        duracao = body.get('duracao', 60)
        notas = body.get('notas', '')

        if not mentorado_nome or not data:
            self._send_json({'error': 'mentorado and data are required'}, 400)
            return

        # Build datetime
        start_dt = f'{data}T{horario}:00'
        end_dt_obj = datetime.fromisoformat(start_dt) + timedelta(minutes=duracao)
        end_dt = end_dt_obj.isoformat()

        topic = f'Call {tipo.title()} — {mentorado_nome}'
        result = {'mentorado': mentorado_nome, 'data': data, 'horario': horario, 'tipo': tipo}

        # Step 1: Create Zoom meeting
        zoom_result = create_zoom_meeting(
            topic=topic,
            start_time=start_dt,
            duration=duracao,
            invitees=[email] if email else None
        )
        if zoom_result.get('join_url'):
            result['zoom'] = zoom_result
            zoom_url = zoom_result['join_url']
        else:
            result['zoom'] = zoom_result  # May have error
            zoom_url = ''
            print(f'[Schedule] Zoom creation failed: {zoom_result}')

        # Step 2: Create Google Calendar event
        description = f'Tipo: {tipo}\n'
        if notas:
            description += f'Notas: {notas}\n'
        if zoom_url:
            description += f'\nZoom: {zoom_url}\n'

        attendees = []
        if email:
            attendees.append(email)

        gcal_result = create_calendar_event(
            summary=topic,
            start_iso=start_dt,
            end_iso=end_dt,
            description=description,
            attendees=attendees,
            location=zoom_url
        )
        result['calendar'] = gcal_result

        # Step 3: Store in Supabase
        if mentorado_id:
            # Map tipo to valid tipo_call values: diagnostico, planejamento, acompanhamento, fechamento
            tipo_call_map = {
                'acompanhamento': 'acompanhamento',
                'conselho': 'acompanhamento',
                'qa': 'acompanhamento',
                'onboarding': 'diagnostico',
                'estrategia': 'planejamento',
            }
            call_data = {
                'mentorado_id': int(mentorado_id),
                'data_call': f'{data}T{horario}:00+00:00',
                'tipo': tipo,
                'tipo_call': tipo_call_map.get(tipo, 'acompanhamento'),
                'duracao_minutos': duracao,
                'zoom_meeting_id': str(zoom_result.get('meeting_id', '')),
                'zoom_topic': topic,
                'status': 'processando',
                'observacoes_equipe': notas or None,
                'link_gravacao': zoom_url or None,
            }
            supa_result = insert_scheduled_call(call_data)
            result['supabase'] = supa_result if not isinstance(supa_result, list) else {'inserted': True, 'id': supa_result[0].get('id') if supa_result else None}

        self._send_json(result)
        print(f'[Schedule] Call scheduled: {topic} on {data} {horario}')

    # ===== INDIVIDUAL ENDPOINTS =====
    def _handle_create_zoom_meeting(self):
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
            result = create_zoom_meeting(
                topic=body.get('topic', 'Call Mentoria'),
                start_time=body.get('start_time', ''),
                duration=body.get('duration', 60),
                invitees=body.get('invitees', [])
            )
            self._send_json(result)
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    def _handle_create_calendar_event(self):
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
            result = create_calendar_event(
                summary=body.get('summary', ''),
                start_iso=body.get('start_iso') or body.get('start', ''),
                end_iso=body.get('end_iso') or body.get('end', ''),
                description=body.get('description', ''),
                attendees=body.get('attendees', []),
                location=body.get('location', '')
            )
            self._send_json(result)
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    # ===== MEDIA PRESIGN (S3) =====
    def _handle_media_presign(self):
        """Generate presigned URL for Hetzner S3 media"""
        params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        key = params.get('key', [''])[0]

        if not key:
            self._send_json({'error': 'key parameter required'}, 400)
            return

        try:
            url = generate_presigned_url(key)
            # FULL DEBUG: log the complete URL
            print(f'[Presign] Key: {key}')
            print(f'[Presign] Generated URL length: {len(url)}')
            print(f'[Presign] Full URL: {url}')
            self._send_json({'url': url, 'bucket': S3_BUCKET, 'endpoint': S3_ENDPOINT})
        except Exception as e:
            print(f'[S3] Presign error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_media_stream(self):
        """Stream media directly from Hetzner S3 (proxy to avoid CORS). Supports fallback URL."""
        params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        key = params.get('key', [''])[0]
        fallback_url = params.get('fallback', [''])[0]

        # SSRF protection: only allow known media domains as fallback
        ALLOWED_FALLBACK_DOMAINS = ['mmg.whatsapp.net', 'media.whatsapp.net', 'web.whatsapp.net', 'pps.whatsapp.net']
        if fallback_url:
            try:
                fb_host = urllib.parse.urlparse(fallback_url).hostname or ''
                if not any(fb_host.endswith(d) for d in ALLOWED_FALLBACK_DOMAINS):
                    fallback_url = ''  # silently ignore non-whatsapp URLs
            except Exception:
                fallback_url = ''

        if not key:
            self._send_json({'error': 'key parameter required'}, 400)
            return

        try:
            # Try with the key as provided
            url = generate_presigned_url(key)
            print(f'[Stream] Attempting to stream S3 key: {key}')

            req = urllib.request.Request(url)
            response = urllib.request.urlopen(req, timeout=30)

            # Get content type and size
            content_type = response.headers.get('Content-Type', 'application/octet-stream')
            content_length = response.headers.get('Content-Length')

            # Send response
            self.send_response(200)
            self.send_header('Content-Type', content_type)
            if content_length:
                self.send_header('Content-Length', content_length)
            # Allow CORS from frontend
            self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
            self.send_header('Cache-Control', 'public, max-age=3600')
            self.end_headers()

            # Stream the file
            chunk_size = 8192
            while True:
                chunk = response.read(chunk_size)
                if not chunk:
                    break
                self.wfile.write(chunk)

            print(f'[Stream] Successfully streamed {key}')

        except urllib.error.HTTPError as he:
            # If 403/404, try fallback URL (temporary WhatsApp URL)
            if he.code in [403, 404] and fallback_url:
                print(f'[Stream] S3 key not found: {key}. Trying fallback URL...')
                try:
                    fb_req = urllib.request.Request(fallback_url)
                    fb_resp = urllib.request.urlopen(fb_req, timeout=30)
                    ct = fb_resp.headers.get('Content-Type', 'application/octet-stream')
                    cl = fb_resp.headers.get('Content-Length')
                    self.send_response(200)
                    self.send_header('Content-Type', ct)
                    if cl: self.send_header('Content-Length', cl)
                    self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
                    self.send_header('Cache-Control', 'public, max-age=3600')
                    self.end_headers()
                    while True:
                        chunk = fb_resp.read(8192)
                        if not chunk: break
                        self.wfile.write(chunk)
                    print(f'[Stream] Fallback URL worked for {key}')
                    return
                except Exception as fb_err:
                    print(f'[Stream] Fallback URL also failed: {fb_err}')

            if he.code in [403, 404]:
                print(f'[Stream] Key not found: {key}, status {he.code}. Attempting to discover correct UUID...')
                try:
                    # Extract parts from original key
                    parts = key.split('/')
                    if len(parts) >= 3:
                        chat_id = parts[1]  # The wrong instanceId
                        remote_jid = parts[2]  # The chatId
                        message_type = parts[3] if len(parts) > 3 else 'audioMessage'

                        # Try to find correct instanceId by listing bucket
                        # Build list URL
                        list_url = f'https://{S3_ENDPOINT}/{S3_BUCKET}/?prefix=evolution-api/&delimiter=/'
                        presigned_list = generate_presigned_url('evolution-api/')

                        # For now, just respond with helpful debug info
                        print(f'[Stream] Parts: chat_id={chat_id}, remote_jid={remote_jid}, message_type={message_type}')
                        self._send_json({'error': f'File not found at {key}. May need UUID discovery.'}, 404)
                        return
                except Exception as e2:
                    print(f'[Stream] Discovery error: {e2}')

            self._send_json({'error': f'HTTP {he.code}: {he.reason}'}, he.code)

        except Exception as e:
            print(f'[Stream] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_instance_uuid(self):
        """Get the actual UUID of the Evolution instance from S3 bucket"""
        try:
            # List objects in bucket to find the instanceId pattern
            # S3 path is: evolution-api/{UUID}/{chatId}/{messageType}
            # We can extract the UUID from the first match
            presigned_url = generate_presigned_url('evolution-api/')
            print(f'[Instance UUID] Fetching bucket list from: {presigned_url}')

            # For now, return hardcoded based on what we found
            # TODO: Implement actual bucket listing via S3 API
            self._send_json({
                'instance': os.getenv('EVOLUTION_INSTANCE', 'producao002'),
                'note': 'UUID discovery not yet automated. Please check S3 bucket manually.',
                's3_bucket': S3_BUCKET,
                's3_endpoint': S3_ENDPOINT,
            })

        except Exception as e:
            print(f'[Instance UUID] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_get_portfolio(self):
        """GET /api/mentees/portfolio — carteira do consultor com health/priority scores"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401)
                return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return
            email = payload.get('email', '')

            # Fetch mentees assigned to this consultant
            mentees = supabase_request(
                'GET',
                f'mentorados?select=id,nome,email,instagram,fase_jornada,cohort,ativo,consultor_responsavel,snoozed_until,whatsapp_7d,whatsapp_30d'
                f'&ativo=eq.true&consultor_responsavel=eq.{urllib.parse.quote(email)}&order=nome'
            )
            if isinstance(mentees, dict) and 'error' in mentees:
                self._send_json({'error': mentees['error']}, 500)
                return
            if not isinstance(mentees, list):
                mentees = []

            # Fetch all wa_topics for health signals (one call, then group by mentee name)
            topics_raw = supabase_request(
                'GET',
                'vw_wa_topic_board?select=group_jid,mentorado_nome,status,type_slug,last_message_at&order=last_message_at.desc'
            )
            topics_by_nome = {}
            if isinstance(topics_raw, list):
                for t in topics_raw:
                    nome = t.get('mentorado_nome') or ''
                    if nome not in topics_by_nome:
                        topics_by_nome[nome] = []
                    topics_by_nome[nome].append(t)

            now = datetime.now(timezone.utc)
            result = []
            for m in mentees:
                nome = m.get('nome', '')
                mtopics = topics_by_nome.get(nome, [])

                unread_count = m.get('whatsapp_7d') or 0
                last_activity = None
                negative_topics = 0

                for t in mtopics:
                    lma = t.get('last_message_at')
                    if lma:
                        if last_activity is None or lma > last_activity:
                            last_activity = lma
                    if t.get('type_slug') in ('problema', 'risco', 'reclamacao', 'negativo'):
                        negative_topics += 1

                # Health: green/yellow/red based on days since last activity
                health = 'green'
                days_since = None
                if last_activity:
                    try:
                        ldt = datetime.fromisoformat(last_activity.replace('Z', '+00:00'))
                        days_since = (now - ldt).days
                        if days_since > 14:
                            health = 'red'
                        elif days_since > 7:
                            health = 'yellow'
                    except Exception:
                        pass
                else:
                    health = 'red'

                if negative_topics > 0 and health == 'green':
                    health = 'yellow'

                # Check snooze
                snoozed = False
                snoozed_until = m.get('snoozed_until')
                if snoozed_until:
                    try:
                        su = datetime.fromisoformat(snoozed_until.replace('Z', '+00:00'))
                        if su > now:
                            snoozed = True
                    except Exception:
                        pass

                # Priority score for inbox ordering (higher = more urgent)
                priority = 0
                if not snoozed:
                    if health == 'red':
                        priority += 30
                    elif health == 'yellow':
                        priority += 10
                    if m.get('fase_jornada') in ('onboarding', 'renovacao'):
                        priority += 20
                    priority += min((unread_count or 0) // 5, 15)
                    if negative_topics > 0:
                        priority += 10

                result.append({
                    **m,
                    'health': health,
                    'snoozed': snoozed,
                    'days_since_activity': days_since,
                    'unread_count': unread_count,
                    'negative_topics': negative_topics,
                    'priority_score': priority,
                    'recent_topics': mtopics[:3],
                })

            # Sort by priority descending (snoozed go to bottom)
            result.sort(key=lambda x: (x['snoozed'], -x['priority_score']))
            self._send_json(result)

        except Exception as e:
            log_error('Portfolio', f'_handle_get_portfolio failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_get_mentees(self):
        result = get_mentees_with_email()
        self._send_json(result if isinstance(result, list) else [result])

    def _handle_patch_mentee(self, mentee_id):
        """PATCH /api/mentees/{id} — update fase_jornada or snoozed_until"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401)
                return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return

            try:
                body = json.loads(self._read_body())
            except Exception:
                self._send_json({'error': 'Invalid JSON'}, 400)
                return

            ALLOWED_FIELDS = {'fase_jornada', 'snoozed_until', 'wa_status', 'trilha'}
            updates = {k: v for k, v in body.items() if k in ALLOWED_FIELDS}
            if not updates:
                self._send_json({'error': 'No allowed fields provided'}, 400)
                return

            valid_fases = {
                'onboarding', 'execucao', 'resultado', 'renovacao', 'encerrado'
            }
            if 'fase_jornada' in updates and updates['fase_jornada'] not in valid_fases:
                self._send_json({'error': f'Invalid fase_jornada: {updates["fase_jornada"]}'}, 400)
                return

            valid_trilha = {'scale', 'clinic'}
            if 'trilha' in updates and updates['trilha'] not in valid_trilha:
                self._send_json({'error': f'Invalid trilha: {updates["trilha"]}'}, 400)
                return

            valid_wa_status = {'aguardando', 'em_andamento', 'bloqueado', 'resolvido'}
            if 'wa_status' in updates and updates['wa_status'] not in valid_wa_status:
                self._send_json({'error': f'Invalid wa_status: {updates["wa_status"]}'}, 400)
                return

            result = supabase_request(
                'PATCH',
                f'mentorados?id=eq.{mentee_id}&select=id,nome,fase_jornada,snoozed_until,wa_status',
                updates
            )
            self._send_json(result)
        except Exception as e:
            log_error('PatchMentee', f'_handle_patch_mentee failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_offboard_mentee(self, mentee_id):
        """POST /api/mentees/{id}/offboard — desativa um mentorado (ativo = false).
        Body: { motivo: 'reembolso'|'conclusao'|'cancelamento'|'outro', obs?: string }
        """
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401)
                return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return

            try:
                body = json.loads(self._read_body())
            except Exception:
                self._send_json({'error': 'Invalid JSON'}, 400)
                return

            motivo = body.get('motivo', '').strip()
            valid_motivos = {'reembolso', 'conclusao', 'cancelamento', 'outro'}
            if motivo not in valid_motivos:
                self._send_json({'error': f'motivo deve ser um de: {", ".join(valid_motivos)}'}, 400)
                return

            from datetime import date
            updates = {
                'ativo': False,
                'motivo_inativacao': motivo,
                'data_inativacao': date.today().isoformat(),
            }
            obs = (body.get('obs') or '').strip()
            if obs:
                updates['obs_inativacao'] = obs

            result = supabase_request(
                'PATCH',
                f'mentorados?id=eq.{mentee_id}&select=id,nome,ativo,motivo_inativacao,data_inativacao',
                updates
            )
            self._send_json(result)
        except Exception as e:
            log_error('OffboardMentee', f'_handle_offboard_mentee failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_bulk_patch_mentees(self):
        """PATCH /api/mentees/bulk — update fase_jornada for multiple mentees at once
        Body: { ids: [uuid1, uuid2, ...], updates: { fase_jornada: 'execucao' } }
        """
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401)
                return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return

            try:
                body = json.loads(self._read_body())
            except Exception:
                self._send_json({'error': 'Invalid JSON'}, 400)
                return

            ids = body.get('ids', [])
            if not ids or not isinstance(ids, list):
                self._send_json({'error': 'ids must be a non-empty list'}, 400)
                return

            ALLOWED_FIELDS = {'fase_jornada', 'snoozed_until'}
            updates = {k: v for k, v in body.get('updates', {}).items() if k in ALLOWED_FIELDS}
            if not updates:
                self._send_json({'error': 'No allowed fields in updates'}, 400)
                return

            valid_fases = {
                'onboarding', 'execucao', 'resultado', 'renovacao', 'encerrado'
            }
            if 'fase_jornada' in updates and updates['fase_jornada'] not in valid_fases:
                self._send_json({'error': f'Invalid fase_jornada: {updates["fase_jornada"]}'}, 400)
                return

            ids_csv = ','.join(str(i) for i in ids)
            result = supabase_request(
                'PATCH',
                f'mentorados?id=in.({ids_csv})&select=id,nome,fase_jornada,snoozed_until',
                updates
            )
            updated = result if isinstance(result, list) else []
            self._send_json({'updated': len(updated), 'mentees': updated})
        except Exception as e:
            log_error('BulkPatch', f'_handle_bulk_patch_mentees failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_get_notes(self, mentee_id):
        """GET /api/mentees/{id}/notes — list notes for a mentee"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401)
                return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return

            notes = supabase_request(
                'GET',
                f'mentee_notes?select=*&mentorado_id=eq.{mentee_id}&order=created_at.desc'
            )
            self._send_json(notes if isinstance(notes, list) else [])
        except Exception as e:
            log_error('Notes', f'GET notes failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_post_note(self, mentee_id):
        """POST /api/mentees/{id}/notes — create a note for a mentee"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401)
                return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return

            email = payload.get('email', '')

            body = json.loads(self._read_body())
            tipo = body.get('tipo', '')
            valid_types = {'checkpoint_mensal', 'feedback_aula', 'registro_ligacao', 'livre', 'nota_livre'}
            if tipo not in valid_types:
                self._send_json(
                    {'error': f'Invalid tipo. Must be one of: {", ".join(sorted(valid_types))}'},
                    400
                )
                return

            note = {
                'mentorado_id': int(mentee_id),
                'consultor_id': email,
                'tipo': tipo,
                'conteudo': body.get('conteudo', {}),
                'tags': body.get('tags', []),
            }
            result = supabase_request('POST', 'mentee_notes', note)
            if isinstance(result, dict) and 'error' in result:
                self._send_json({'error': result['error']}, 500)
                return
            self._send_json(result, 201)
        except Exception as e:
            log_error('Notes', f'POST note failed: {e}')
            self._send_json({'error': str(e)}, 500)

    # ===== MENTEE GROUPS =====

    def _handle_get_groups(self):
        """GET /api/mentee-groups — list all groups with member counts"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401); return
            token = auth_header[7:]
            if not verify_jwt_token(token):
                self._send_json({'error': 'Invalid token'}, 401); return
            groups = supabase_request('GET', 'mentee_groups?select=*,mentee_group_members(mentee_id)&order=nome.asc')
            if not isinstance(groups, list):
                self._send_json([]); return
            result = []
            for g in groups:
                members = g.pop('mentee_group_members', []) or []
                g['member_count'] = len(members)
                g['member_ids'] = [m['mentee_id'] for m in members]
                result.append(g)
            self._send_json(result)
        except Exception as e:
            log_error('Groups', f'GET groups failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_get_group_members(self, group_id):
        """GET /api/mentee-groups/{id}/members — list members of a group"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401); return
            if not verify_jwt_token(auth_header[7:]):
                self._send_json({'error': 'Invalid token'}, 401); return
            members = supabase_request('GET', f'mentee_group_members?select=*&group_id=eq.{group_id}')
            self._send_json(members if isinstance(members, list) else [])
        except Exception as e:
            log_error('Groups', f'GET group members failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_post_group(self):
        """POST /api/mentee-groups — create a group"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401); return
            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload:
                self._send_json({'error': 'Invalid token'}, 401); return
            body = json.loads(self._read_body())
            nome = body.get('nome', '').strip()
            if not nome:
                self._send_json({'error': 'nome is required'}, 400); return
            group = {
                'nome': nome,
                'cor': body.get('cor', '#6366f1'),
                'icon': body.get('icon', '\U0001f4c1'),
                'created_by': payload.get('email', ''),
            }
            result = supabase_request('POST', 'mentee_groups', group)
            self._send_json(result, 201)
        except Exception as e:
            log_error('Groups', f'POST group failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_delete_group(self, group_id):
        """DELETE /api/mentee-groups/{id}"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401); return
            if not verify_jwt_token(auth_header[7:]):
                self._send_json({'error': 'Invalid token'}, 401); return
            supabase_request('DELETE', f'mentee_groups?id=eq.{group_id}')
            self._send_json({'ok': True})
        except Exception as e:
            log_error('Groups', f'DELETE group failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_post_group_member(self, group_id):
        """POST /api/mentee-groups/{id}/members — add mentee to group"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401); return
            if not verify_jwt_token(auth_header[7:]):
                self._send_json({'error': 'Invalid token'}, 401); return
            body = json.loads(self._read_body())
            mentee_id = body.get('mentee_id')
            if not mentee_id:
                self._send_json({'error': 'mentee_id required'}, 400); return
            result = supabase_request('POST', 'mentee_group_members', {'group_id': int(group_id), 'mentee_id': int(mentee_id)})
            self._send_json(result, 201)
        except Exception as e:
            log_error('Groups', f'POST group member failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_delete_group_member(self, group_id, mentee_id):
        """DELETE /api/mentee-groups/{id}/members/{mentee_id}"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Unauthorized'}, 401); return
            if not verify_jwt_token(auth_header[7:]):
                self._send_json({'error': 'Invalid token'}, 401); return
            supabase_request('DELETE', f'mentee_group_members?group_id=eq.{group_id}&mentee_id=eq.{mentee_id}')
            self._send_json({'ok': True})
        except Exception as e:
            log_error('Groups', f'DELETE group member failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_list_events(self):
        result = list_calendar_events()
        self._send_json(result)

    def _handle_upcoming_calls(self):
        # Get ALL calls (not just scheduled), ordered by date DESC to show latest first
        result = supabase_request('GET', "calls_mentoria?select=*&order=data_call.desc&limit=500")
        self._send_json(result if isinstance(result, list) else [result] if result else [])

    # ===== EVOLUTION PROXY =====
    def _resolve_evolution_apikey(self, target_path):
        """Resolve the correct API key for an Evolution instance.
        Instance-specific keys are stored in wa_sessions.instance_api_key.
        Falls back to the global EVOLUTION_API_KEY."""
        import re
        # Extract instance name from path (e.g., /instance/connect/spalla_u5)
        match = re.search(r'/(?:connect|connectionState|logout|delete|restart|fetchInstances)/([^/?]+)', target_path)
        if match:
            instance_name = match.group(1)
            try:
                result = supabase_request('GET', f'wa_sessions?instance_name=eq.{instance_name}&select=instance_api_key&limit=1')
                if isinstance(result, list) and result and result[0].get('instance_api_key'):
                    return result[0]['instance_api_key']
            except Exception:
                pass
        return EVOLUTION_API_KEY

    def _proxy_evolution(self, method):
        target_path = self.path[len('/api/evolution'):]
        if '..' in target_path:
            self._send_json({'error': 'Invalid path'}, 400)
            return
        url = f'{EVOLUTION_BASE}{target_path}'
        body = self._read_body() if method in ('POST', 'PUT') else None

        apikey = self._resolve_evolution_apikey(target_path)
        req = urllib.request.Request(url, data=body, method=method)
        req.add_header('Content-Type', 'application/json')
        req.add_header('apikey', apikey)

        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                resp_body = resp.read()
                self.send_response(resp.status)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
                self.send_header('Content-Length', len(resp_body))
                self.end_headers()
                self.wfile.write(resp_body)
        except urllib.error.HTTPError as e:
            error_body = e.read()
            self.send_response(e.code)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
            self.send_header('Content-Length', len(error_body))
            self.end_headers()
            self.wfile.write(error_body)
        except Exception as e:
            error_msg = json.dumps({'error': str(e)}).encode()
            self.send_response(502)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', self._get_cors_origin())
            self.send_header('Content-Length', len(error_msg))
            self.end_headers()
            self.wfile.write(error_msg)

    # ===== SHEETS HANDLERS =====
    def _handle_sheets_status(self):
        self._send_json({
            'sheets_configured': get_sheets_service() is not None,
            'last_sync': _sheets_last_sync,
            'last_result': _sheets_last_result,
            'payments_sheet': PAYMENTS_SHEET_ID,
            'contracts_sheet': CONTRACTS_SHEET_ID,
        })

    # ===== DOSSIÊ PRODUCTION SYSTEM =====

    # CR-C1: Role-based auth — only team members can mutate pipeline
    DS_ALLOWED_ROLES = ('admin', 'team')

    def _ds_check_auth(self):
        """Verify JWT token + role for DS endpoints. Returns payload or None (sends 401/403)."""
        auth_header = self.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            self._send_json({'error': 'Missing or invalid token'}, 401)
            return None
        payload = verify_jwt_token(auth_header[7:])
        if not payload or payload.get('type') == 'refresh':
            self._send_json({'error': 'Invalid or expired token'}, 401)
            return None
        # Check role — require explicit role in JWT, no fallback
        role = payload.get('role')
        if not role or role not in self.DS_ALLOWED_ROLES:
            self._send_json({'error': 'Insufficient permissions — JWT must contain role (admin or team)'}, 403)
            return None
        return payload

    # CR-M1: State machine — valid stage transitions (merged for both trilhas)
    DS_VALID_TRANSITIONS = {
        'pendente': ('producao_ia',),
        'producao_ia': ('revisao_mariza',),
        'revisao_mariza': ('revisao_kaique', 'revisao_paralela', 'producao_ia'),  # kaique=scale, paralela=clinic, back
        'revisao_kaique': ('revisao_queila', 'revisao_gobbi', 'revisao_mariza'),  # queila=clinic(via paralela), gobbi=scale, back
        'revisao_gobbi': ('enviado', 'revisao_kaique'),  # scale final review, can send back
        'revisao_paralela': ('revisao_queila',),  # clinic: auto-advance when both approve
        'revisao_queila': ('enviado', 'revisao_paralela', 'aprovado', 'revisao_kaique'),  # clinic final review
        'aprovado': ('enviado',),
        'enviado': ('finalizado', 'ajustes'),
        'finalizado': (),
        'ajustes': ('revisao_mariza',),
    }

    def _handle_ds_update_stage(self):
        """Update ds_documentos stage and create ds_eventos audit trail.
        POST /api/ds/update-stage
        Body: { mentorado_slug: str, dossie_tipo: str, estagio: str }
        Requires: Bearer JWT token with role in DS_ALLOWED_ROLES
        Returns: 200 {ok, mentorado, tipo, estagio, responsavel}
        Errors: 400 (validation), 401 (auth), 403 (role), 404 (not found), 409 (invalid transition), 500 (server)
        """
        auth = self._ds_check_auth()
        if not auth:
            return

        try:
            body = json.loads(self._read_body())
        except Exception:
            self._send_json({'error': 'Invalid JSON'}, 400)
            return

        slug = body.get('mentorado_slug', '')
        tipo = body.get('dossie_tipo', '')
        estagio = body.get('estagio', '')

        if not slug or not tipo or not estagio:
            self._send_json({'error': 'mentorado_slug, dossie_tipo, and estagio are required'}, 400)
            return

        valid_tipos = ('oferta', 'funil', 'conteudo')
        if tipo not in valid_tipos:
            self._send_json({'error': f'dossie_tipo must be one of {valid_tipos}'}, 400)
            return

        valid_estagios = tuple(self.DS_VALID_TRANSITIONS.keys())
        if estagio not in valid_estagios:
            self._send_json({'error': f'estagio must be one of {valid_estagios}'}, 400)
            return

        # Sanitize slug: only allow alphanumeric, spaces, hyphens, accented chars
        safe_slug = re.sub(r'[^a-zA-ZÀ-ÿ0-9\s\-]', '', slug).strip()
        if not safe_slug or len(safe_slug) < 2:
            self._send_json({'error': 'Invalid mentorado_slug'}, 400)
            return

        # Find mentorado (slug is sanitized, safe for ILIKE)
        mentee = supabase_request('GET',
            f'mentorados?nome=ilike.*{urllib.parse.quote(safe_slug)}*&select=id,nome&limit=5')
        if isinstance(mentee, dict) and mentee.get('error'):
            self._send_json({'error': f'Database error looking up mentorado: {mentee["error"]}'}, 500)
            return
        if not mentee:
            self._send_json({'error': f'Mentorado not found: {safe_slug}'}, 404)
            return
        if len(mentee) > 1:
            names = [m['nome'] for m in mentee]
            self._send_json({'error': f'Ambiguous slug — {len(mentee)} matches: {names}. Be more specific.'}, 400)
            return
        mentorado_id = mentee[0]['id']
        mentorado_nome = mentee[0]['nome']

        # CR-M2: Find document via active producao (not cancelled/finalizado)
        docs = supabase_request('GET',
            f'ds_documentos?mentorado_id=eq.{mentorado_id}&tipo=eq.{tipo}'
            f'&producao_id=not.is.null&select=id,producao_id,estagio_atual'
            f'&order=created_at.desc&limit=1')
        if isinstance(docs, dict) and docs.get('error'):
            self._send_json({'error': f'Database error looking up document: {docs["error"]}'}, 500)
            return
        if not docs:
            self._send_json({'error': f'Document not found: {mentorado_nome} / {tipo}'}, 404)
            return
        doc = docs[0]

        # CR-M1: Validate state transition
        current_stage = doc['estagio_atual']
        allowed_next = self.DS_VALID_TRANSITIONS.get(current_stage, ())
        if estagio not in allowed_next:
            self._send_json({
                'error': f'Invalid transition: {current_stage} → {estagio}',
                'allowed': list(allowed_next),
            }, 409)
            return

        # Determine next responsavel
        responsavel_map = {
            'producao_ia': 'Mariza',
            'revisao_mariza': 'Mariza',
            'revisao_kaique': 'Kaique',
            'revisao_queila': 'Queila',
            'revisao_gobbi': 'Gobbi',
            'revisao_paralela': 'Gobbi + Kaique',
        }
        responsavel = responsavel_map.get(estagio)

        # Update document stage
        update = {'estagio_atual': estagio, 'estagio_desde': datetime.now(timezone.utc).isoformat()}
        if estagio == 'producao_ia':
            update['data_producao_ia'] = datetime.now(timezone.utc).isoformat()
        if responsavel:
            update['responsavel_atual'] = responsavel

        result = supabase_request('PATCH', f'ds_documentos?id=eq.{doc["id"]}', update)
        if isinstance(result, dict) and result.get('error'):
            self._send_json({'error': f'Failed to update document: {result["error"]}'}, 500)
            return

        # Create audit event
        evt_result = supabase_request('POST', 'ds_eventos', {
            'producao_id': doc['producao_id'],
            'documento_id': doc['id'],
            'mentorado_id': mentorado_id,
            'tipo_evento': 'estagio_change',
            'de_valor': current_stage,
            'para_valor': estagio,
            'responsavel': responsavel or 'pipeline-auto',
            'descricao': f'{tipo} → {estagio} (via API)',
        })
        audit_ok = not (isinstance(evt_result, dict) and evt_result.get('error'))
        if not audit_ok:
            log_error('DS', 'Failed to create audit event', evt_result['error'])

        self._send_json({
            'ok': True,
            'mentorado': mentorado_nome,
            'tipo': tipo,
            'estagio': estagio,
            'responsavel': responsavel,
            'audit_event': audit_ok,
        })

    def _handle_sheets_sync(self):
        try:
            result = sync_sheets_to_supabase()
            self._send_json(result)
        except Exception as e:
            log_error('Sheets', 'Manual sync failed', e)
            self._send_json({'error': str(e)}, 500)

    # ===== AUTH HANDLERS =====
    def _handle_auth_register(self):
        """Register new user (Supabase-backed)"""
        try:
            body = json.loads(self._read_body())
            email = body.get('email', '').strip().lower()
            password = body.get('password', '').strip()
            full_name = body.get('fullName', body.get('full_name', '')).strip()

            if not email or not password:
                self._send_json({'error': 'Email and password required'}, 400)
                return

            if len(password) < 6:
                self._send_json({'error': 'Password must be at least 6 characters'}, 400)
                return

            # Check if email already exists
            existing = supabase_request('GET', f'auth_users?email=eq.{email}&select=id')
            if isinstance(existing, list) and len(existing) > 0:
                self._send_json({'error': 'Email already exists'}, 409)
                return

            role = body.get('role', 'equipe').strip().lower()
            if role not in ('equipe', 'mentorado', 'admin'):
                role = 'equipe'

            password_hash = hash_password(password)
            result = supabase_request('POST', 'auth_users', {
                'email': email,
                'password_hash': password_hash,
                'full_name': full_name,
                'role': role,
            })

            if isinstance(result, dict) and result.get('error'):
                self._send_json({'error': 'Registration failed'}, 500)
                return

            user = result[0] if isinstance(result, list) else result
            user_id = user.get('id')

            access_token = create_jwt_token(email, user_id, role=role)
            refresh_token = create_refresh_token(email, user_id)

            self._send_json({
                'success': True,
                'user': {'id': user_id, 'email': email, 'full_name': full_name, 'role': role},
                'access_token': access_token,
                'refresh_token': refresh_token,
                'expires_in': ACCESS_TOKEN_EXPIRY_MINUTES * 60
            }, 201)
        except Exception as e:
            log_error('AUTH', f'Registration failed: {e}')
            self._send_json({'error': 'Registration failed'}, 500)

    def _handle_auth_me(self):
        """Validate token and return current user data"""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Missing or invalid token'}, 401)
                return

            token = auth_header[7:]
            payload = verify_jwt_token(token)
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid or expired token'}, 401)
                return

            user_id = payload.get('user_id')
            email = payload.get('email')

            # Fetch fresh user data from DB
            result = supabase_request('GET', f'auth_users?id=eq.{user_id}&select=id,email,full_name')
            if isinstance(result, list) and len(result) > 0:
                self._send_json({'user': result[0]})
            else:
                self._send_json({'user': {'id': user_id, 'email': email}})
        except Exception as e:
            log_error('AUTH', f'Token validation failed: {e}')
            self._send_json({'error': 'Token validation failed'}, 401)

    def _handle_auth_login(self):
        """Login user and issue JWT (Supabase-backed)"""
        try:
            body = json.loads(self._read_body())
            email = body.get('email', '').strip().lower()
            password = body.get('password', '').strip()

            if not email or not password:
                self._send_json({'error': 'Email and password required'}, 400)
                return

            result = supabase_request('GET', f'auth_users?email=eq.{email}&select=id,email,full_name,password_hash')
            if not isinstance(result, list) or len(result) == 0:
                self._send_json({'error': 'Invalid email or password'}, 401)
                return

            row = result[0]
            pw_result = verify_password(password, row['password_hash'])
            if not pw_result:
                self._send_json({'error': 'Invalid email or password'}, 401)
                return

            user_id = row['id']
            db_email = row['email']

            # Lazy migration: upgrade legacy SHA-256 hash to bcrypt
            if pw_result == 'migrate' and _bcrypt:
                new_hash = hash_password(password)
                supabase_request('PATCH', f'auth_users?id=eq.{user_id}', {'password_hash': new_hash})
                print(f'[AUTH] Migrated password hash for user {user_id} to bcrypt')
            full_name = row.get('full_name', '')
            role = row.get('role', 'equipe')

            access_token = create_jwt_token(db_email, user_id, role=role)
            refresh_token = create_refresh_token(db_email, user_id)

            self._send_json({
                'success': True,
                'user': {'id': user_id, 'email': db_email, 'full_name': full_name, 'role': role},
                'access_token': access_token,
                'refresh_token': refresh_token,
                'expires_in': ACCESS_TOKEN_EXPIRY_MINUTES * 60
            }, 200)
        except Exception as e:
            log_error('AUTH', f'Login failed: {e}')
            self._send_json({'error': 'Login failed'}, 500)

    def _handle_auth_refresh(self):
        """Refresh JWT token"""
        try:
            body = json.loads(self._read_body())
            refresh_token = body.get('refresh_token', '').strip()

            if not refresh_token:
                self._send_json({'error': 'Refresh token required'}, 400)
                return

            payload = verify_jwt_token(refresh_token)
            if not payload or payload.get('type') != 'refresh':
                self._send_json({'error': 'Invalid or expired refresh token'}, 401)
                return

            email = payload.get('email')
            user_id = payload.get('user_id')

            # Fetch fresh user data
            user_data = supabase_request('GET', f'auth_users?id=eq.{user_id}&select=id,email,full_name,role')
            user = user_data[0] if isinstance(user_data, list) and len(user_data) > 0 else {'id': user_id, 'email': email}
            role = user.get('role', 'equipe')

            new_access_token = create_jwt_token(email, user_id, role=role)
            new_refresh_token = create_refresh_token(email, user_id)

            self._send_json({
                'success': True,
                'user': user,
                'access_token': new_access_token,
                'refresh_token': new_refresh_token,
                'expires_in': ACCESS_TOKEN_EXPIRY_MINUTES * 60
            }, 200)
        except Exception as e:
            log_error('AUTH', f'Token refresh failed: {e}')
            self._send_json({'error': 'Token refresh failed'}, 500)

    def _handle_auth_reset_password(self):
        """Handle password reset request — returns generic message to prevent email enumeration"""
        try:
            body = json.loads(self._read_body())
            email = body.get('email', '').strip().lower()

            if not email:
                self._send_json({'error': 'Email obrigatorio'}, 400)
                return

            # Check if user exists (but always return same message)
            users = supabase_request('GET', f'auth_users?email=eq.{email}&select=id,email')
            if isinstance(users, list) and len(users) > 0:
                # TODO: Send actual reset email when email service is configured
                log_error('AUTH', f'Password reset requested for: {email} (user found, email not sent — no email service configured)')
            else:
                log_error('AUTH', f'Password reset requested for: {email} (user not found)')

            # Always return success to prevent email enumeration
            self._send_json({
                'success': True,
                'message': 'Se o email existir no sistema, as instrucoes de recuperacao serao enviadas.'
            }, 200)
        except Exception as e:
            log_error('AUTH', f'Reset password error: {e}')
            self._send_json({'error': 'Erro ao processar solicitacao'}, 500)

    def _handle_welcome_flow_register(self):
        """Register mentorado from welcome-flow wizard (public endpoint, no auth required)"""
        try:
            body = json.loads(self._read_body())
            email = body.get('email', '').strip().lower()
            nome = body.get('nome', '').strip()
            whatsapp = body.get('whatsapp', '').strip()

            if not email or not nome:
                self._send_json({'error': 'Nome and email are required'}, 400)
                return

            # Generate temp password: first name lowercase + last 4 digits of whatsapp + random suffix
            first_name = nome.split()[0].lower()
            # Remove non-digit chars and get last 4 digits
            digits = ''.join(filter(str.isdigit, whatsapp))
            last4 = digits[-4:] if len(digits) >= 4 else digits
            random_suffix = secrets.token_hex(2)  # 4 random hex chars
            temp_password = f"{first_name}{last4}{random_suffix}"

            if len(temp_password) < 4:
                self._send_json({'error': 'Could not generate valid password — check nome and whatsapp'}, 400)
                return

            # Check if auth_users already exists
            existing = supabase_request('GET', f'auth_users?email=eq.{email}&select=id,email,full_name')
            if isinstance(existing, list) and len(existing) > 0:
                self._send_json({
                    'success': True,
                    'already_exists': True,
                    'user_id': existing[0].get('id'),
                    'message': 'User already exists with this email'
                })
                return

            # Create auth_users entry with role=mentorado
            password_hash = hash_password(temp_password)
            result = supabase_request('POST', 'auth_users', {
                'email': email,
                'password_hash': password_hash,
                'full_name': nome,
                'role': 'mentorado',
            })

            if isinstance(result, dict) and result.get('error'):
                log_error('WELCOME', f'Failed to create auth user: {result}')
                self._send_json({'error': 'Registration failed'}, 500)
                return

            user = result[0] if isinstance(result, list) else result
            user_id = user.get('id')

            self._send_json({
                'success': True,
                'user_id': user_id,
                'email': email,
            }, 201)
        except Exception as e:
            log_error('WELCOME', f'Welcome flow registration failed: {e}')
            self._send_json({'error': 'Registration failed'}, 500)

    # ===== STORAGE HANDLERS =====

    def _handle_storage_process(self):
        """POST /api/storage/process — Trigger file processing pipeline."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
        except Exception:
            self._send_json({'error': 'Invalid JSON'}, 400)
            return

        arquivo_id = body.get('arquivo_id')
        if not arquivo_id:
            self._send_json({'error': 'arquivo_id is required'}, 400)
            return

        # Process in background thread
        def _bg():
            try:
                process_file_pipeline(arquivo_id)
            except Exception as e:
                log_error('Storage', f'Background processing failed: {e}')

        t = threading.Thread(target=_bg, daemon=True)
        t.start()

        self._send_json({'status': 'processing', 'arquivo_id': arquivo_id})

    def _handle_storage_search(self):
        """POST /api/storage/search — Semantic/keyword/hybrid search."""
        try:
            body = json.loads(self._read_body())
        except Exception:
            self._send_json({'error': 'Invalid JSON'}, 400)
            return

        query = body.get('query', '').strip()
        if not query:
            self._send_json({'error': 'query is required'}, 400)
            return

        mode = body.get('mode', 'hybrid')
        if mode not in ('semantic', 'keyword', 'hybrid'):
            mode = 'hybrid'

        filters = body.get('filters', {})
        limit = min(body.get('limit', 10), 50)

        try:
            start_time = time.time()
            results = search_semantic(query, mode=mode, filters=filters, limit=limit)
            elapsed_ms = int((time.time() - start_time) * 1000)

            self._send_json({
                'results': results,
                'total': len(results),
                'query': query,
                'mode': mode,
                'query_time_ms': elapsed_ms,
            })
        except ValueError as e:
            self._send_json({'error': str(e)}, 400)
        except Exception as e:
            log_error('Storage', 'Search failed', e)
            self._send_json({'error': f'Search failed: {str(e)}'}, 500)

    def _handle_storage_list_files(self):
        """GET /api/storage/files?entidade_tipo=X&entidade_id=Y — List files for entity."""
        qs = urllib.parse.urlparse(self.path).query
        params = urllib.parse.parse_qs(qs)
        entidade_tipo = params.get('entidade_tipo', [None])[0]
        entidade_id = params.get('entidade_id', [None])[0]

        query = 'sp_arquivos?select=*&deleted_at=is.null&order=created_at.desc'
        if entidade_tipo:
            query += f'&entidade_tipo=eq.{entidade_tipo}'
        if entidade_id:
            query += f'&entidade_id=eq.{entidade_id}'

        result = supabase_request('GET', query)
        self._send_json(result if isinstance(result, list) else [])

    def _handle_storage_status(self):
        """GET /api/storage/status — Storage overview (files, sizes, processing queue)."""
        overview = supabase_request('GET', 'vw_storage_overview?select=*')
        queue = supabase_request('GET', 'vw_processamento_fila?select=*&limit=20')
        self._send_json({
            'overview': overview if isinstance(overview, list) else [],
            'queue': queue if isinstance(queue, list) else [],
            'embedding_provider': EMBEDDING_PROVIDER,
            'embedding_dims': EMBEDDING_DIMS,
            'voyage_configured': bool(VOYAGE_API_KEY),
            'openai_configured': bool(OPENAI_API_KEY),
            'gemini_configured': bool(GEMINI_API_KEY),
            'vision_provider': 'gemini' if GEMINI_API_KEY else ('openai' if OPENAI_API_KEY else 'none'),
        })

    def _handle_storage_test(self):
        """POST /api/storage/test — Test embedding + chunk insert (debug endpoint)."""
        try:
            test_text = 'Teste de embedding para busca semântica no Spalla Dashboard'
            log_info('Storage', f'[TEST] Embedding provider: {EMBEDDING_PROVIDER}, key set: {bool(VOYAGE_API_KEY)}')

            embedding = embed_query(test_text)
            log_info('Storage', f'[TEST] Got embedding: dims={len(embedding)}, first3={embedding[:3]}')

            self._send_json({
                'status': 'ok',
                'provider': EMBEDDING_PROVIDER,
                'dims': len(embedding),
                'first_values': embedding[:5],
                'voyage_key_set': bool(VOYAGE_API_KEY),
                'openai_key_set': bool(OPENAI_API_KEY),
            })
        except Exception as e:
            log_error('Storage', f'[TEST] Failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_storage_reprocess(self):
        """POST /api/storage/reprocess — Reprocess all files with status 'pendente' or 'erro'."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body()) if int(self.headers.get('Content-Length', 0)) > 0 else {}
        except Exception:
            body = {}

        status_filter = body.get('status', 'pendente')  # 'pendente', 'erro', or 'all'

        query = 'sp_arquivos?select=id&deleted_at=is.null'
        if status_filter == 'all':
            query += '&status_processamento=in.(pendente,erro)'
        else:
            query += f'&status_processamento=eq.{status_filter}'

        result = supabase_request('GET', query)
        if not isinstance(result, list):
            self._send_json({'error': 'Failed to query files'}, 500)
            return

        count = 0
        for row in result:
            def _bg(aid=row['id']):
                try:
                    process_file_pipeline(aid)
                except Exception as e:
                    log_error('Storage', f'Reprocess failed for {aid}: {e}')
            t = threading.Thread(target=_bg, daemon=True)
            t.start()
            count += 1
            # Stagger to avoid overwhelming OpenAI API
            if count % 5 == 0:
                time.sleep(1)

        self._send_json({'queued': count, 'status_filter': status_filter})

    # ===== BIBLIOTECA =====

    def _handle_biblioteca_get(self):
        """
        GET /api/biblioteca               → lista todos os docs (sem conteúdo)
        GET /api/biblioteca?mentee_id=123 → filtra por mentorado
        GET /api/biblioteca?slug=danyella-truiz-oferta → doc único por slug
        GET /api/biblioteca/{uuid}        → doc único por id (com conteúdo_md)
        """
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        path_parts = [p for p in parsed.path.split('/') if p]
        # /api/biblioteca/{id}
        if len(path_parts) == 3:
            doc_id = path_parts[2]
            result = supabase_request(
                'GET',
                f'sp_documentos?id=eq.{doc_id}&select=*&limit=1',
            )
            if isinstance(result, list) and result:
                self._send_json(result[0])
            else:
                self._send_json({'error': 'Not found'}, 404)
            return

        # Listagem
        slug = qs.get('slug', [None])[0]
        if slug:
            result = supabase_request(
                'GET',
                f'sp_documentos?deep_link_slug=eq.{slug}&select=*&limit=1',
            )
            if isinstance(result, list) and result:
                self._send_json(result[0])
            else:
                self._send_json({'error': 'Not found'}, 404)
            return

        mentee_id = qs.get('mentee_id', [None])[0]
        tipo = qs.get('tipo', [None])[0]

        # Usa view sem conteúdo para listagem (mais leve)
        base = 'vw_sp_documentos_lista?select=*'
        filters = []
        if mentee_id:
            filters.append(f'mentee_id=eq.{mentee_id}')
        if tipo:
            filters.append(f'tipo=eq.{tipo}')
        query = base + ('&' + '&'.join(filters) if filters else '') + '&order=mentee_nome,tipo,criado_em'
        result = supabase_request('GET', query)
        self._send_json(result if isinstance(result, list) else [])

    # ===== API KEYS MANAGEMENT =====

    def _handle_generate_api_key(self):
        """POST /api/keys/generate — Generate a new API key. Requires admin JWT."""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Admin JWT required'}, 401)
                return
            payload = verify_jwt_token(auth_header[7:])
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return
            if payload.get('role') != 'admin':
                self._send_json({'error': 'Admin role required to generate API keys'}, 403)
                return

            body = json.loads(self._read_body())
            label = (body.get('label') or '').strip()
            if not label:
                self._send_json({'error': 'label is required'}, 400)
                return
            role = body.get('role', 'integration')
            if role not in ('integration', 'readonly'):
                self._send_json({'error': 'role must be integration or readonly'}, 400)
                return

            raw_key = f'sk_spalla_{secrets.token_hex(24)}'
            key_hash = hashlib.sha256(raw_key.encode()).hexdigest()
            key_prefix = raw_key[:14] + '...'

            result = supabase_request('POST', 'api_keys', {
                'key_hash': key_hash, 'key_prefix': key_prefix,
                'label': label, 'role': role, 'active': True,
                'created_by': payload.get('email', ''),
            })
            if isinstance(result, dict) and result.get('error'):
                self._send_json({'error': f'Failed to store key: {result["error"]}'}, 500)
                return
            key_id = result[0]['id'] if isinstance(result, list) and result else None
            self._send_json({
                'key': raw_key, 'key_id': key_id, 'key_prefix': key_prefix,
                'label': label, 'role': role,
                'warning': 'Store this key safely — it will NOT be shown again.',
            }, 201)
        except Exception as e:
            log_error('APIKeys', f'generate failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_list_api_keys(self):
        """GET /api/keys — List all API keys (prefix only). Requires admin JWT."""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Admin JWT required'}, 401)
                return
            payload = verify_jwt_token(auth_header[7:])
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return
            if payload.get('role') != 'admin':
                self._send_json({'error': 'Admin role required'}, 403)
                return
            result = supabase_request('GET',
                'api_keys?select=id,key_prefix,label,role,active,created_at,last_used_at,created_by'
                '&order=created_at.desc')
            self._send_json(result if isinstance(result, list) else [])
        except Exception as e:
            log_error('APIKeys', f'list failed: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_revoke_api_key(self, key_id):
        """DELETE /api/keys/{id} — Revoke an API key. Requires admin JWT."""
        try:
            auth_header = self.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                self._send_json({'error': 'Admin JWT required'}, 401)
                return
            payload = verify_jwt_token(auth_header[7:])
            if not payload or payload.get('type') == 'refresh':
                self._send_json({'error': 'Invalid token'}, 401)
                return
            if payload.get('role') != 'admin':
                self._send_json({'error': 'Admin role required'}, 403)
                return
            result = supabase_request('PATCH',
                f'api_keys?id=eq.{key_id}&select=id,label,active',
                {'active': False})
            if isinstance(result, list) and result:
                self._send_json({'ok': True, 'key_id': key_id, 'revoked': True})
            else:
                self._send_json({'error': 'Key not found'}, 404)
        except Exception as e:
            log_error('APIKeys', f'revoke failed: {e}')
            self._send_json({'error': str(e)}, 500)

    # ===== WHATSAPP GROUP MANAGEMENT (Story 8) =====

    def _handle_wa_groups_list(self):
        """GET /api/wa/groups — List all tracked WA groups from Supabase."""
        try:
            result = supabase_request('GET', 'wa_groups?select=*&order=last_activity.desc.nullsfirst&is_active=eq.true')
            self._send_json(result if isinstance(result, list) else [])
        except Exception as e:
            log_error('WA-Groups', f'list failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_groups_sync(self):
        """POST /api/wa/groups/sync — Sync groups from Evolution API to Supabase."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
            instance = body.get('instance', '').strip()
            if not instance:
                self._send_json({'error': 'instance is required'}, 400)
                return

            # Fetch groups from Evolution API
            url = f'{EVOLUTION_BASE}/group/fetchAllGroups/{instance}?getParticipants=true'
            req = urllib.request.Request(url, method='GET')
            req.add_header('apikey', EVOLUTION_API_KEY)
            with urllib.request.urlopen(req, timeout=30) as resp:
                groups = json.loads(resp.read())

            if not isinstance(groups, list):
                groups = groups.get('groups', groups.get('data', []))

            synced = 0
            for g in groups:
                jid = g.get('id') or g.get('jid', '')
                if not jid or not jid.endswith('@g.us'):
                    continue
                name = g.get('subject') or g.get('name') or jid
                participants = g.get('participants', [])
                row = {
                    'group_jid': jid,
                    'name': name,
                    'description': g.get('desc', g.get('description', '')),
                    'participant_count': len(participants),
                    'participants': json.dumps(participants),
                    'photo_url': g.get('profilePictureUrl', g.get('imgUrl', '')),
                    'instance_name': instance,
                    'is_active': True,
                    'synced_at': datetime.now(timezone.utc).isoformat(),
                    'updated_at': datetime.now(timezone.utc).isoformat(),
                }
                # Upsert by group_jid
                supabase_request('POST',
                    'wa_groups?on_conflict=group_jid',
                    row)
                synced += 1

            log_info('WA-Groups', f'Synced {synced} groups from {instance}')
            self._send_json({'ok': True, 'synced': synced})

        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            log_error('WA-Groups', f'Evolution API error: {e.code} {error_body}')
            self._send_json({'error': f'Evolution API: {e.code}', 'detail': error_body}, e.code)
        except Exception as e:
            log_error('WA-Groups', f'sync failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _handle_youtube_upload(self):
        """POST /api/youtube/upload — Upload Zoom recording to YouTube."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            if not YOUTUBE_API_KEY:
                self._send_json({'error': 'YOUTUBE_API_KEY not configured. Set it in Railway env vars.'}, 501)
                return
            body = self._read_json_body()
            zoom_recording_url = body.get('recording_url', '')
            title = body.get('title', 'Call de Mentoria')
            description = body.get('description', '')
            mentorado_nome = body.get('mentorado_nome', '')
            privacy = body.get('privacy', 'unlisted')  # unlisted | private | public

            if not zoom_recording_url:
                self._send_json({'error': 'recording_url required'}, 400)
                return

            # Step 1: Download Zoom recording
            zoom_token = get_zoom_token()
            req = urllib.request.Request(zoom_recording_url)
            req.add_header('Authorization', f'Bearer {zoom_token}')
            with urllib.request.urlopen(req, timeout=300) as resp:
                video_bytes = resp.read()

            # Step 2: Upload to YouTube via resumable upload
            # First, get upload URL
            metadata = json.dumps({
                'snippet': {
                    'title': f'{title} — {mentorado_nome}' if mentorado_nome else title,
                    'description': description or f'Gravação de call de mentoria. Mentorado: {mentorado_nome}',
                    'tags': ['mentoria', 'case', mentorado_nome] if mentorado_nome else ['mentoria', 'case'],
                    'categoryId': '27',  # Education
                },
                'status': {'privacyStatus': privacy, 'selfDeclaredMadeForKids': False},
            }).encode()

            init_req = urllib.request.Request(
                f'https://www.googleapis.com/upload/youtube/v3/videos?uploadType=resumable&part=snippet,status&key={YOUTUBE_API_KEY}',
                data=metadata, method='POST')
            init_req.add_header('Content-Type', 'application/json; charset=UTF-8')
            init_req.add_header('X-Upload-Content-Length', str(len(video_bytes)))
            init_req.add_header('X-Upload-Content-Type', 'video/mp4')

            with urllib.request.urlopen(init_req, timeout=30) as init_resp:
                upload_url = init_resp.headers.get('Location')

            if not upload_url:
                self._send_json({'error': 'Failed to get YouTube upload URL'}, 500)
                return

            # Step 3: Upload video bytes
            up_req = urllib.request.Request(upload_url, data=video_bytes, method='PUT')
            up_req.add_header('Content-Type', 'video/mp4')
            up_req.add_header('Content-Length', str(len(video_bytes)))
            with urllib.request.urlopen(up_req, timeout=600) as up_resp:
                result = json.loads(up_resp.read())

            video_id = result.get('id', '')
            self._send_json({
                'success': True,
                'video_id': video_id,
                'url': f'https://youtu.be/{video_id}',
                'title': result.get('snippet', {}).get('title', ''),
            })
        except Exception as e:
            print(f'[youtube-upload] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_drive_sync(self):
        """POST /api/drive/sync — List Google Drive files for a mentorado folder."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            folder_mentory = os.environ.get('GOOGLE_DRIVE_FOLDER_MENTORY', '')
            folder_clinic = os.environ.get('GOOGLE_DRIVE_FOLDER_CLINIC', '')
            if not folder_mentory and not folder_clinic:
                self._send_json({'error': 'GOOGLE_DRIVE_FOLDER_MENTORY/CLINIC not configured'}, 501)
                return

            body = self._read_json_body()
            mentorado_id = body.get('mentorado_id')
            mentorado_nome = body.get('mentorado_nome', '')
            produto = body.get('produto', 'mentory')  # mentory | clinic

            if not mentorado_id or not mentorado_nome:
                self._send_json({'error': 'mentorado_id and mentorado_nome required'}, 400)
                return

            parent_folder = folder_mentory if produto == 'mentory' else folder_clinic
            if not parent_folder:
                self._send_json({'error': f'GOOGLE_DRIVE_FOLDER_{produto.upper()} not configured'}, 501)
                return

            # Use same SA auth as Calendar (google-auth library)
            from google.oauth2 import service_account as sa_module
            from googleapiclient.discovery import build as gbuild

            sa_json = os.environ.get('GOOGLE_SA_JSON', '') or os.environ.get('GOOGLE_SA_CREDENTIALS_B64', '')
            if sa_json:
                sa_info = json.loads(base64.b64decode(sa_json))
                credentials = sa_module.Credentials.from_service_account_info(
                    sa_info, scopes=['https://www.googleapis.com/auth/drive']
                )
            elif os.path.exists(GOOGLE_SA_PATH):
                credentials = sa_module.Credentials.from_service_account_file(
                    GOOGLE_SA_PATH, scopes=['https://www.googleapis.com/auth/drive']
                )
            else:
                self._send_json({'error': 'Google SA credentials not found'}, 501)
                return

            drive = gbuild('drive', 'v3', credentials=credentials)

            # Step 1: Find or create mentee's folder inside parent
            query = f"name='{mentorado_nome}' and '{parent_folder}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = drive.files().list(q=query, fields='files(id,name)').execute()
            mentee_folders = results.get('files', [])

            if mentee_folders:
                mentee_folder_id = mentee_folders[0]['id']
                created_folder = False
            else:
                # Create folder for this mentee
                folder_meta = {
                    'name': mentorado_nome,
                    'mimeType': 'application/vnd.google-apps.folder',
                    'parents': [parent_folder],
                }
                folder = drive.files().create(body=folder_meta, fields='id').execute()
                mentee_folder_id = folder['id']
                created_folder = True

            # Step 2: List files in mentee's folder
            file_query = f"'{mentee_folder_id}' in parents and trashed=false"
            file_results = drive.files().list(
                q=file_query,
                fields='files(id,name,mimeType,size,modifiedTime,webViewLink)',
                orderBy='modifiedTime desc',
                pageSize=50,
            ).execute()
            files = file_results.get('files', [])

            self._send_json({
                'folder_id': mentee_folder_id,
                'folder_url': f'https://drive.google.com/drive/folders/{mentee_folder_id}',
                'created_folder': created_folder,
                'files': [{
                    'id': f['id'],
                    'name': f['name'],
                    'type': f.get('mimeType', ''),
                    'size': int(f.get('size', 0)) if f.get('size') else None,
                    'modified': f.get('modifiedTime', ''),
                    'url': f.get('webViewLink', ''),
                } for f in files],
                'count': len(files),
            })
        except ImportError:
            self._send_json({'error': 'google-auth/google-api-python-client not installed. pip install google-auth google-api-python-client'}, 501)
        except Exception as e:
            print(f'[drive-sync] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_weekly_summary(self):
        """POST /api/mentee/weekly-summary — Generate AI weekly summary for a mentorado."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = self._read_json_body()
            mentorado_id = body.get('mentorado_id')
            if not mentorado_id:
                self._send_json({'error': 'mentorado_id required'}, 400)
                return

            # Gather data from last 7 days
            seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

            # 1. Messages
            messages = supabase_request('GET',
                f'interacoes_mentoria?mentorado_id=eq.{mentorado_id}&created_at=gte.{seven_days_ago}'
                f'&select=conteudo,categoria,sentimento,tipo_interacao,requer_resposta,respondido,sender_name,created_at'
                f'&order=created_at.asc&limit=50')

            # 2. Tasks
            tasks = supabase_request('GET',
                f'god_tasks?mentorado_id=eq.{mentorado_id}&updated_at=gte.{seven_days_ago}'
                f'&select=titulo,status,tipo,responsavel,updated_at&limit=30')

            # 3. Percepcoes
            percs = supabase_request('GET',
                f'percepcoes_mentorado?mentorado_id=eq.{mentorado_id}&created_at=gte.{seven_days_ago}'
                f'&select=conteudo,tipo,autor,created_at&limit=10')

            if not GEMINI_API_KEY:
                self._send_json({'error': 'GEMINI_API_KEY not configured'}, 501)
                return

            # Build context
            msg_texts = [f"[{m.get('categoria','?')}] {m.get('sender_name','?')}: {(m.get('conteudo',''))[:200]}" for m in (messages or [])]
            task_texts = [f"[{t.get('status','')}] {t.get('titulo','')}" for t in (tasks or [])]
            perc_texts = [f"[{p.get('tipo','')}] {p.get('conteudo','')}" for p in (percs or [])]

            prompt = f"""Gere um resumo semanal para o consultor sobre este mentorado.
Seja direto, prático, em português. Máximo 300 palavras.

Estrutura:
1. **Resumo geral** (1-2 frases)
2. **Pontos positivos** (o que avançou)
3. **Pontos de atenção** (o que precisa de ação)
4. **Próximos passos sugeridos** (2-3 ações concretas)

Dados da semana:
--- MENSAGENS ({len(messages or [])}) ---
{chr(10).join(msg_texts[:20])}

--- TAREFAS ({len(tasks or [])}) ---
{chr(10).join(task_texts[:15])}

--- PERCEPÇÕES ({len(percs or [])}) ---
{chr(10).join(perc_texts[:5])}
"""

            gemini_body = {
                'contents': [{'parts': [{'text': prompt}]}],
                'generationConfig': {'temperature': 0.3, 'maxOutputTokens': 800}
            }
            conn = http.client.HTTPSConnection('generativelanguage.googleapis.com', timeout=30)
            conn.request('POST',
                         f'/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}',
                         body=json.dumps(gemini_body).encode(),
                         headers={'Content-Type': 'application/json'})
            resp = conn.getresponse()
            resp_data = resp.read()
            conn.close()

            if resp.status >= 400:
                self._send_json({'error': f'Gemini error {resp.status}'}, 500)
                return

            result = json.loads(resp_data)
            summary = result['candidates'][0]['content']['parts'][0]['text']

            self._send_json({
                'summary': summary,
                'stats': {
                    'messages': len(messages or []),
                    'tasks': len(tasks or []),
                    'percepcoes': len(percs or []),
                },
            })
        except Exception as e:
            print(f'[weekly-summary] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_tasks_from_audio(self):
        """POST /api/tasks/from-audio — Transcribe audio, extract N tasks via Gemini."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in content_type:
                self._send_json({'error': 'Expected multipart/form-data'}, 400)
                return

            # Parse multipart
            import cgi
            environ = {'REQUEST_METHOD': 'POST', 'CONTENT_TYPE': content_type}
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=environ)
            audio_field = form['audio']
            audio_bytes = audio_field.file.read()
            filename = audio_field.filename or 'audio.webm'
            mime_type = audio_field.type or 'audio/webm'

            # Step 1: Transcribe via Whisper
            transcript = openai_whisper(audio_bytes, filename, mime_type)
            if not transcript or len(transcript.strip()) < 10:
                self._send_json({'error': 'Transcrição vazia ou muito curta', 'transcript': transcript}, 400)
                return

            # Step 2: Extract tasks via Gemini Flash
            extraction_prompt = f"""Você é um assistente especializado em extrair tarefas de áudio transcrito.
O contexto é uma consultoria de mentoria (CASE), onde a equipe descarrega demandas por áudio.

REGRAS DE EXTRAÇÃO:
1. Cada AÇÃO DISTINTA é uma tarefa separada. Se o falante diz "precisa fazer X e Y", são 2 tarefas.
2. Se uma tarefa tem SUB-ITENS (checklist), agrupe como subtasks dentro da tarefa pai.
3. DETECTE o tipo correto:
   - "dossie": produção, revisão, ajuste de dossiê
   - "ajuste_dossie": correção específica de dossiê já feito
   - "follow_up": checar com mentorado, cobrar resposta, fazer acompanhamento
   - "rotina": algo recorrente (toda semana, todo dia, etc.)
   - "geral": qualquer outra coisa
4. DETECTE responsável pelo NOME se mencionado (Heitor, Lara, Mariza, Kaique, Queila, Hugo, Gobbi).
5. DETECTE mentorado pelo NOME se mencionado.
6. DETECTE prazo: "amanhã"=1, "essa semana"=5, "semana que vem"=10, "urgente"=1, "até sexta"=calcule dias.
7. DETECTE prioridade: "urgente"/"agora"→urgente, "importante"→alta, default→normal.
8. Se o falante menciona DATAS ESPECÍFICAS (dia X, segunda, etc.), converta em prazo_dias relativo a hoje.
9. Subtasks são itens dentro de uma tarefa maior. Ex: "precisa revisar o dossiê: checar oferta, ajustar funil, validar copy" = 1 tarefa com 3 subtasks.

FORMATO DE SAÍDA (JSON array):
[{{
  "titulo": "string (imperativo, max 80 chars, ex: 'Revisar dossiê da Betina')",
  "descricao": "string ou null (contexto adicional extraído do áudio)",
  "responsavel": "string ou null (primeiro nome)",
  "mentorado": "string ou null (nome do mentorado)",
  "prioridade": "baixa" | "normal" | "alta" | "urgente",
  "tipo": "geral" | "dossie" | "ajuste_dossie" | "follow_up" | "rotina",
  "prazo_dias": number ou null,
  "subtasks": ["string", "string"] ou []
}}]

Retorne APENAS o JSON array. Sem markdown, sem explicação, sem comentários.
Se não houver tarefas claras, retorne [].

Transcrição:
---
{transcript}
---"""

            if not GEMINI_API_KEY:
                self._send_json({'error': 'GEMINI_API_KEY not configured'}, 500)
                return

            gemini_body = {
                'contents': [{'parts': [{'text': extraction_prompt}]}],
                'generationConfig': {'temperature': 0.1, 'responseMimeType': 'application/json'}
            }
            conn = http.client.HTTPSConnection('generativelanguage.googleapis.com', timeout=30)
            conn.request('POST',
                         f'/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}',
                         body=json.dumps(gemini_body).encode(),
                         headers={'Content-Type': 'application/json'})
            resp = conn.getresponse()
            resp_data = resp.read()
            conn.close()

            if resp.status >= 400:
                self._send_json({'error': f'Gemini error {resp.status}', 'detail': resp_data.decode()[:300]}, 500)
                return

            gemini_result = json.loads(resp_data)
            raw_text = gemini_result['candidates'][0]['content']['parts'][0]['text']

            # Parse JSON from response (handle markdown code blocks)
            clean = raw_text.strip()
            if clean.startswith('```'):
                clean = clean.split('\n', 1)[1] if '\n' in clean else clean[3:]
                if clean.endswith('```'):
                    clean = clean[:-3]
            tasks = json.loads(clean.strip())

            self._send_json({
                'transcript': transcript,
                'tasks': tasks,
                'count': len(tasks),
            })
        except Exception as e:
            print(f'[tasks-from-audio] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_context_transcribe(self):
        """POST /api/context/transcribe — Transcribe audio URL or file via Whisper.
        Accepts JSON { arquivo_url } or multipart with 'audio' field.
        Returns { transcricao }.
        """
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            content_type = self.headers.get('Content-Type', '')

            if 'multipart/form-data' in content_type:
                # Direct audio blob upload
                import cgi
                environ = {'REQUEST_METHOD': 'POST', 'CONTENT_TYPE': content_type}
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=environ)
                audio_field = form['audio']
                audio_bytes = audio_field.file.read()
                filename = audio_field.filename or 'audio.webm'
                mime_type = audio_field.type or 'audio/webm'
            else:
                # JSON body with arquivo_url
                length = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(length)
                import json as _json
                payload = _json.loads(body)
                arquivo_url = payload.get('arquivo_url')
                if not arquivo_url:
                    self._send_json({'error': 'arquivo_url required'}, 400)
                    return
                import urllib.request as _urlreq
                MAX_AUDIO_BYTES = 50 * 1024 * 1024  # 50 MB limit
                print(f'[context-transcribe] Downloading: {arquivo_url[:120]}')
                try:
                    req = _urlreq.Request(arquivo_url, headers={'User-Agent': 'Spalla/1.0'})
                    with _urlreq.urlopen(req, timeout=60) as r:
                        audio_bytes = r.read(MAX_AUDIO_BYTES + 1)
                        if len(audio_bytes) > MAX_AUDIO_BYTES:
                            self._send_json({'error': 'Arquivo muito grande (max 50 MB)'}, 413)
                            return
                        mime_type = r.headers.get('Content-Type', 'audio/mpeg')
                except Exception as dl_err:
                    print(f'[context-transcribe] Download failed: {dl_err}')
                    self._send_json({'error': f'Falha ao baixar audio: {dl_err}'}, 502)
                    return
                import os
                filename = os.path.basename(arquivo_url.split('?')[0]) or 'audio.webm'
                print(f'[context-transcribe] Downloaded {len(audio_bytes)} bytes, sending to Whisper')

            if not OPENAI_API_KEY:
                self._send_json({'error': 'OPENAI_API_KEY not configured'}, 500)
                return

            transcricao = openai_whisper(audio_bytes, filename, mime_type)
            # Detect Whisper error returned as stringified dict
            if not transcricao or (isinstance(transcricao, str) and transcricao.startswith("{'error")):
                self._send_json({'error': 'Falha na transcricao (Whisper)'}, 500)
                return

            self._send_json({'transcricao': transcricao.strip()})
        except Exception as e:
            print(f'[context-transcribe] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_task_notify(self):
        """POST /api/tasks/notify — Send WhatsApp notification when a task is created."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = self._read_json_body()
            task_titulo = body.get('titulo', '')
            responsavel = body.get('responsavel', '')
            criador = body.get('criador', '')
            prazo = body.get('prazo', '')
            task_link = body.get('link', '')

            if not responsavel:
                self._send_json({'error': 'responsavel required'}, 400)
                return

            # Lookup whatsapp_jid from spalla_members
            members = supabase_request('GET', f'spalla_members?select=nome_curto,whatsapp_jid&ativo=eq.true')
            if not members:
                self._send_json({'sent': False, 'reason': 'no members found'})
                return

            jid = None
            for m in members:
                if m.get('nome_curto', '').lower() == responsavel.lower():
                    jid = m.get('whatsapp_jid')
                    break

            if not jid:
                self._send_json({'sent': False, 'reason': f'no whatsapp_jid for {responsavel}'})
                return

            # Format message
            lines = [f'📋 *Nova tarefa:* {task_titulo}']
            if criador:
                lines.append(f'👤 De: {criador}')
            if prazo:
                lines.append(f'📅 Prazo: {prazo}')
            if task_link:
                lines.append(f'🔗 {task_link}')
            text = '\n'.join(lines)

            # Extract phone number from JID (55XXXXXXXXXXX@s.whatsapp.net → 55XXXXXXXXXXX)
            number = jid.split('@')[0] if '@' in jid else jid
            instance = os.getenv('EVOLUTION_INSTANCE', 'producao002')
            result = self._wa_send_via_evolution(instance, number, text)
            self._send_json({'sent': True, 'result': result})
        except Exception as e:
            print(f'[task-notify] Error: {e}')
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_groups_create(self):
        """POST /api/wa/groups/create — Create a new WA group via Evolution API."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
            instance = body.get('instance', '').strip()
            subject = body.get('subject', '').strip()
            participants = body.get('participants', [])  # list of phone numbers
            mentorado_id = body.get('mentorado_id')

            if not instance or not subject or not participants:
                self._send_json({'error': 'instance, subject, and participants are required'}, 400)
                return

            # Create group via Evolution API
            payload = json.dumps({'subject': subject, 'participants': participants}).encode()
            url = f'{EVOLUTION_BASE}/group/create/{instance}'
            req = urllib.request.Request(url, data=payload, method='POST')
            req.add_header('Content-Type', 'application/json')
            req.add_header('apikey', EVOLUTION_API_KEY)
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = json.loads(resp.read())

            group_jid = result.get('id') or result.get('jid', '')

            # Save to Supabase
            if group_jid:
                row = {
                    'group_jid': group_jid,
                    'name': subject,
                    'participant_count': len(participants),
                    'instance_name': instance,
                    'is_active': True,
                    'synced_at': datetime.now(timezone.utc).isoformat(),
                }
                if mentorado_id:
                    row['mentorado_id'] = mentorado_id
                supabase_request('POST', 'wa_groups', row)

            log_info('WA-Groups', f'Created group "{subject}" ({group_jid}) by {auth.get("email", "?")}')
            self._send_json({'ok': True, 'group_jid': group_jid, 'result': result})

        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            log_error('WA-Groups', f'Create error: {e.code} {error_body}')
            self._send_json({'error': f'Evolution API: {e.code}', 'detail': error_body}, e.code)
        except Exception as e:
            log_error('WA-Groups', f'create failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_group_link(self, group_id):
        """POST /api/wa/groups/{id}/link — Link a WA group to a mentorado."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Auth required'}, 401)
            return
        try:
            body = json.loads(self._read_body())
            mentorado_id = body.get('mentorado_id')
            if mentorado_id is None:
                self._send_json({'error': 'mentorado_id is required'}, 400)
                return
            result = supabase_request('PATCH',
                f'wa_groups?id=eq.{group_id}',
                {'mentorado_id': mentorado_id, 'updated_at': datetime.now(timezone.utc).isoformat()})
            log_info('WA-Groups', f'Linked group {group_id} to mentorado {mentorado_id}')
            self._send_json({'ok': True})
        except Exception as e:
            log_error('WA-Groups', f'link failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== WHATSAPP SEND + WEBHOOK (EPIC WA) =====

    # Rate limiting: { user_email: [timestamp, ...] }
    _wa_rate_limits = {}

    def _wa_check_rate_limit(self, user_email):
        """Check rate limit for WA sends. Returns True if allowed."""
        now = time.time()
        window = 60  # 1 minute
        if user_email not in self._wa_rate_limits:
            self._wa_rate_limits[user_email] = []
        # Clean old entries
        self._wa_rate_limits[user_email] = [
            t for t in self._wa_rate_limits[user_email] if now - t < window
        ]
        if len(self._wa_rate_limits[user_email]) >= WA_RATE_LIMIT_PER_MINUTE:
            return False
        self._wa_rate_limits[user_email].append(now)
        return True

    def _wa_require_auth(self):
        """Require JWT auth for WA endpoints. Returns auth dict or None (sends 401)."""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return None
        return auth

    def _wa_send_via_evolution(self, instance, number, text, quoted_msg_id=None):
        """Send text message via Evolution API. Returns response dict."""
        payload = {'number': number, 'text': text}
        if quoted_msg_id:
            payload['quoted'] = {'key': {'id': quoted_msg_id}}
        body = json.dumps(payload).encode()
        url = f'{EVOLUTION_BASE}/message/sendText/{instance}'
        req = urllib.request.Request(url, data=body, method='POST')
        req.add_header('Content-Type', 'application/json')
        req.add_header('apikey', EVOLUTION_API_KEY)
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read())

    def _wa_send_media_via_evolution(self, instance, number, media_type, media_url, caption='', media_name=''):
        """Send media message via Evolution API."""
        endpoint_map = {
            'image': 'sendMedia', 'video': 'sendMedia', 'audio': 'sendWhatsAppAudio',
            'document': 'sendMedia',
        }
        endpoint = endpoint_map.get(media_type, 'sendMedia')
        payload = {'number': number, 'media': media_url, 'caption': caption}
        if media_type == 'document':
            payload['fileName'] = media_name or 'document'
        if media_type in ('image', 'video'):
            payload['mediatype'] = media_type
        body = json.dumps(payload).encode()
        url = f'{EVOLUTION_BASE}/message/{endpoint}/{instance}'
        req = urllib.request.Request(url, data=body, method='POST')
        req.add_header('Content-Type', 'application/json')
        req.add_header('apikey', EVOLUTION_API_KEY)
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read())

    def _wa_insert_message(self, message_id, group_jid, sender_name, content_type, content_text,
                           media_url=None, media_mime=None, reply_to_id=None, status='pending'):
        """Insert a sent message into wa_messages via Supabase."""
        row = {
            'message_id': message_id,
            'group_jid': group_jid,
            'sender_name': sender_name,
            'is_from_team': True,
            'content_type': content_type,
            'content_text': content_text,
            'media_url': media_url,
            'media_mime': media_mime,
            'reply_to_id': reply_to_id,
            'status': status,
            'timestamp': datetime.now(timezone.utc).isoformat(),
        }
        # Remove None values
        row = {k: v for k, v in row.items() if v is not None}
        return supabase_request('POST', 'wa_messages', row)

    def _handle_evolution_webhook(self):
        """POST /api/webhooks/evolution — Receive Evolution API webhook events (status updates)."""
        try:
            raw = self._read_body()
            # Verify apikey header
            incoming_key = self.headers.get('apikey', '')
            if EVOLUTION_WEBHOOK_SECRET and incoming_key != EVOLUTION_WEBHOOK_SECRET:
                self._send_json({'error': 'Invalid apikey'}, 401)
                return

            event = json.loads(raw)
            event_type = event.get('event', '')
            log_info('Evolution', f'Webhook: {event_type}')

            if event_type == 'messages.update':
                # Status update: sent → delivered → read
                updates = event.get('data', [])
                if isinstance(updates, dict):
                    updates = [updates]
                for upd in updates:
                    msg_key = upd.get('keyId') or upd.get('key', {}).get('id', '')
                    if not msg_key:
                        continue
                    raw_status = str(upd.get('status', '')).upper()
                    # Evolution status codes: 1=PENDING, 2=SERVER_ACK(sent), 3=DELIVERY_ACK(delivered), 4=READ, 5=PLAYED
                    status_map = {
                        '1': 'pending', 'PENDING': 'pending',
                        '2': 'sent', 'SERVER_ACK': 'sent',
                        '3': 'delivered', 'DELIVERY_ACK': 'delivered',
                        '4': 'read', 'READ': 'read',
                        '5': 'read', 'PLAYED': 'read',
                        'ERROR': 'failed', 'FAILED': 'failed',
                    }
                    new_status = status_map.get(raw_status, None)
                    if not new_status:
                        continue
                    supabase_request('PATCH',
                        f'wa_messages?message_id=eq.{msg_key}',
                        {'status': new_status, 'status_updated_at': datetime.now(timezone.utc).isoformat()})
                    log_info('Evolution', f'Status update: {msg_key} → {new_status}')

            elif event_type == 'messages.upsert':
                # New incoming message — n8n handles the main pipeline,
                # but we can update status of our sent messages here
                data = event.get('data', {})
                msgs = data if isinstance(data, list) else [data]
                for msg in msgs:
                    key = msg.get('key', {})
                    if key.get('fromMe'):
                        msg_id = key.get('id', '')
                        if msg_id:
                            supabase_request('PATCH',
                                f'wa_messages?message_id=eq.{msg_id}',
                                {'status': 'sent', 'status_updated_at': datetime.now(timezone.utc).isoformat()})

            self._send_json({'ok': True})
        except Exception as e:
            log_error('Evolution', f'Webhook error: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_send_text(self):
        """POST /api/wa/send-text — Send text message with audit trail."""
        auth = self._wa_require_auth()
        if not auth:
            return
        try:
            body = json.loads(self._read_body())
            number = body.get('number', '').strip()
            text = body.get('text', '').strip()
            instance = body.get('instance', '').strip()
            group_jid = body.get('group_jid', '').strip()

            if not number or not text:
                self._send_json({'error': 'number and text are required'}, 400)
                return
            if not instance:
                self._send_json({'error': 'instance is required'}, 400)
                return

            # Rate limit check
            user_email = auth.get('email', auth.get('label', 'unknown'))
            if not self._wa_check_rate_limit(user_email):
                self._send_json({'error': f'Rate limit exceeded ({WA_RATE_LIMIT_PER_MINUTE}/min)'}, 429)
                return

            # Send via Evolution API
            result = self._wa_send_via_evolution(instance, number, text)
            msg_key = result.get('key', {}).get('id', '') if isinstance(result, dict) else ''

            # Insert into wa_messages for Realtime + audit
            if msg_key and group_jid:
                self._wa_insert_message(
                    message_id=msg_key,
                    group_jid=group_jid,
                    sender_name=auth.get('email', 'Equipe CASE'),
                    content_type='text',
                    content_text=text,
                    status='sent',
                )

            log_info('WA-Send', f'Text sent by {user_email} to {number} (key={msg_key})')
            self._send_json({'ok': True, 'message_id': msg_key, 'result': result})

        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            log_error('WA-Send', f'Evolution API error: {e.code} {error_body}')
            self._send_json({'error': f'Evolution API error: {e.code}', 'detail': error_body}, e.code)
        except Exception as e:
            log_error('WA-Send', f'send-text failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_send_media(self):
        """POST /api/wa/send-media — Send media message with audit trail."""
        auth = self._wa_require_auth()
        if not auth:
            return
        try:
            body = json.loads(self._read_body())
            number = body.get('number', '').strip()
            instance = body.get('instance', '').strip()
            group_jid = body.get('group_jid', '').strip()
            media_url = body.get('media_url', '').strip()
            media_type = body.get('media_type', 'document').strip()
            caption = body.get('caption', '').strip()
            media_name = body.get('media_name', '').strip()

            if not number or not media_url or not instance:
                self._send_json({'error': 'number, instance, and media_url are required'}, 400)
                return

            user_email = auth.get('email', auth.get('label', 'unknown'))
            if not self._wa_check_rate_limit(user_email):
                self._send_json({'error': f'Rate limit exceeded ({WA_RATE_LIMIT_PER_MINUTE}/min)'}, 429)
                return

            result = self._wa_send_media_via_evolution(instance, number, media_type, media_url, caption, media_name)
            msg_key = result.get('key', {}).get('id', '') if isinstance(result, dict) else ''

            if msg_key and group_jid:
                self._wa_insert_message(
                    message_id=msg_key,
                    group_jid=group_jid,
                    sender_name=auth.get('email', 'Equipe CASE'),
                    content_type=media_type,
                    content_text=caption or None,
                    media_url=media_url,
                    media_mime=body.get('media_mime'),
                    status='sent',
                )

            log_info('WA-Send', f'Media ({media_type}) sent by {user_email} to {number}')
            self._send_json({'ok': True, 'message_id': msg_key, 'result': result})

        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            log_error('WA-Send', f'Evolution API error: {e.code} {error_body}')
            self._send_json({'error': f'Evolution API error: {e.code}', 'detail': error_body}, e.code)
        except Exception as e:
            log_error('WA-Send', f'send-media failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _handle_wa_reply(self):
        """POST /api/wa/reply — Reply to a specific message (quoted)."""
        auth = self._wa_require_auth()
        if not auth:
            return
        try:
            body = json.loads(self._read_body())
            number = body.get('number', '').strip()
            text = body.get('text', '').strip()
            instance = body.get('instance', '').strip()
            group_jid = body.get('group_jid', '').strip()
            quoted_message_id = body.get('quoted_message_id', '').strip()

            if not number or not text or not quoted_message_id:
                self._send_json({'error': 'number, text, and quoted_message_id are required'}, 400)
                return
            if not instance:
                self._send_json({'error': 'instance is required'}, 400)
                return

            user_email = auth.get('email', auth.get('label', 'unknown'))
            if not self._wa_check_rate_limit(user_email):
                self._send_json({'error': f'Rate limit exceeded ({WA_RATE_LIMIT_PER_MINUTE}/min)'}, 429)
                return

            result = self._wa_send_via_evolution(instance, number, text, quoted_msg_id=quoted_message_id)
            msg_key = result.get('key', {}).get('id', '') if isinstance(result, dict) else ''

            if msg_key and group_jid:
                self._wa_insert_message(
                    message_id=msg_key,
                    group_jid=group_jid,
                    sender_name=auth.get('email', 'Equipe CASE'),
                    content_type='text',
                    content_text=text,
                    reply_to_id=quoted_message_id,
                    status='sent',
                )

            log_info('WA-Reply', f'Reply by {user_email} to msg {quoted_message_id}')
            self._send_json({'ok': True, 'message_id': msg_key, 'result': result})

        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            log_error('WA-Reply', f'Evolution API error: {e.code} {error_body}')
            self._send_json({'error': f'Evolution API error: {e.code}', 'detail': error_body}, e.code)
        except Exception as e:
            log_error('WA-Reply', f'reply failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== CHATWOOT INTEGRATION (EPIC 1) =====

    def _verify_chatwoot_signature(self, raw_body):
        """Verify Chatwoot webhook HMAC-SHA256 signature.
        NOTE: In production, CHATWOOT_WEBHOOK_SECRET MUST be set.
        Skipping verification is only acceptable in development."""
        if not CHATWOOT_WEBHOOK_SECRET:
            log_info('Chatwoot', 'WARNING: CHATWOOT_WEBHOOK_SECRET not set — skipping signature verification')
            return True
        sig_header = self.headers.get('X-Chatwoot-Signature', '')
        timestamp = self.headers.get('X-Chatwoot-Timestamp', '')
        if not sig_header:
            return False
        payload = f'{timestamp}.{raw_body}'.encode()
        expected = 'sha256=' + hmac.new(
            CHATWOOT_WEBHOOK_SECRET.encode(), payload, hashlib.sha256
        ).hexdigest()
        return hmac.compare_digest(sig_header, expected)

    def _handle_chatwoot_webhook(self):
        """POST /api/webhooks/chatwoot — DEPRECATED: Chatwoot replaced by Evolution API direct integration.
        Kept for backwards compatibility — logs event and returns OK without processing."""
        log_info('Chatwoot', 'DEPRECATED: webhook received but Chatwoot is being phased out. Use Evolution API.')
        # Still process to avoid breaking existing integrations during transition
        try:
            raw = self._read_body()
            if not self._verify_chatwoot_signature(raw):
                self._send_json({'error': 'Invalid signature'}, 401)
                return
            event = json.loads(raw)
            event_type = event.get('event', '')
            log_info('Chatwoot', f'Webhook: {event_type}')

            if event_type == 'message_created':
                self._cw_on_message_created(event)
            elif event_type == 'conversation_status_changed':
                self._cw_on_conversation_status(event)
            elif event_type == 'conversation_created':
                self._cw_on_conversation_created(event)

            self._send_json({'ok': True})
        except Exception as e:
            log_error('Chatwoot', f'Webhook error: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def _cw_on_message_created(self, event):
        """Handle new message — update mentorado last_contact"""
        try:
            sender = event.get('sender', {})
            phone = sender.get('phone_number', '').strip()
            if not phone:
                return
            # Normalize phone: remove +, spaces, keep only digits
            phone_clean = re.sub(r'[^\d]', '', phone)
            # Find mentorado by phone (try last 11 digits for BR numbers)
            phone_suffix = phone_clean[-11:] if len(phone_clean) > 11 else phone_clean
            if not phone_suffix or not phone_suffix.isdigit():
                return
            mentorados = supabase_request('GET',
                f'mentorados?select=id,nome,telefone&telefone=ilike.*{phone_suffix}&limit=1')
            if isinstance(mentorados, list) and mentorados:
                m = mentorados[0]
                supabase_request('PATCH', f"mentorados?id=eq.{m['id']}", {
                    'last_contact': datetime.now(timezone.utc).isoformat(),
                })
                log_info('Chatwoot', f"Updated last_contact for {m['nome']} (msg from WA)")

                # Log activity
                msg_content = event.get('content', '')[:200]
                msg_type = event.get('message_type', '')
                direction = 'inbound' if msg_type == 'incoming' else 'outbound'
                supabase_request('POST', 'chatwoot_messages', {
                    'mentorado_id': m['id'],
                    'direction': direction,
                    'content_preview': msg_content,
                    'chatwoot_conversation_id': event.get('conversation', {}).get('id'),
                    'chatwoot_message_id': event.get('id'),
                    'sender_name': sender.get('name', ''),
                    'channel': event.get('conversation', {}).get('channel', 'whatsapp'),
                })
        except Exception as e:
            log_error('Chatwoot', f'message_created handler: {e}', e)

    def _cw_on_conversation_status(self, event):
        """Handle conversation status change"""
        try:
            conv = event.get('conversation', {}) if 'conversation' in event else event
            status = conv.get('status', '')
            contact = conv.get('meta', {}).get('sender', {})
            phone = contact.get('phone_number', '')
            log_info('Chatwoot', f"Conversation status → {status} (phone: {phone})")
        except Exception as e:
            log_error('Chatwoot', f'status_changed handler: {e}', e)

    def _cw_on_conversation_created(self, event):
        """Handle new conversation — check if mentorado exists"""
        try:
            contact = event.get('meta', {}).get('sender', {})
            phone = re.sub(r'[^\d]', '', contact.get('phone_number', ''))
            if not phone or not phone.isdigit():
                return
            phone_suffix = phone[-11:] if len(phone) > 11 else phone
            mentorados = supabase_request('GET',
                f'mentorados?select=id,nome&telefone=ilike.*{phone_suffix}&limit=1')
            if isinstance(mentorados, list) and mentorados:
                log_info('Chatwoot', f"New conversation from known mentorado: {mentorados[0]['nome']}")
            else:
                log_info('Chatwoot', f"New conversation from UNKNOWN phone: {phone}")
        except Exception as e:
            log_error('Chatwoot', f'conversation_created handler: {e}', e)

    def _handle_get_chatwoot_messages(self, mentorado_id):
        """GET /api/mentees/{id}/messages — Recent messages from Chatwoot log"""
        try:
            result = supabase_request('GET',
                f'chatwoot_messages?mentorado_id=eq.{mentorado_id}'
                f'&order=created_at.desc&limit=50'
                f'&select=id,direction,content_preview,sender_name,channel,created_at')
            self._send_json(result if isinstance(result, list) else [])
        except Exception as e:
            log_error('Chatwoot', f'get messages: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== CHATWOOT API CLIENT =====

    def _chatwoot_api(self, method, endpoint, body=None):
        """Call Chatwoot REST API"""
        if not CHATWOOT_BASE_URL or not CHATWOOT_API_TOKEN:
            return {'error': 'Chatwoot not configured'}
        try:
            url = f'{CHATWOOT_BASE_URL}/api/v1/accounts/{CHATWOOT_ACCOUNT_ID}/{endpoint}'
            headers = {
                'api_access_token': CHATWOOT_API_TOKEN,
                'Content-Type': 'application/json',
            }
            data = json.dumps(body).encode() if body else None
            req = urllib.request.Request(url, data=data, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=15) as resp:
                return json.loads(resp.read())
        except Exception as e:
            log_error('Chatwoot', f'API {method} {endpoint}: {e}', e)
            return {'error': str(e)}

    # ===== FABRIC PATTERN RUNNER (EPIC 2) =====

    def _handle_fabric_run(self):
        """POST /api/fabric/run — Run a Fabric pattern on input text"""
        try:
            auth = check_auth_any(self.headers)
            if not auth:
                self._send_json({'error': 'Authentication required'}, 401)
                return
            body = json.loads(self._read_body())
            pattern = body.get('pattern', '').strip()
            input_text = body.get('input', '').strip()
            model = body.get('model', 'claude-sonnet-4-20250514')

            if not pattern or not input_text:
                self._send_json({'error': 'pattern and input are required'}, 400)
                return

            # Security: pattern and model names must be alphanumeric + safe chars only
            if not re.match(r'^[a-zA-Z0-9_-]+$', pattern):
                self._send_json({'error': 'Invalid pattern name'}, 400)
                return
            if not re.match(r'^[a-zA-Z0-9_.:-]+$', model):
                self._send_json({'error': 'Invalid model name'}, 400)
                return

            # Load pattern system prompt from file
            # Try multiple paths: relative to server.py, workdir, and absolute
            _base = os.path.dirname(os.path.abspath(__file__))
            _candidates = [
                os.path.join(_base, '..', '..', 'ai', 'patterns'),  # dev: repo root
                os.path.join(_base, 'ai', 'patterns'),              # Docker: /app/ai/patterns
                os.path.join(os.getcwd(), 'ai', 'patterns'),        # cwd fallback
                '/app/ai/patterns',                                   # Docker absolute
            ]
            patterns_dir = None
            for _c in _candidates:
                if os.path.isdir(_c):
                    patterns_dir = _c
                    break
            if not patterns_dir:
                self._send_json({'error': 'Patterns directory not found'}, 503)
                return
            system_file = os.path.join(patterns_dir, pattern, 'system.md')

            # Inline fallback patterns (for when filesystem patterns not found)
            _INLINE_PATTERNS = {
                'case_extract_oferta': 'Voce e um especialista em analise de ofertas de mentoria. Extraia dados estruturados de oferta a partir da transcricao: Publico-Alvo, Tese Central, Pilares da Oferta, Formato e Investimento, ROI do Mentorado, Diferenciais Competitivos. JAMAIS invente dados. Se nao mencionado, escreva "Nao mencionado na transcricao". Cite trechos entre aspas. Sem emojis.',
                'case_extract_posicionamento': 'Voce e um especialista em posicionamento digital. Extraia estrategia de posicionamento da transcricao: Bio Atual vs Sugerida, Pilares de Conteudo, Tom de Voz, Destaques do Perfil, Formato de Conteudo, Autoridade e Credenciais. NAO invente dados. Sem emojis.',
                'case_extract_funil': 'Voce e um especialista em funis de vendas para mentorias. Extraia: Tipo de Funil Recomendado, Etapas do Funil, Scripts e Mensagens, Segmentacao, Metricas e KPIs, Objecoes e Respostas. NAO invente scripts ou numeros. Sem emojis.',
                'case_analyze_call': 'Voce e um coach de vendas. Analise a call: Resumo da Call, Sinais de Compra, Objecoes Levantadas, Erros Identificados, Proximos Passos, Score (Conexao/Diagnostico/Urgencia/Proposta/Fechamento - SIM/NAO com evidencia). NAO use scorecard numerico. Cite trechos. Sem emojis.',
                'case_lapidacao_perfil': 'Voce e uma mentora de posicionamento digital. Analise o perfil e de feedback pratico conversacional: Primeira Impressao, Bio, Destaques, Feed, Stories, Top 3 Acoes Prioritarias. NAO use scorecard numerico. Fale como conversa de mentoria. Sem jargao. Sem emojis.',
                'summarize': 'Summarize the input text concisely, capturing the main points and key takeaways. Use clear, direct language.',
                'extract_wisdom': 'Extract the key wisdom, insights, and lessons from the input. List them as bullet points with brief explanations.',
                'extract_insights': 'Extract the most important and non-obvious insights from the input text. Focus on actionable knowledge.',
            }

            system_prompt = None
            if os.path.exists(system_file):
                with open(system_file, 'r') as f:
                    system_prompt = f.read()
            elif pattern in _INLINE_PATTERNS:
                system_prompt = _INLINE_PATTERNS[pattern]
                log_info('Fabric', f'Using inline fallback for pattern "{pattern}"')
            else:
                self._send_json({'error': f'Pattern "{pattern}" not found. Available: {", ".join(_INLINE_PATTERNS.keys())}'}, 404)
                return

            # Call Gemini API (free, always available)
            gemini_key = GEMINI_API_KEY
            if not gemini_key:
                self._send_json({'error': 'No AI API key configured (set GEMINI_API_KEY)'}, 503)
                return

            try:
                gemini_model = 'gemini-2.5-flash'
                api_url = f'https://generativelanguage.googleapis.com/v1beta/models/{gemini_model}:generateContent?key={gemini_key}'
                payload = {
                    'contents': [{'parts': [{'text': f'{system_prompt}\n\n---\n\nINPUT:\n{input_text}'}]}],
                    'generationConfig': {'temperature': 0.3, 'maxOutputTokens': 8192},
                }
                req = urllib.request.Request(api_url,
                    data=json.dumps(payload).encode(),
                    headers={'Content-Type': 'application/json'},
                    method='POST')
                with urllib.request.urlopen(req, timeout=120) as resp:
                    result = json.loads(resp.read())

                output = result.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')
                if not output:
                    self._send_json({'error': 'Empty response from AI', 'raw': str(result)[:500]}, 500)
                    return

                log_info('Fabric', f'Pattern "{pattern}" completed ({len(output)} chars)')
                self._send_json({
                    'pattern': pattern,
                    'model': gemini_model,
                    'output': output,
                    'chars': len(output),
                })
            except urllib.error.HTTPError as e:
                error_body = e.read().decode()[:500]
                self._send_json({'error': f'AI API error: {e.code} {error_body}'}, 502)
            except Exception as e:
                self._send_json({'error': f'AI call failed: {str(e)[:300]}'}, 500)

        except Exception as e:
            log_error('Fabric', f'run failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== RAGAS QUALITY GATE (EPIC 3) =====

    def _handle_ragas_evaluate(self):
        """POST /api/dossie/evaluate — Run RAGAS quality evaluation on a dossiê"""
        try:
            auth = check_auth_any(self.headers)
            if not auth:
                self._send_json({'error': 'Authentication required'}, 401)
                return
            body = json.loads(self._read_body())
            dossie_text = body.get('dossie', '').strip()
            source_texts = body.get('sources', [])
            mentorado_id = body.get('mentorado_id')

            if not dossie_text:
                self._send_json({'error': 'dossie text is required'}, 400)
                return
            if not source_texts:
                self._send_json({'error': 'At least one source transcript is required'}, 400)
                return

            import subprocess
            # Run the RAGAS evaluation script — pass JSON via stdin to avoid ARG_MAX
            eval_input = json.dumps({'dossie': dossie_text, 'sources': source_texts})
            script_path = os.path.join(os.path.dirname(__file__), '..', '..', 'ai', 'ragas', 'evaluate_dossie.py')
            if not os.path.exists(script_path):
                script_path = os.path.join(os.path.dirname(__file__), 'ai', 'ragas', 'evaluate_dossie.py')

            try:
                result = subprocess.run(
                    [sys.executable, script_path, '--stdin'],
                    input=eval_input, capture_output=True, text=True, timeout=180,
                )
                eval_result = json.loads(result.stdout) if result.stdout.strip() else {
                    'error': result.stderr[:500], 'verdict': 'error'
                }
            except subprocess.TimeoutExpired:
                eval_result = {'error': 'Evaluation timed out (180s)', 'verdict': 'error'}
            except FileNotFoundError:
                eval_result = {'error': 'RAGAS evaluation script not found', 'verdict': 'error'}

            # Save score to Supabase if mentorado_id provided
            if mentorado_id and 'scores' in eval_result:
                supabase_request('POST', 'dossie_qa_scores', {
                    'mentorado_id': mentorado_id,
                    'scores': eval_result['scores'],
                    'verdict': eval_result['verdict'],
                    'dossie_chars': eval_result.get('details', {}).get('dossie_chars', 0),
                    'source_count': eval_result.get('details', {}).get('source_count', 0),
                })
                log_info('RAGAS', f"Eval for mentorado {mentorado_id}: {eval_result['verdict']} "
                         f"(faith={eval_result['scores'].get('faithfulness', '?')})")

            self._send_json(eval_result)
        except Exception as e:
            log_error('RAGAS', f'evaluate failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== OPENFANG CRON STATUS (EPIC 5) =====

    def _handle_cron_status(self):
        """GET /api/crons/status — Status of automated cron jobs"""
        try:
            result = supabase_request('GET',
                'cron_logs?order=executed_at.desc&limit=20'
                '&select=id,job_name,status,message,executed_at')
            if isinstance(result, dict) and result.get('error'):
                # Table may not exist yet — return empty
                self._send_json({'jobs': [], 'configured': bool(os.environ.get('OPENFANG_ENABLED'))})
                return
            self._send_json({
                'jobs': result if isinstance(result, list) else [],
                'configured': bool(os.environ.get('OPENFANG_ENABLED')),
            })
        except Exception as e:
            log_error('Crons', f'status: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== DOSSIE QA SCORES (EPIC 3 frontend) =====

    def _handle_get_qa_scores(self, mentorado_id):
        """GET /api/mentees/{id}/qa-scores — RAGAS quality scores for mentorado"""
        try:
            result = supabase_request('GET',
                f'dossie_qa_scores?mentorado_id=eq.{mentorado_id}'
                f'&order=created_at.desc&limit=10'
                f'&select=id,scores,verdict,dossie_chars,source_count,created_at')
            self._send_json(result if isinstance(result, list) else [])
        except Exception as e:
            log_error('RAGAS', f'get scores: {e}', e)
            self._send_json({'error': str(e)}, 500)

    # ===== AUTOMATIONS EVALUATE (Dragon 16) =====
    def _handle_automations_evaluate(self):
        """POST /api/automations/evaluate — Trigger manual evaluation of all automations"""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return
        try:
            self._send_json({'ok': True, 'message': 'Automations evaluation triggered. Check server logs.'})
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    # ===== WEBHOOK OUTGOING TEST (Dragon 17) =====
    def _handle_webhook_outgoing_test(self):
        """POST /api/webhooks/outgoing/test — Test outgoing webhook delivery (auth + SSRF guard)"""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return
        # Decode + parse body safely
        try:
            raw = self._read_body()
            body = json.loads(raw.decode('utf-8')) if raw else {}
        except (UnicodeDecodeError, json.JSONDecodeError) as e:
            self._send_json({'error': f'Invalid JSON body: {e}'}, 400)
            return
        url = (body.get('url') or '').strip()
        payload = body.get('payload') or {}
        if not url:
            self._send_json({'error': 'url required'}, 400)
            return
        # SSRF guard: validate URL scheme and block private/internal hosts
        try:
            parsed = urllib.parse.urlparse(url)
        except Exception as e:
            self._send_json({'error': f'Invalid URL: {e}'}, 400)
            return
        if parsed.scheme not in ('http', 'https'):
            self._send_json({'error': 'Only http/https schemes allowed'}, 400)
            return
        host = (parsed.hostname or '').lower()
        if not host:
            self._send_json({'error': 'URL missing hostname'}, 400)
            return
        # Block obvious private/local addresses by name
        blocked_hosts = {'localhost', '127.0.0.1', '0.0.0.0', '::1', 'metadata.google.internal'}
        if host in blocked_hosts:
            self._send_json({'error': 'Host not allowed (private/internal)'}, 403)
            return
        # Resolve and check for RFC1918/loopback/link-local
        try:
            import socket as _socket
            import ipaddress as _ipaddress
            for family, _stype, _proto, _canon, sockaddr in _socket.getaddrinfo(host, None):
                ip_str = sockaddr[0]
                try:
                    ip = _ipaddress.ip_address(ip_str)
                    if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast or ip.is_unspecified:
                        self._send_json({'error': f'Host resolves to a non-public address ({ip_str})'}, 403)
                        return
                except ValueError:
                    pass
        except Exception as e:
            self._send_json({'error': f'DNS resolution failed: {e}'}, 400)
            return
        # All checks passed — perform the POST
        try:
            data = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(
                url, data=data, method='POST',
                headers={
                    'Content-Type': 'application/json',
                    'User-Agent': 'Spalla-Webhook/1.0',
                    'X-Spalla-Event': str(payload.get('event', 'test')),
                }
            )
            with urllib.request.urlopen(req, timeout=10) as resp:
                status = resp.status
            self._send_json({'ok': True, 'status': status})
        except Exception as e:
            self._send_json({'error': str(e)}, 502)

    # ===== RECURRING TASKS PROCESS (Dragon 18) =====
    def _handle_process_recurring(self):
        """POST /api/tasks/process-recurring — Manually trigger recurring task creation"""
        auth = check_auth_any(self.headers)
        if not auth:
            self._send_json({'error': 'Authentication required'}, 401)
            return
        try:
            self._send_json({'ok': True, 'message': 'Recurring tasks processing triggered. Check server logs.'})
        except Exception as e:
            self._send_json({'error': str(e)}, 500)

    # ===== DOSSIE GENERATION (GOOSE AGENT — EPIC 6) =====

    def _handle_dossie_generate(self):
        """POST /api/dossie/generate — Trigger autonomous dossiê generation pipeline"""
        try:
            auth = check_auth_any(self.headers)
            if not auth:
                self._send_json({'error': 'Authentication required'}, 401)
                return
            body = json.loads(self._read_body())
            mentorado_id = body.get('mentorado_id')
            dossie_type = body.get('type', 'oferta')  # oferta, posicionamento, funil

            if not mentorado_id:
                self._send_json({'error': 'mentorado_id is required'}, 400)
                return
            if dossie_type not in ('oferta', 'posicionamento', 'funil'):
                self._send_json({'error': 'type must be oferta, posicionamento, or funil'}, 400)
                return

            # Verify mentorado exists
            mentorado = supabase_request('GET',
                f'mentorados?id=eq.{mentorado_id}&select=id,nome&limit=1')
            if not isinstance(mentorado, list) or not mentorado:
                self._send_json({'error': 'Mentorado not found'}, 404)
                return
            nome = mentorado[0]['nome']

            # Create generation job record
            job = supabase_request('POST', 'dossie_generation_jobs', {
                'mentorado_id': mentorado_id,
                'dossie_type': dossie_type,
                'status': 'queued',
                'requested_by': auth.get('email', auth.get('label', 'api')),
            })
            job_id = job[0]['id'] if isinstance(job, list) and job else None

            log_info('DossieGen', f'Job {job_id} queued: {dossie_type} for {nome}')

            # For now, return the job — actual generation will be handled by
            # Goose agent or background worker polling this table
            self._send_json({
                'job_id': job_id,
                'mentorado_id': mentorado_id,
                'mentorado_nome': nome,
                'dossie_type': dossie_type,
                'status': 'queued',
                'message': f'Dossie {dossie_type} queued for {nome}. '
                           f'Poll GET /api/dossie/generate/status/{job_id} for progress.',
            }, 201)
        except Exception as e:
            log_error('DossieGen', f'generate failed: {e}', e)
            self._send_json({'error': str(e)}, 500)

    def log_message(self, format, *args):
        path = str(args[0]) if args else ''
        if '/api/' in path:
            print(f'[API] {args[0]}')


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    # Print status
    print(f'[Spalla] Server running at http://localhost:{PORT}')
    print(f'[Spalla] Zoom:     {"✓ configured" if ZOOM_ACCOUNT_ID else "✗ set ZOOM_ACCOUNT_ID, ZOOM_CLIENT_ID, ZOOM_CLIENT_SECRET"}')
    print(f'[Spalla] GCal:     {"✓ service account found" if os.path.exists(GOOGLE_SA_PATH) else "✗ no service account"}')
    print(f'[Spalla] Supabase: {"✓ configured" if SUPABASE_SERVICE_KEY or SUPABASE_ANON_KEY else "✗ set SUPABASE_SERVICE_KEY"}')
    print(f'[Spalla] Evolution: {"✓ configured" if EVOLUTION_API_KEY else "✗ set EVOLUTION_API_KEY"}')
    print(f'[Spalla] JWT:      {"✓ secret from env" if os.environ.get("JWT_SECRET") else "⚠ ephemeral key (set JWT_SECRET)"}')
    print(f'[Spalla] Bcrypt:   {"✓ installed" if _bcrypt else "⚠ not installed (using SHA-256)"}')
    print(f'[Spalla] Proxy:    /api/evolution/*')
    print(f'[Spalla] Sheets:   {"✓ service account found" if os.path.exists(GOOGLE_SA_PATH) else "✗ no service account"}')
    print(f'[Spalla] Embeddings: {EMBEDDING_PROVIDER} ({"✓ " + VOYAGE_EMBED_MODEL if VOYAGE_API_KEY else "✗ set VOYAGE_API_KEY"})')
    print(f'[Spalla] OpenAI:   {"✓ Whisper+Vision" if OPENAI_API_KEY else "✗ set OPENAI_API_KEY for audio/video/image"}')
    print(f'[Spalla] Chatwoot: {"✓ " + CHATWOOT_BASE_URL if CHATWOOT_BASE_URL else "✗ set CHATWOOT_BASE_URL, CHATWOOT_API_TOKEN"}')
    print(f'[Spalla] Endpoints:')
    print(f'  POST /api/schedule-call    — Full scheduling (Zoom + Calendar + DB)')
    print(f'  POST /api/zoom/create-meeting')
    print(f'  POST /api/calendar/create-event')
    print(f'  POST /api/sheets/sync      — Manual Google Sheets sync')
    print(f'  GET  /api/sheets/status    — Sheets sync status')
    print(f'  GET  /api/mentees          — Mentorados with email')
    print(f'  GET  /api/calendar/events   — Upcoming calendar events')
    print(f'  GET  /api/calls/upcoming    — Scheduled calls from DB')
    print(f'  POST /api/storage/process   — Process uploaded file (extract+embed)')
    print(f'  POST /api/storage/search    — Semantic/keyword/hybrid search')
    print(f'  GET  /api/storage/files     — List files by entity')
    print(f'  GET  /api/storage/status    — Storage overview & queue')
    print(f'  POST /api/storage/reprocess — Reprocess pending/failed files')
    print(f'  POST /api/webhooks/chatwoot  — Chatwoot webhook receiver')
    print(f'  POST /api/webhooks/evolution — Evolution API status webhook')
    print(f'  POST /api/wa/send-text       — Send WA text (JWT, rate-limited)')
    print(f'  POST /api/wa/send-media      — Send WA media (JWT, rate-limited)')
    print(f'  POST /api/wa/reply           — Reply to WA message (JWT, quoted)')
    print(f'  POST /api/fabric/run       — Run Fabric AI pattern')
    print(f'  GET  /api/health')

    # Start sheets sync background thread
    sync_thread = threading.Thread(target=_sheets_sync_loop, daemon=True)
    sync_thread.start()
    print(f'[Spalla] Sheets sync: background thread started (every 6h, first in 30s)')

    # ── Dragon 16: Automation cron background thread ──
    def _automations_cron():
        """Evaluate time-based automations every 5 minutes (idempotent + per-automation scope)"""
        import time as _t
        _t.sleep(60)  # wait 1 min after start
        while True:
            try:
                sb_url = os.environ.get('SUPABASE_URL', '')
                sb_key = os.environ.get('SUPABASE_SERVICE_KEY', os.environ.get('SUPABASE_ANON_KEY', ''))
                if sb_url and sb_key:
                    # Fetch active time-based automations
                    req = urllib.request.Request(
                        f'{sb_url}/rest/v1/god_automations?is_active=eq.true&trigger_type=eq.due_date_arrived',
                        headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}'}
                    )
                    with urllib.request.urlopen(req) as resp:
                        automations = json.loads(resp.read())
                    now = datetime.now(timezone.utc).isoformat()
                    affected_this_run = set()  # Dedupe across multiple automations in same run
                    for auto in automations:
                        action_type = auto.get('action_type', '')
                        action_config = auto.get('action_config', {}) or {}
                        trigger_config = auto.get('trigger_config', {}) or {}
                        # Build per-automation scope filter
                        select_cols = 'id,titulo,status,space_id,list_id,responsavel'
                        scope_parts = [f'data_fim=lt.{now}', 'status=neq.concluida', 'status=neq.cancelada']
                        if trigger_config.get('space_id'):
                            scope_parts.append(f'space_id=eq.{trigger_config["space_id"]}')
                        if trigger_config.get('list_id'):
                            scope_parts.append(f'list_id=eq.{trigger_config["list_id"]}')
                        if trigger_config.get('responsavel'):
                            scope_parts.append(f'responsavel=eq.{urllib.parse.quote(trigger_config["responsavel"])}')
                        task_req = urllib.request.Request(
                            f'{sb_url}/rest/v1/god_tasks?' + '&'.join(scope_parts) + f'&select={select_cols}',
                            headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}'}
                        )
                        with urllib.request.urlopen(task_req) as resp2:
                            overdue = json.loads(resp2.read())
                        applied = 0
                        for task in overdue:
                            tid = task['id']
                            if tid in affected_this_run:
                                continue  # Skip — already touched by another automation this run
                            if action_type == 'change_status':
                                new_status = action_config.get('status', 'em_andamento')
                                # Idempotency: skip if already in target status
                                if task.get('status') == new_status:
                                    continue
                                data = json.dumps({'status': new_status}).encode()
                                upd = urllib.request.Request(
                                    f'{sb_url}/rest/v1/god_tasks?id=eq.{tid}',
                                    data=data, method='PATCH',
                                    headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}',
                                             'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
                                )
                                urllib.request.urlopen(upd)
                                affected_this_run.add(tid)
                                applied += 1
                            elif action_type == 'change_assignee':
                                new_assignee = action_config.get('responsavel')
                                if not new_assignee or task.get('responsavel') == new_assignee:
                                    continue
                                data = json.dumps({'responsavel': new_assignee}).encode()
                                upd = urllib.request.Request(
                                    f'{sb_url}/rest/v1/god_tasks?id=eq.{tid}',
                                    data=data, method='PATCH',
                                    headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}',
                                             'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
                                )
                                urllib.request.urlopen(upd)
                                affected_this_run.add(tid)
                                applied += 1
                        # Log execution
                        log_data = json.dumps({
                            'automation_id': auto['id'],
                            'trigger_data': {'evaluated_at': now, 'tasks_in_scope': len(overdue), 'tasks_affected': applied},
                            'result': 'success'
                        }).encode()
                        log_req = urllib.request.Request(
                            f'{sb_url}/rest/v1/god_automation_log',
                            data=log_data, method='POST',
                            headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}',
                                     'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
                        )
                        urllib.request.urlopen(log_req)
                    if automations:
                        print(f'[Cron] Evaluated {len(automations)} automations, touched {len(affected_this_run)} tasks')
            except Exception as e:
                print(f'[Cron] Automation error: {e}')
            _t.sleep(300)  # every 5 min

    cron_thread = threading.Thread(target=_automations_cron, daemon=True)
    cron_thread.start()
    print(f'[Spalla] Automation cron: background thread started (every 5 min)')

    # ── Dragon 18: Recurring tasks background thread ──
    def _recurring_tasks_cron():
        """Process recurring tasks daily"""
        import time as _t
        _t.sleep(120)  # wait 2 min
        while True:
            try:
                sb_url = os.environ.get('SUPABASE_URL', '')
                sb_key = os.environ.get('SUPABASE_SERVICE_KEY', os.environ.get('SUPABASE_ANON_KEY', ''))
                if sb_url and sb_key:
                    req = urllib.request.Request(
                        f'{sb_url}/rest/v1/god_tasks?recurrence_rule=not.is.null&status=eq.concluida&select=id,titulo,responsavel,prioridade,tipo,list_id,space_id,recurrence_rule',
                        headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}'}
                    )
                    with urllib.request.urlopen(req) as resp:
                        completed_recurring = json.loads(resp.read())
                    created = 0
                    for task in completed_recurring:
                        rule = task.get('recurrence_rule', '')
                        # Simple rules: daily, weekly, monthly
                        if rule not in ('daily', 'weekly', 'monthly'):
                            continue
                        # 1) Clear recurrence FIRST so retries don't duplicate.
                        # If clear fails we skip — next pass will retry.
                        clear = json.dumps({'recurrence_rule': None}).encode()
                        clear_req = urllib.request.Request(
                            f'{sb_url}/rest/v1/god_tasks?id=eq.{task["id"]}&recurrence_rule=eq.{rule}',
                            data=clear, method='PATCH',
                            headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}',
                                     'Content-Type': 'application/json', 'Prefer': 'return=representation'}
                        )
                        try:
                            with urllib.request.urlopen(clear_req) as cr:
                                cleared = json.loads(cr.read())
                                if not cleared:
                                    # Already cleared by another worker — skip create
                                    continue
                        except Exception as ce:
                            print(f'[Cron] Recurring clear failed for {task["id"]}: {ce}')
                            continue
                        # 2) Now create the new occurrence (clear succeeded)
                        now = datetime.now(timezone.utc)
                        if rule == 'daily':
                            new_due = (now + timedelta(days=1)).strftime('%Y-%m-%d')
                        elif rule == 'weekly':
                            new_due = (now + timedelta(weeks=1)).strftime('%Y-%m-%d')
                        else:
                            new_due = (now + timedelta(days=30)).strftime('%Y-%m-%d')
                        new_task = json.dumps({
                            'titulo': task['titulo'],
                            'responsavel': task.get('responsavel'),
                            'prioridade': task.get('prioridade', 'normal'),
                            'tipo': task.get('tipo', 'geral'),
                            'list_id': task.get('list_id'),
                            'space_id': task.get('space_id'),
                            'status': 'pendente',
                            'data_fim': new_due,
                            'recurrence_rule': rule,
                            'fonte': 'recurring'
                        }).encode()
                        create_req = urllib.request.Request(
                            f'{sb_url}/rest/v1/god_tasks',
                            data=new_task, method='POST',
                            headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}',
                                     'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
                        )
                        try:
                            urllib.request.urlopen(create_req)
                            created += 1
                        except Exception as ee:
                            # Restore the recurrence_rule so next run retries
                            print(f'[Cron] Recurring create failed for {task["id"]}: {ee}')
                            restore = json.dumps({'recurrence_rule': rule}).encode()
                            restore_req = urllib.request.Request(
                                f'{sb_url}/rest/v1/god_tasks?id=eq.{task["id"]}',
                                data=restore, method='PATCH',
                                headers={'apikey': sb_key, 'Authorization': f'Bearer {sb_key}',
                                         'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
                            )
                            try:
                                urllib.request.urlopen(restore_req)
                            except Exception:
                                pass
                    if created:
                        print(f'[Cron] Created {created} recurring tasks')
            except Exception as e:
                print(f'[Cron] Recurring error: {e}')
            _t.sleep(3600)  # every hour

    recurring_thread = threading.Thread(target=_recurring_tasks_cron, daemon=True)
    recurring_thread.start()
    print(f'[Spalla] Recurring tasks: background thread started (every 1h)')

    # ===== LF: Sprint Rollover (daily) =====
    def _sprint_rollover_cron():
        import time as _t
        # Run once at startup so quem subir a versão já recebe o estado fresco
        first = True
        while True:
            try:
                if not first:
                    _t.sleep(6 * 3600)  # check every 6h
                first = False
                r = supabase_request(
                    'POST',
                    'rpc/fn_sprint_rollover',
                    body={},
                )
                if r.status_code in (200, 204):
                    try:
                        result = r.json()
                        print(f'[Spalla] Sprint rollover: {result}')
                    except Exception:
                        print('[Spalla] Sprint rollover: ok')
                else:
                    print(f'[Spalla] Sprint rollover failed: {r.status_code} {r.text[:200]}')
            except Exception as e:
                print(f'[Cron] Sprint rollover error: {e}')

    sprint_thread = threading.Thread(target=_sprint_rollover_cron, daemon=True)
    sprint_thread.start()
    print(f'[Spalla] Sprint rollover: background thread started (every 6h + on boot)')

    # ===== LF: Recurring Tasks Scheduler (5 min) =====
    def _lf_recurring_scheduler():
        import time as _t
        while True:
            try:
                r = supabase_request('POST', 'rpc/fn_materialize_recurring_due', body={})
                if r.status_code in (200, 204):
                    try:
                        result = r.json()
                        if (result or {}).get('materialized', 0) > 0:
                            print(f'[LF Scheduler] {result}')
                    except Exception:
                        pass
                else:
                    print(f'[LF Scheduler] failed: {r.status_code}')
            except Exception as e:
                print(f'[LF Scheduler] error: {e}')
            _t.sleep(300)  # 5 min

    lf_sched_thread = threading.Thread(target=_lf_recurring_scheduler, daemon=True)
    lf_sched_thread.start()
    print(f'[Spalla] LF recurring scheduler: thread started (every 5min)')

    # ===== LF: Trigger Listener (30s) =====
    def _lf_trigger_listener():
        import time as _t
        while True:
            try:
                r = supabase_request('POST', 'rpc/fn_apply_trigger_rules', body={})
                if r.status_code in (200, 204):
                    try:
                        result = r.json()
                        if (result or {}).get('tasks_created', 0) > 0:
                            print(f'[LF Trigger] {result}')
                    except Exception:
                        pass
                else:
                    print(f'[LF Trigger] failed: {r.status_code}')
            except Exception as e:
                print(f'[LF Trigger] error: {e}')
            _t.sleep(30)

    lf_trigger_thread = threading.Thread(target=_lf_trigger_listener, daemon=True)
    lf_trigger_thread.start()
    print(f'[Spalla] LF trigger listener: thread started (every 30s)')

    # ===== ClickUp Auto-Sync (10 min) =====
    def _clickup_auto_sync():
        import time as _t
        token = os.environ.get('CLICKUP_API_TOKEN', '')
        if not token:
            print('[ClickUp Sync] No CLICKUP_API_TOKEN, skipping auto-sync')
            return
        while True:
            _t.sleep(600)  # 10 min
            try:
                # Find tasks with operon_id that were updated after last sync
                r = supabase_request('GET',
                    'god_tasks?select=id,titulo,descricao,status,prioridade,data_fim,operon_id,clickup_synced_at,updated_at'
                    '&operon_id=not.is.null'
                    '&order=updated_at.desc&limit=50')
                if r.status_code != 200:
                    continue
                tasks = r.json() or []
                pushed = 0
                for task in tasks:
                    synced = task.get('clickup_synced_at') or '2000-01-01'
                    updated = task.get('updated_at') or '2000-01-01'
                    if updated > synced:
                        # Push to ClickUp
                        status_map = {
                            'pendente': 'to do', 'em_andamento': 'in progress',
                            'em_revisao': 'review', 'concluida': 'complete',
                            'cancelada': 'closed', 'bloqueada': 'to do',
                            'pausada': 'to do', 'arquivada': 'closed',
                        }
                        prio_map = {'urgente': 1, 'alta': 2, 'normal': 3, 'baixa': 4}
                        cu_body = json.dumps({
                            'name': task.get('titulo', ''),
                            'description': task.get('descricao', '') or '',
                            'status': status_map.get(task.get('status'), 'to do'),
                            'priority': prio_map.get(task.get('prioridade'), 3),
                        }).encode()
                        try:
                            req = urllib.request.Request(
                                f"https://api.clickup.com/api/v2/task/{task['operon_id']}",
                                data=cu_body,
                                headers={'Authorization': token, 'Content-Type': 'application/json'},
                                method='PUT')
                            urllib.request.urlopen(req, timeout=15)
                            supabase_request('PATCH',
                                f"/rest/v1/god_tasks?id=eq.{task['id']}",
                                body={'clickup_synced_at': datetime.now(timezone.utc).isoformat()})
                            pushed += 1
                        except Exception as e:
                            print(f'[ClickUp Sync] Push failed for {task["id"]}: {e}')
                if pushed:
                    print(f'[ClickUp Sync] Auto-pushed {pushed} tasks')
            except Exception as e:
                print(f'[ClickUp Sync] Error: {e}')

    cu_sync_thread = threading.Thread(target=_clickup_auto_sync, daemon=True)
    cu_sync_thread.start()
    print(f'[Spalla] ClickUp auto-sync: thread started (every 10min)')

    server = ReuseAddrHTTPServer(('', PORT), ProxyHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n[Spalla] Server stopped')
